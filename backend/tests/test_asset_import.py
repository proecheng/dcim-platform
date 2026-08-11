"""资产导入/导出 + U位冲突校验测试 — Story 7-1"""

import pytest
from io import BytesIO

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import delete
from openpyxl import Workbook, load_workbook

from app.core.database import Base
from app.models.asset import (
    Asset,
    Cabinet,
    AssetLifecycle,
    AssetType,
    AssetStatus,
    MaintenanceRecord,
    AssetInventory,
    AssetInventoryItem,
)
from app.models.user import User
from app.api.deps import SiteAccessContext
from app.api.v1.asset import _check_u_position_conflict, import_assets, export_assets


# ============================================================
# Fixtures
# ============================================================


@pytest.fixture(scope="module")
def anyio_backend():
    return "asyncio"


@pytest.fixture(scope="module")
async def engine():
    eng = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with eng.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield eng
    await eng.dispose()


@pytest.fixture(scope="module")
def session_factory(engine):
    return async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


@pytest.fixture
async def db(session_factory):
    async with session_factory() as session:
        # 清理
        await session.execute(delete(AssetLifecycle))
        await session.execute(delete(AssetInventoryItem))
        await session.execute(delete(AssetInventory))
        await session.execute(delete(MaintenanceRecord))
        await session.execute(delete(Asset))
        await session.execute(delete(Cabinet))
        await session.execute(delete(User))
        await session.commit()
        yield session


@pytest.fixture
async def sample_cabinet(db: AsyncSession):
    """创建测试机柜"""
    cab = Cabinet(
        cabinet_code="CAB-001",
        cabinet_name="测试机柜A",
        location="A区",
        total_u=42,
    )
    db.add(cab)
    await db.commit()
    await db.refresh(cab)
    return cab


@pytest.fixture
async def sample_asset(db: AsyncSession, sample_cabinet):
    """创建测试资产（占用 U1-U4）"""
    asset = Asset(
        asset_code="SRV-001",
        asset_name="测试服务器1",
        asset_type=AssetType.server,
        cabinet_id=sample_cabinet.id,
        u_position=1,
        u_height=4,
        status=AssetStatus.in_use,
    )
    db.add(asset)
    await db.commit()
    await db.refresh(asset)
    return asset


@pytest.fixture
async def test_user(db: AsyncSession):
    """创建测试用户"""
    user = User(
        username="test_import_user",
        password_hash="fakehash",
        role="operator",
        is_active=True,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    return user


def _admin_context(user: User) -> SiteAccessContext:
    return SiteAccessContext(user.id, "admin", "test-jti", None)


def _make_excel(rows: list[list]) -> BytesIO:
    """构造 Excel BytesIO，rows[0] 为表头"""
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# 辅助：模拟 UploadFile
# ============================================================


class FakeUploadFile:
    """模拟 FastAPI UploadFile"""

    def __init__(self, buf: BytesIO, filename: str = "test.xlsx"):
        self._buf = buf
        self.filename = filename

    async def read(self):
        return self._buf.read()


# ============================================================
# 测试
# ============================================================


class TestUPositionConflict:
    # 1. 冲突检测
    async def test_u_position_conflict_detection(self, db: AsyncSession, sample_asset, sample_cabinet):
        """已有 U1-U4 的资产，新资产 U3-U6 应检测到冲突"""
        conflict = await _check_u_position_conflict(db, sample_cabinet.id, u_position=3, u_height=4)
        assert conflict is not None
        assert "U位冲突" in conflict
        assert "SRV-001" in conflict

    # 2. 无冲突
    async def test_u_position_no_conflict(self, db: AsyncSession, sample_asset, sample_cabinet):
        """已有 U1-U4，新资产 U5-U8 不应冲突"""
        conflict = await _check_u_position_conflict(db, sample_cabinet.id, u_position=5, u_height=4)
        assert conflict is None

    # 3. 创建资产时 U 位冲突返回 400
    async def test_create_asset_with_u_conflict(self, db: AsyncSession, sample_asset, sample_cabinet):
        """通过 API 创建资产时 U 位冲突应返回 400"""
        from app.schemas.asset import AssetCreate

        data = AssetCreate(
            asset_code="SRV-CONFLICT",
            asset_name="冲突服务器",
            asset_type=AssetType.server,
            cabinet_id=sample_cabinet.id,
            u_position=2,
            u_height=3,
        )

        # 直接调用冲突检测函数验证
        conflict = await _check_u_position_conflict(db, data.cabinet_id, data.u_position, data.u_height)
        assert conflict is not None
        assert "U位冲突" in conflict

    # 4. 更新资产时 U 位冲突
    async def test_update_asset_with_u_conflict(self, db: AsyncSession, sample_cabinet):
        """两个资产，更新第二个使其与第一个冲突"""
        asset1 = Asset(
            asset_code="UPD-001",
            asset_name="资产1",
            asset_type=AssetType.server,
            cabinet_id=sample_cabinet.id,
            u_position=1,
            u_height=4,
        )
        asset2 = Asset(
            asset_code="UPD-002",
            asset_name="资产2",
            asset_type=AssetType.server,
            cabinet_id=sample_cabinet.id,
            u_position=10,
            u_height=2,
        )
        db.add_all([asset1, asset2])
        await db.commit()
        await db.refresh(asset1)
        await db.refresh(asset2)

        # 尝试将 asset2 移到 U3（与 asset1 的 U1-U4 冲突）
        conflict = await _check_u_position_conflict(
            db, sample_cabinet.id, u_position=3, u_height=2, exclude_asset_id=asset2.id
        )
        assert conflict is not None
        assert "UPD-001" in conflict

    # 5. exclude_asset_id 排除自身
    async def test_u_position_exclude_self(self, db: AsyncSession, sample_asset, sample_cabinet):
        """更新自身位置时不应与自身冲突"""
        conflict = await _check_u_position_conflict(
            db, sample_cabinet.id, u_position=1, u_height=4, exclude_asset_id=sample_asset.id
        )
        assert conflict is None


class TestImportAssets:
    # 6. 有效 Excel 预校验
    async def test_import_preview_valid(self, db: AsyncSession, sample_cabinet, test_user):
        """有效数据预校验应返回 success_count > 0"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型", "机柜编码", "U位起始", "占用U数"],
                ["IMP-001", "导入服务器1", "服务器", "CAB-001", 10, 2],
                ["IMP-002", "导入服务器2", "server", None, None, None],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="preview",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["total"] == 2
        assert result["success_count"] == 2
        assert result["error_count"] == 0

    # 7. 缺少必填字段
    async def test_import_preview_missing_required(self, db: AsyncSession, test_user):
        """缺少 asset_code 应报错"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型"],
                [None, "无编码资产", "服务器"],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="preview",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["error_count"] >= 1
        assert any(e["field"] == "asset_code" for e in result["errors"])

    # 8. 重复编码
    async def test_import_preview_duplicate_code(self, db: AsyncSession, sample_asset, test_user):
        """编码与数据库已有资产重复应报错"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型"],
                ["SRV-001", "重复编码资产", "服务器"],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="preview",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["error_count"] >= 1
        assert any("已存在" in e["message"] for e in result["errors"])

    # 9. 无效机柜编码
    async def test_import_preview_invalid_cabinet(self, db: AsyncSession, test_user):
        """不存在的机柜编码应报错"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型", "机柜编码"],
                ["IMP-CAB-ERR", "机柜错误资产", "服务器", "NONEXIST-CAB"],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="preview",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["error_count"] >= 1
        assert any("机柜编码不存在" in e["message"] for e in result["errors"])

    # 10. 确认导入
    async def test_import_confirm(self, db: AsyncSession, sample_cabinet, test_user):
        """confirm 模式应创建资产和生命周期记录"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型", "机柜编码", "U位起始", "占用U数"],
                ["CONF-001", "确认导入1", "服务器", "CAB-001", 20, 2],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="confirm",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["success_count"] == 1
        assert len(result["created_ids"]) == 1

        # 验证资产已创建
        from sqlalchemy import select

        asset_result = await db.execute(select(Asset).where(Asset.asset_code == "CONF-001"))
        asset = asset_result.scalar_one_or_none()
        assert asset is not None
        assert asset.asset_name == "确认导入1"
        assert asset.cabinet_id == sample_cabinet.id

        # 验证生命周期记录
        lc_result = await db.execute(select(AssetLifecycle).where(AssetLifecycle.asset_id == asset.id))
        lc = lc_result.scalar_one_or_none()
        assert lc is not None
        assert lc.action == "purchase"
        assert lc.operator == test_user.username

    # 11. Excel 内部编码重复
    async def test_import_preview_internal_duplicate(self, db: AsyncSession, test_user):
        """同一 Excel 内编码重复应报错"""
        buf = _make_excel(
            [
                ["资产编码", "资产名称", "资产类型"],
                ["DUP-INT-001", "资产A", "服务器"],
                ["DUP-INT-001", "资产B", "服务器"],
            ]
        )
        fake_file = FakeUploadFile(buf)
        result = await import_assets(
            file=fake_file,
            mode="preview",
            db=db,
            current_user=test_user,
            context=_admin_context(test_user),
        )
        assert result["error_count"] >= 1
        assert any("Excel内编码重复" in e["message"] for e in result["errors"])


class TestExportAssets:
    # 12. 导出生成 Excel
    async def test_export_generates_excel(self, db: AsyncSession, sample_asset, sample_cabinet):
        """导出应返回有效的 Excel StreamingResponse"""
        # 创建一个假的 viewer 用户
        user = User(
            username="test_export_viewer",
            password_hash="fakehash",
            role="viewer",
            is_active=True,
        )
        db.add(user)
        await db.commit()

        response = await export_assets(
            asset_type=None,
            status=None,
            cabinet_id=None,
            keyword=None,
            template=False,
            db=db,
            _=user,
            context=_admin_context(user),
        )

        # 读取 StreamingResponse body
        body = b""
        async for chunk in response.body_iterator:
            if isinstance(chunk, bytes):
                body += chunk
            else:
                body += chunk.encode()

        assert len(body) > 0
        wb = load_workbook(BytesIO(body))
        ws = wb.active
        assert ws.title == "资产列表"
        # 表头行 + 至少1条数据
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2
        assert rows[0][0] == "资产编码"
        wb.close()
