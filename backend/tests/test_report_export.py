"""对接报告导出测试 — Story 3.5"""
import pytest
import pytest_asyncio
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.database import Base
from app.models.gateway import DataSource, DataSourcePoint
from app.services.report_export import generate_integration_report


# ============================================================
# Fixtures
# ============================================================

@pytest_asyncio.fixture
async def db_session():
    engine = create_async_engine("sqlite+aiosqlite://", echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    async with session_factory() as session:
        yield session
    await engine.dispose()


# ============================================================
# 测试
# ============================================================

class TestReportExport:
    """对接报告导出测试"""

    def test_generate_valid_excel(self):
        """测试生成有效 Excel 文件"""
        datasources = [
            {"name": "DS-1", "protocol_type": "modbus_tcp", "connection_config": {"host": "192.168.1.1", "port": 502}, "status": "connected", "last_communication": "2026-01-01 12:00:00", "created_at": "2026-01-01 00:00:00", "is_enabled": True},
        ]
        points = [
            {"datasource_name": "DS-1", "address": "40001", "data_type": "float32", "scale": 1.0, "offset": 0.0, "is_dry_contact": False},
        ]
        result = generate_integration_report(datasources, points)
        assert isinstance(result, bytes)
        assert len(result) > 0
        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(result))
        assert wb is not None
        wb.close()

    def test_correct_sheets_and_headers(self):
        """测试 Sheet 名称和列头正确"""
        result = generate_integration_report([], [])
        wb = load_workbook(BytesIO(result))
        sheet_names = wb.sheetnames
        assert "数据源清单" in sheet_names
        assert "点位映射表" in sheet_names

        ws1 = wb["数据源清单"]
        headers1 = [ws1.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers1 == ["名称", "协议类型", "连接参数", "连接状态", "最后通信时间", "创建时间", "启用状态"]

        ws2 = wb["点位映射表"]
        headers2 = [ws2.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers2 == ["数据源名称", "地址", "数据类型", "缩放系数", "偏移量", "是否干接点"]
        wb.close()

    def test_data_filled_correctly(self):
        """测试数据正确填充"""
        datasources = [
            {"name": "温湿度传感器", "protocol_type": "modbus_tcp", "connection_config": {"host": "10.0.0.1"}, "status": "connected", "last_communication": "2026-02-01", "created_at": "2026-01-15", "is_enabled": True},
            {"name": "UPS监控", "protocol_type": "snmp_v2c", "connection_config": {"host": "10.0.0.2"}, "status": "disconnected", "last_communication": None, "created_at": "2026-01-20", "is_enabled": False},
        ]
        points = [
            {"datasource_name": "温湿度传感器", "address": "40001", "data_type": "float32", "scale": 1.0, "offset": 0.0, "is_dry_contact": False},
            {"datasource_name": "温湿度传感器", "address": "40003", "data_type": "uint16", "scale": 0.1, "offset": 0.0, "is_dry_contact": False},
            {"datasource_name": "UPS监控", "address": "1.3.6.1.2.1.1.1.0", "data_type": "string", "scale": 1.0, "offset": 0.0, "is_dry_contact": True},
        ]
        result = generate_integration_report(datasources, points)
        wb = load_workbook(BytesIO(result))

        ws1 = wb["数据源清单"]
        assert ws1.cell(row=2, column=1).value == "温湿度传感器"
        assert ws1.cell(row=2, column=2).value == "modbus_tcp"
        assert ws1.cell(row=2, column=7).value == "是"
        assert ws1.cell(row=3, column=1).value == "UPS监控"
        assert ws1.cell(row=3, column=7).value == "否"

        ws2 = wb["点位映射表"]
        assert ws2.cell(row=2, column=1).value == "温湿度传感器"
        assert ws2.cell(row=2, column=2).value == "40001"
        assert ws2.cell(row=4, column=6).value == "是"  # UPS dry contact
        wb.close()

    def test_empty_report(self):
        """测试无数据时返回空报告（仅表头）"""
        result = generate_integration_report([], [])
        wb = load_workbook(BytesIO(result))

        ws1 = wb["数据源清单"]
        assert ws1.cell(row=1, column=1).value == "名称"
        assert ws1.cell(row=2, column=1).value is None  # No data rows

        ws2 = wb["点位映射表"]
        assert ws2.cell(row=1, column=1).value == "数据源名称"
        assert ws2.cell(row=2, column=1).value is None
        wb.close()
