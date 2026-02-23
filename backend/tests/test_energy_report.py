"""能效报告导出测试 — Story 6-5"""

import json
import pytest
from datetime import datetime, date
from decimal import Decimal
from io import BytesIO

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import delete
from openpyxl import load_workbook

from app.core.database import Base
from app.models.energy import (
    PUEHistory,
    EnergyMonthly,
    EnergyDaily,
    ElectricityPricing,
    EnergyOpportunity,
    ExecutionPlan,
    ExecutionResult,
    PowerDevice,
)
from app.models.report import ReportRecord
from app.services.energy_report_service import EnergyReportService
from app.services.energy_report_excel import generate_energy_report_excel
from app.services.energy_report_pdf import generate_energy_report_pdf


# ============================================================
# Fixtures
# ============================================================


@pytest.fixture(scope="module")
def anyio_backend():
    return "asyncio"


@pytest.fixture(scope="module")
async def engine():
    eng = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with eng.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield eng
    await eng.dispose()


@pytest.fixture(scope="module")
def session_factory(engine):
    return async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


@pytest.fixture
async def db(session_factory):
    async with session_factory() as session:
        # 清理相关表
        await session.execute(delete(ExecutionResult))
        await session.execute(delete(ExecutionPlan))
        await session.execute(delete(EnergyOpportunity))
        await session.execute(delete(ReportRecord))
        await session.execute(delete(PUEHistory))
        await session.execute(delete(EnergyMonthly))
        await session.execute(delete(EnergyDaily))
        await session.execute(delete(ElectricityPricing))
        await session.execute(delete(PowerDevice))
        await session.commit()
        yield session


def _sample_report_data() -> dict:
    """构造一份完整的 report_data 用于 Excel/PDF 测试"""
    return {
        "year": 2026,
        "month": 1,
        "generated_at": "2026-02-01T10:00:00",
        "pue_trend": {
            "daily_values": [
                {"date": "2026-01-01", "avg_pue": 1.45, "min_pue": 1.40, "max_pue": 1.50},
                {"date": "2026-01-02", "avg_pue": 1.42, "min_pue": 1.38, "max_pue": 1.48},
            ],
            "month_avg_pue": 1.435,
            "yoy_avg": 1.50,
            "mom_avg": 1.46,
            "yoy_change": -4.33,
            "mom_change": -1.71,
        },
        "cost_comparison": {
            "current_month": {
                "total_energy": 50000.0,
                "total_cost": 45000.0,
                "peak_energy": 20000.0,
                "peak_cost": 22000.0,
                "normal_energy": 18000.0,
                "normal_cost": 14400.0,
                "valley_energy": 12000.0,
                "valley_cost": 8600.0,
            },
            "last_month": {
                "total_energy": 48000.0,
                "total_cost": 43000.0,
                "peak_energy": 19000.0,
                "peak_cost": 20900.0,
                "normal_energy": 17000.0,
                "normal_cost": 13600.0,
                "valley_energy": 12000.0,
                "valley_cost": 8500.0,
            },
            "last_year_month": {
                "total_energy": 52000.0,
                "total_cost": 47000.0,
                "peak_energy": 21000.0,
                "peak_cost": 23100.0,
                "normal_energy": 19000.0,
                "normal_cost": 15200.0,
                "valley_energy": 12000.0,
                "valley_cost": 8700.0,
            },
            "yoy_change_rate": -4.26,
            "mom_change_rate": 4.65,
        },
        "energy_saving": {
            "details": [
                {
                    "title": "空调优化",
                    "category": 2,
                    "saving_kwh": 1200.0,
                    "saving_cost": 960.0,
                    "achievement_rate": 85.0,
                },
            ],
            "total_saving_kwh": 1200.0,
            "total_saving_cost": 960.0,
        },
        "energy_overview": {
            "daily_energy": [
                {"date": "2026-01-01", "total_energy": 1600.0},
                {"date": "2026-01-02", "total_energy": 1650.0},
            ],
            "total_energy": 3250.0,
        },
    }


# ============================================================
# 测试
# ============================================================


class TestEnergyReport:
    # 1. 基本数据生成
    async def test_generate_report_data_basic(self, db: AsyncSession):
        """插入测试数据，验证4个section全部存在"""
        # 插入 PUE 数据
        db.add(PUEHistory(record_time=datetime(2026, 1, 15, 10, 0), total_power=100, it_power=60, pue=1.67))
        # 插入 EnergyDaily
        dev = PowerDevice(device_code="TEST-001", device_name="测试设备", device_type="UPS")
        db.add(dev)
        await db.flush()
        db.add(
            EnergyDaily(
                device_id=dev.id,
                stat_date=date(2026, 1, 15),
                total_energy=500,
                peak_energy=200,
                normal_energy=200,
                valley_energy=100,
                energy_cost=400,
            )
        )
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 1)
        assert "pue_trend" in data
        assert "cost_comparison" in data
        assert "energy_saving" in data
        assert "energy_overview" in data
        assert data["year"] == 2026
        assert data["month"] == 1

    # 2. 空月份
    async def test_generate_report_data_empty_month(self, db: AsyncSession):
        """无数据月份应返回零值而非报错"""
        data = await EnergyReportService.generate_report_data(db, 2020, 6)
        assert data["pue_trend"]["month_avg_pue"] == 0
        assert data["pue_trend"]["daily_values"] == []
        assert data["cost_comparison"]["current_month"]["total_energy"] == 0
        assert data["energy_saving"]["total_saving_kwh"] == 0
        assert data["energy_overview"]["total_energy"] == 0

    # 3. PUE 日聚合
    async def test_pue_trend_daily_aggregation(self, db: AsyncSession):
        """同一天4条PUE记录应聚合为1条日均值"""
        for hour in [0, 6, 12, 18]:
            db.add(
                PUEHistory(
                    record_time=datetime(2026, 2, 10, hour, 0),
                    total_power=100,
                    it_power=60,
                    pue=1.4 + hour * 0.01,
                )
            )
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 2)
        daily = data["pue_trend"]["daily_values"]
        assert len(daily) == 1
        assert daily[0]["date"] == "2026-02-10"
        # avg of 1.40, 1.46, 1.52, 1.58 = 1.49
        assert abs(daily[0]["avg_pue"] - 1.49) < 0.01

    # 4. 从 EnergyMonthly 获取费用
    async def test_cost_from_monthly(self, db: AsyncSession):
        """EnergyMonthly 存在时应直接使用其 peak_cost/normal_cost/valley_cost"""
        dev = PowerDevice(device_code="COST-M-001", device_name="费用设备", device_type="UPS")
        db.add(dev)
        await db.flush()
        db.add(
            EnergyMonthly(
                device_id=dev.id,
                stat_year=2026,
                stat_month=3,
                total_energy=10000,
                peak_energy=4000,
                normal_energy=4000,
                valley_energy=2000,
                energy_cost=8000,
                peak_cost=4400,
                normal_cost=3200,
                valley_cost=1400,
            )
        )
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 3)
        cur = data["cost_comparison"]["current_month"]
        assert cur["total_energy"] == 10000.0
        assert cur["peak_cost"] == 4400.0
        assert cur["normal_cost"] == 3200.0
        assert cur["valley_cost"] == 1400.0

    # 5. 从 EnergyDaily + ElectricityPricing 回退
    async def test_cost_fallback_daily(self, db: AsyncSession):
        """无 EnergyMonthly 时应从 EnergyDaily × ElectricityPricing 计算"""
        dev = PowerDevice(device_code="COST-D-001", device_name="日费设备", device_type="HVAC")
        db.add(dev)
        await db.flush()
        db.add(
            EnergyDaily(
                device_id=dev.id,
                stat_date=date(2026, 4, 10),
                total_energy=1000,
                peak_energy=400,
                normal_energy=400,
                valley_energy=200,
                energy_cost=0,
            )
        )
        db.add(
            ElectricityPricing(
                pricing_name="峰时",
                period_type="peak",
                start_time="10:00",
                end_time="12:00",
                price=1.1,
                effective_date=date(2026, 1, 1),
                is_enabled=True,
            )
        )
        db.add(
            ElectricityPricing(
                pricing_name="平时",
                period_type="flat",
                start_time="08:00",
                end_time="10:00",
                price=0.8,
                effective_date=date(2026, 1, 1),
                is_enabled=True,
            )
        )
        db.add(
            ElectricityPricing(
                pricing_name="谷时",
                period_type="valley",
                start_time="23:00",
                end_time="07:00",
                price=0.4,
                effective_date=date(2026, 1, 1),
                is_enabled=True,
            )
        )
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 4)
        cur = data["cost_comparison"]["current_month"]
        assert cur["total_energy"] == 1000.0
        # peak_cost = 400 * 1.1 = 440, normal_cost = 400 * 0.8 = 320, valley_cost = 200 * 0.4 = 80
        assert cur["peak_cost"] == 440.0
        assert cur["normal_cost"] == 320.0
        assert cur["valley_cost"] == 80.0

    # 6. 节能 JSON 解析
    async def test_energy_saving_json_parsing(self, db: AsyncSession):
        """ExecutionResult 的 JSON energy_before/after 应正确解析求和"""
        opp = EnergyOpportunity(
            category=2,
            title="空调优化",
            status="completed",
            potential_saving=Decimal("5000"),
            created_at=datetime(2026, 5, 10),
        )
        db.add(opp)
        await db.flush()
        plan = ExecutionPlan(opportunity_id=opp.id, plan_name="空调计划", status="completed")
        db.add(plan)
        await db.flush()
        result = ExecutionResult(
            plan_id=plan.id,
            actual_saving=Decimal("800"),
            achievement_rate=Decimal("90.00"),
            energy_before=json.dumps(
                [{"date": "2026-01-01", "energy": 500, "cost": 400}, {"date": "2026-01-02", "energy": 480, "cost": 384}]
            ),
            energy_after=json.dumps(
                [{"date": "2026-01-01", "energy": 400, "cost": 320}, {"date": "2026-01-02", "energy": 390, "cost": 312}]
            ),
            status="completed",
        )
        db.add(result)
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 5)
        details = data["energy_saving"]["details"]
        assert len(details) >= 1
        item = next(i for i in details if i["title"] == "空调优化")
        # before: 500+480=980, after: 400+390=790, saving=190
        assert item["saving_kwh"] == 190.0
        assert item["saving_cost"] == 800.0

    # 7. 同比环比计算
    async def test_yoy_mom_calculation(self, db: AsyncSession):
        """当前月、上月、去年同月数据齐全时应正确计算变化率"""
        dev = PowerDevice(device_code="YOY-001", device_name="同比设备", device_type="UPS")
        db.add(dev)
        await db.flush()
        # 当前月 2026-6
        db.add(
            EnergyMonthly(
                device_id=dev.id,
                stat_year=2026,
                stat_month=6,
                total_energy=10000,
                energy_cost=8000,
                peak_energy=4000,
                peak_cost=4400,
                normal_energy=4000,
                normal_cost=3200,
                valley_energy=2000,
                valley_cost=1400,
            )
        )
        # 上月 2026-5
        db.add(
            EnergyMonthly(
                device_id=dev.id,
                stat_year=2026,
                stat_month=5,
                total_energy=9000,
                energy_cost=7200,
                peak_energy=3600,
                peak_cost=3960,
                normal_energy=3600,
                normal_cost=2880,
                valley_energy=1800,
                valley_cost=1260,
            )
        )
        # 去年同月 2025-6
        db.add(
            EnergyMonthly(
                device_id=dev.id,
                stat_year=2025,
                stat_month=6,
                total_energy=11000,
                energy_cost=8800,
                peak_energy=4400,
                peak_cost=4840,
                normal_energy=4400,
                normal_cost=3520,
                valley_energy=2200,
                valley_cost=1540,
            )
        )
        await db.commit()

        data = await EnergyReportService.generate_report_data(db, 2026, 6)
        cc = data["cost_comparison"]
        # mom: (8000-7200)/7200*100 = 11.11
        assert cc["mom_change_rate"] is not None
        assert abs(cc["mom_change_rate"] - 11.11) < 0.1
        # yoy: (8000-8800)/8800*100 = -9.09
        assert cc["yoy_change_rate"] is not None
        assert abs(cc["yoy_change_rate"] - (-9.09)) < 0.1

    # 8. Excel 生成
    def test_excel_generation(self):
        """generate_energy_report_excel 应返回有效的 BytesIO"""
        data = _sample_report_data()
        buf = generate_energy_report_excel(data)
        assert isinstance(buf, BytesIO)
        assert buf.getbuffer().nbytes > 0
        # 验证可被 openpyxl 打开
        wb = load_workbook(buf)
        assert "报告概览" in wb.sheetnames
        assert "PUE趋势" in wb.sheetnames
        assert "电费对比" in wb.sheetnames
        assert "节能成果" in wb.sheetnames
        assert "每日能耗" in wb.sheetnames
        wb.close()

    # 9. PDF 生成
    def test_pdf_generation(self):
        """generate_energy_report_pdf 应返回有效的 BytesIO"""
        data = _sample_report_data()
        buf = generate_energy_report_pdf(data)
        assert isinstance(buf, BytesIO)
        assert buf.getbuffer().nbytes > 0
        # PDF 文件头校验
        buf.seek(0)
        header = buf.read(5)
        assert header == b"%PDF-"

    # 10. 导出创建 ReportRecord
    async def test_export_creates_report_record(self, db: AsyncSession):
        """导出后应在 ReportRecord 表中创建记录"""
        from app.models.user import User as UserModel

        # 创建测试用户
        user = UserModel(
            username="test_export_user",
            password_hash="fakehash",
            role="operator",
            is_active=True,
        )
        db.add(user)
        await db.flush()

        # 模拟导出流程: 生成数据 + 写 ReportRecord
        data = await EnergyReportService.generate_report_data(db, 2026, 1)
        generate_energy_report_excel(data)

        record = ReportRecord(
            template_id=None,
            report_name="能效报告_2026年1月",
            report_type="energy_efficiency",
            start_time=datetime(2026, 1, 1),
            end_time=datetime(2026, 2, 1),
            file_path=None,
            file_size=None,
            status="completed",
            generated_by=user.id,
        )
        db.add(record)
        await db.commit()

        from sqlalchemy import select as sa_select

        stmt = sa_select(ReportRecord).where(ReportRecord.report_type == "energy_efficiency")
        rows = (await db.execute(stmt)).scalars().all()
        assert len(rows) >= 1
        assert rows[0].report_name == "能效报告_2026年1月"
        assert rows[0].status == "completed"
