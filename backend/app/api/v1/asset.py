"""
资产管理 API - v1
"""
from datetime import datetime, timedelta, date
from typing import Optional, List, Dict, Any
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, update
from starlette.responses import StreamingResponse
from pydantic import BaseModel as PydanticBaseModel
import openpyxl

from ..deps import get_db, require_viewer, require_operator
from ...models.user import User
from ...models.asset import (
    Asset, Cabinet, AssetLifecycle, MaintenanceRecord,
    AssetInventory, AssetInventoryItem, AssetStatus, AssetType
)
from ...schemas.asset import (
    CabinetCreate, CabinetUpdate, CabinetResponse,
    AssetCreate, AssetUpdate, AssetResponse,
    LifecycleResponse,
    MaintenanceCreate, MaintenanceResponse,
    InventoryCreate, InventoryItemUpdate, InventoryResponse, InventoryItemResponse,
    AssetStatistics,
    WarrantyAlertItem, WarrantyAlertResponse
)
from ...schemas.common import PageResponse


# ==================== 导入/导出映射 ====================

IMPORT_COLUMN_MAP = {
    "资产编码": "asset_code",
    "资产名称": "asset_name",
    "资产类型": "asset_type",
    "品牌": "brand",
    "型号": "model",
    "序列号": "serial_number",
    "机柜编码": "_cabinet_code",
    "U位起始": "u_position",
    "占用U数": "u_height",
    "采购日期": "purchase_date",
    "保修开始": "warranty_start",
    "保修截止": "warranty_end",
    "供应商": "supplier",
    "负责人": "owner",
    "部门": "department",
    "采购价格": "purchase_price",
    "备注": "remark",
}

ASSET_TYPE_MAP = {
    "服务器": "server", "网络设备": "network", "存储设备": "storage",
    "UPS": "ups", "PDU": "pdu", "空调": "ac", "机柜": "cabinet",
    "传感器": "sensor", "其他": "other",
    "server": "server", "network": "network", "storage": "storage",
    "ups": "ups", "pdu": "pdu", "ac": "ac", "cabinet": "cabinet",
    "sensor": "sensor", "other": "other",
}

EXPORT_COLUMNS = [
    ("资产编码", "asset_code"),
    ("资产名称", "asset_name"),
    ("资产类型", "asset_type"),
    ("品牌", "brand"),
    ("型号", "model"),
    ("序列号", "serial_number"),
    ("机柜编码", "_cabinet_code"),
    ("U位起始", "u_position"),
    ("占用U数", "u_height"),
    ("状态", "status"),
    ("采购日期", "purchase_date"),
    ("保修开始", "warranty_start"),
    ("保修截止", "warranty_end"),
    ("供应商", "supplier"),
    ("负责人", "owner"),
    ("部门", "department"),
    ("采购价格", "purchase_price"),
    ("备注", "remark"),
]

STATUS_CN_MAP = {
    "in_stock": "库存中", "in_use": "使用中", "borrowed": "借出",
    "maintenance": "维护中", "scrapped": "已报废",
}


# ==================== U 位冲突校验 ====================

async def _check_u_position_conflict(
    db: AsyncSession,
    cabinet_id: int,
    u_position: int,
    u_height: int,
    exclude_asset_id: Optional[int] = None
) -> Optional[str]:
    """检查 U 位是否冲突，返回冲突信息或 None"""
    if cabinet_id is None or u_position is None or u_height is None:
        return None

    query = select(Asset).where(
        Asset.cabinet_id == cabinet_id,
        Asset.u_position.isnot(None),
        Asset.u_height.isnot(None)
    )
    if exclude_asset_id:
        query = query.where(Asset.id != exclude_asset_id)

    result = await db.execute(query)
    existing_assets = result.scalars().all()

    new_start = u_position
    new_end = u_position + u_height - 1

    for existing in existing_assets:
        ex_start = existing.u_position
        ex_end = existing.u_position + existing.u_height - 1
        if new_start <= ex_end and new_end >= ex_start:
            return f"U位冲突: U{new_start}-U{new_end} 与资产 {existing.asset_code}({existing.asset_name}) 的 U{ex_start}-U{ex_end} 重叠"

    return None


router = APIRouter(prefix="/asset", tags=["资产管理"])


# ==================== 机柜管理 ====================

@router.get("/cabinets", response_model=List[CabinetResponse], summary="获取机柜列表")
async def get_cabinets(
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回记录数"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取机柜列表（分页）
    """
    query = select(Cabinet).offset(skip).limit(limit)
    result = await db.execute(query)
    cabinets = result.scalars().all()

    # 计算每个机柜的已使用U数和可用U数
    cabinet_list = []
    for cabinet in cabinets:
        # 获取该机柜中所有资产占用的U数
        used_u_result = await db.execute(
            select(func.sum(Asset.u_height)).where(
                Asset.cabinet_id == cabinet.id,
                Asset.u_height.isnot(None)
            )
        )
        used_u = used_u_result.scalar() or 0

        cabinet_data = CabinetResponse.model_validate(cabinet)
        cabinet_data.used_u = used_u
        cabinet_data.available_u = (cabinet.total_u or 42) - used_u
        cabinet_list.append(cabinet_data)

    return cabinet_list


@router.get("/cabinets/{cabinet_id}", response_model=CabinetResponse, summary="获取机柜详情")
async def get_cabinet(
    cabinet_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    根据ID获取机柜详情
    """
    result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = result.scalar_one_or_none()

    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    # 计算已使用U数
    used_u_result = await db.execute(
        select(func.sum(Asset.u_height)).where(
            Asset.cabinet_id == cabinet_id,
            Asset.u_height.isnot(None)
        )
    )
    used_u = used_u_result.scalar() or 0

    cabinet_data = CabinetResponse.model_validate(cabinet)
    cabinet_data.used_u = used_u
    cabinet_data.available_u = (cabinet.total_u or 42) - used_u

    return cabinet_data


@router.get("/cabinets/{cabinet_id}/usage", summary="获取机柜U位使用情况")
async def get_cabinet_usage(
    cabinet_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
) -> Dict[str, Any]:
    """
    获取机柜U位使用情况，包含U位映射表
    """
    result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = result.scalar_one_or_none()

    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    total_u = cabinet.total_u or 42

    # 获取该机柜中所有资产
    assets_result = await db.execute(
        select(Asset).where(
            Asset.cabinet_id == cabinet_id,
            Asset.u_position.isnot(None),
            Asset.u_height.isnot(None)
        )
    )
    assets = assets_result.scalars().all()

    # 构建U位映射表
    u_map = {}
    used_u = 0

    for asset in assets:
        if asset.u_position and asset.u_height:
            for u in range(asset.u_position, asset.u_position + asset.u_height):
                if u <= total_u:
                    u_map[str(u)] = {
                        "asset_id": asset.id,
                        "asset_code": asset.asset_code,
                        "asset_name": asset.asset_name,
                        "asset_type": asset.asset_type.value if asset.asset_type else None
                    }
            used_u += asset.u_height

    available_u = total_u - used_u
    usage_rate = round((used_u / total_u * 100), 2) if total_u > 0 else 0

    # 构建资产列表（用于 U 位可视化）
    assets_list = []
    for asset in assets:
        if asset.u_position is not None and asset.u_height is not None:
            assets_list.append({
                "asset_id": asset.id,
                "asset_code": asset.asset_code,
                "asset_name": asset.asset_name,
                "asset_type": asset.asset_type.value if asset.asset_type else None,
                "model": asset.model or "",
                "brand": asset.brand or "",
                "status": asset.status.value if asset.status else None,
                "u_position": asset.u_position,
                "u_height": asset.u_height,
            })

    return {
        "cabinet_id": cabinet_id,
        "cabinet_name": cabinet.cabinet_name,
        "total_u": total_u,
        "used_u": used_u,
        "available_u": available_u,
        "usage_rate": usage_rate,
        "u_map": u_map,
        "assets": assets_list
    }



class MoveAssetRequest(PydanticBaseModel):
    asset_id: int
    new_u_position: int


@router.put("/cabinets/{cabinet_id}/move-asset", summary="拖拽移动资产U位")
async def move_asset_in_cabinet(
    cabinet_id: int,
    data: MoveAssetRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """在同一机柜内移动资产到新的U位位置"""
    # 1. 校验机柜存在
    cab_result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = cab_result.scalar_one_or_none()
    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    total_u = cabinet.total_u or 42

    # 2. 校验资产存在且属于该机柜
    asset_result = await db.execute(select(Asset).where(Asset.id == data.asset_id))
    asset = asset_result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")
    if asset.cabinet_id != cabinet_id:
        raise HTTPException(status_code=400, detail="资产不属于该机柜")

    u_height = asset.u_height or 1

    # 3. 校验范围
    if data.new_u_position < 1:
        raise HTTPException(status_code=400, detail="U位起始位置不能小于1")
    if data.new_u_position + u_height - 1 > total_u:
        raise HTTPException(status_code=400, detail=f"U位超出机柜范围（最大{total_u}U）")

    # 4. U位冲突校验
    conflict = await _check_u_position_conflict(db, cabinet_id, data.new_u_position, u_height, exclude_asset_id=asset.id)
    if conflict:
        raise HTTPException(status_code=400, detail=conflict)

    # 5. 记录旧位置并更新
    old_u_position = asset.u_position
    asset.u_position = data.new_u_position

    # 6. 创建生命周期记录
    lifecycle = AssetLifecycle(
        asset_id=asset.id,
        action="move",
        action_date=datetime.now(),
        operator=current_user.username,
        from_location=f"U{old_u_position}" if old_u_position else "",
        to_location=f"U{data.new_u_position}",
        remark="U位拖拽移动"
    )
    db.add(lifecycle)
    await db.commit()

    # 7. 返回更新后的 usage 数据
    return await get_cabinet_usage(cabinet_id, db, current_user)


@router.post("/cabinets", response_model=CabinetResponse, summary="创建机柜")
async def create_cabinet(
    data: CabinetCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    创建新机柜
    """
    # 检查编码是否已存在
    existing = await db.execute(
        select(Cabinet).where(Cabinet.cabinet_code == data.cabinet_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="机柜编码已存在")

    cabinet = Cabinet(**data.model_dump())
    db.add(cabinet)
    await db.commit()
    await db.refresh(cabinet)

    cabinet_data = CabinetResponse.model_validate(cabinet)
    cabinet_data.used_u = 0
    cabinet_data.available_u = cabinet.total_u or 42

    return cabinet_data


@router.put("/cabinets/{cabinet_id}", response_model=CabinetResponse, summary="更新机柜")
async def update_cabinet(
    cabinet_id: int,
    data: CabinetUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    更新机柜信息
    """
    result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = result.scalar_one_or_none()

    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    # 如果更新编码，检查是否已存在
    update_data = data.model_dump(exclude_unset=True)
    if "cabinet_code" in update_data and update_data["cabinet_code"] != cabinet.cabinet_code:
        existing = await db.execute(
            select(Cabinet).where(Cabinet.cabinet_code == update_data["cabinet_code"])
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="机柜编码已存在")

    for key, value in update_data.items():
        if value is not None:
            setattr(cabinet, key, value)

    cabinet.updated_at = datetime.now()
    await db.commit()
    await db.refresh(cabinet)

    # 计算已使用U数
    used_u_result = await db.execute(
        select(func.sum(Asset.u_height)).where(
            Asset.cabinet_id == cabinet_id,
            Asset.u_height.isnot(None)
        )
    )
    used_u = used_u_result.scalar() or 0

    cabinet_data = CabinetResponse.model_validate(cabinet)
    cabinet_data.used_u = used_u
    cabinet_data.available_u = (cabinet.total_u or 42) - used_u

    return cabinet_data


@router.delete("/cabinets/{cabinet_id}", summary="删除机柜")
async def delete_cabinet(
    cabinet_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    删除机柜（如果有关联资产则不允许删除）
    """
    result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = result.scalar_one_or_none()

    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    # 检查是否有关联资产
    asset_count_result = await db.execute(
        select(func.count(Asset.id)).where(Asset.cabinet_id == cabinet_id)
    )
    asset_count = asset_count_result.scalar()

    if asset_count > 0:
        raise HTTPException(status_code=400, detail="机柜下存在关联资产，无法删除")

    await db.delete(cabinet)
    await db.commit()

    return {"message": "机柜删除成功"}


# ==================== 资产管理 ====================

@router.get("/assets", response_model=List[AssetResponse], summary="获取资产列表")
async def get_assets(
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回记录数"),
    asset_type: Optional[AssetType] = Query(None, description="资产类型"),
    status: Optional[AssetStatus] = Query(None, description="资产状态"),
    cabinet_id: Optional[int] = Query(None, description="机柜ID"),
    keyword: Optional[str] = Query(None, description="关键词(资产编码、名称、品牌、型号)"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取资产列表（多条件筛选、分页）
    """
    query = select(Asset)

    conditions = []
    if asset_type:
        conditions.append(Asset.asset_type == asset_type)
    if status:
        conditions.append(Asset.status == status)
    if cabinet_id:
        conditions.append(Asset.cabinet_id == cabinet_id)
    if keyword:
        keyword_filter = or_(
            Asset.asset_code.contains(keyword),
            Asset.asset_name.contains(keyword),
            Asset.brand.contains(keyword),
            Asset.model.contains(keyword)
        )
        conditions.append(keyword_filter)

    if conditions:
        query = query.where(and_(*conditions))

    query = query.order_by(Asset.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    assets = result.scalars().all()

    # 填充机柜名称和保修状态
    asset_list = []
    for asset in assets:
        asset_data = AssetResponse.model_validate(asset)

        # 获取机柜名称
        if asset.cabinet_id:
            cabinet_result = await db.execute(
                select(Cabinet).where(Cabinet.id == asset.cabinet_id)
            )
            cabinet = cabinet_result.scalar_one_or_none()
            if cabinet:
                asset_data.cabinet_name = cabinet.cabinet_name

        # 计算保修状态
        asset_data.warranty_status = _calculate_warranty_status(asset.warranty_end)
        asset_list.append(asset_data)

    return asset_list


@router.post("/assets/import", summary="批量导入资产")
async def import_assets(
    file: UploadFile = File(..., description="Excel文件"),
    mode: str = Query("preview", description="模式: preview(预校验) / confirm(确认导入)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """批量导入资产（预校验 / 确认导入）"""
    # 读取 Excel
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析Excel文件")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel文件无数据行")

    # 解析表头
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map: Dict[int, str] = {}
    for idx, col_name in enumerate(header):
        if col_name in IMPORT_COLUMN_MAP:
            col_map[idx] = IMPORT_COLUMN_MAP[col_name]

    errors: List[Dict[str, Any]] = []
    preview_data: List[Dict[str, Any]] = []
    seen_codes: set = set()

    # 预加载所有已有 asset_code
    existing_codes_result = await db.execute(select(Asset.asset_code))
    existing_codes = {r[0] for r in existing_codes_result.all()}

    # 预加载所有机柜 code -> id 映射
    cabinet_result = await db.execute(select(Cabinet))
    cabinets = cabinet_result.scalars().all()
    cabinet_map = {c.cabinet_code: c for c in cabinets}

    # 同批次 U 位占用追踪: {cabinet_id: [(row_idx, start_u, end_u), ...]}
    pending_u_ranges: Dict[int, List[tuple]] = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        row_data: Dict[str, Any] = {}
        for col_idx, field_name in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                val = str(val).strip() if not isinstance(val, (int, float, datetime, date)) else val
            row_data[field_name] = val

        row_errors: List[Dict[str, str]] = []

        # 必填检查
        if not row_data.get("asset_code"):
            row_errors.append({"row": row_idx, "field": "asset_code", "message": "资产编码不能为空"})
        if not row_data.get("asset_name"):
            row_errors.append({"row": row_idx, "field": "asset_name", "message": "资产名称不能为空"})
        if not row_data.get("asset_type"):
            row_errors.append({"row": row_idx, "field": "asset_type", "message": "资产类型不能为空"})

        # asset_code 唯一性
        code = row_data.get("asset_code")
        if code:
            code = str(code)
            if code in existing_codes:
                row_errors.append({"row": row_idx, "field": "asset_code", "message": "编码已存在"})
            elif code in seen_codes:
                row_errors.append({"row": row_idx, "field": "asset_code", "message": "Excel内编码重复"})
            else:
                seen_codes.add(code)

        # asset_type 枚举校验
        raw_type = str(row_data.get("asset_type", "")) if row_data.get("asset_type") else ""
        mapped_type = ASSET_TYPE_MAP.get(raw_type)
        if raw_type and not mapped_type:
            row_errors.append({"row": row_idx, "field": "asset_type", "message": f"无效的资产类型: {raw_type}"})
        if mapped_type:
            row_data["asset_type"] = mapped_type

        # 机柜编码 -> cabinet_id
        cabinet_code = row_data.pop("_cabinet_code", None)
        cabinet_id = None
        if cabinet_code:
            cabinet_code = str(cabinet_code)
            cab = cabinet_map.get(cabinet_code)
            if not cab:
                row_errors.append({"row": row_idx, "field": "_cabinet_code", "message": f"机柜编码不存在: {cabinet_code}"})
            else:
                cabinet_id = cab.id
                row_data["cabinet_id"] = cabinet_id

        # U 位数值转换
        for int_field in ("u_position", "u_height"):
            if row_data.get(int_field) is not None:
                try:
                    row_data[int_field] = int(row_data[int_field])
                except (ValueError, TypeError):
                    row_errors.append({"row": row_idx, "field": int_field, "message": f"{int_field} 必须为整数"})
                    row_data[int_field] = None

        # U 位冲突校验（含数据库已有 + 同批次 Excel 行间冲突）
        u_pos = row_data.get("u_position")
        u_h = row_data.get("u_height")
        if cabinet_id and u_pos and u_h and not row_errors:
            # 检查与数据库已有资产的冲突
            conflict = await _check_u_position_conflict(db, cabinet_id, u_pos, u_h)
            if conflict:
                row_errors.append({"row": row_idx, "field": "u_position", "message": conflict})
            else:
                # 检查与同批次前面行的冲突
                key = cabinet_id
                new_start = u_pos
                new_end = u_pos + u_h - 1
                for prev_row, prev_start, prev_end in pending_u_ranges.get(key, []):
                    if new_start <= prev_end and new_end >= prev_start:
                        row_errors.append({
                            "row": row_idx,
                            "field": "u_position",
                            "message": f"与第 {prev_row} 行 U 位冲突 (U{prev_start}-U{prev_end})"
                        })
                        break
                else:
                    # 无冲突，记录本行的 U 位范围
                    pending_u_ranges.setdefault(key, []).append((row_idx, new_start, new_end))

        # 日期字段转换
        for date_field in ("purchase_date", "warranty_start", "warranty_end"):
            val = row_data.get(date_field)
            if val is not None and not isinstance(val, date):
                try:
                    row_data[date_field] = datetime.strptime(str(val), "%Y-%m-%d").date()
                except (ValueError, TypeError):
                    try:
                        row_data[date_field] = datetime.strptime(str(val), "%Y/%m/%d").date()
                    except (ValueError, TypeError):
                        row_errors.append({"row": row_idx, "field": date_field, "message": f"日期格式无效: {val}"})
                        row_data[date_field] = None
            elif isinstance(val, datetime):
                row_data[date_field] = val.date()

        errors.extend(row_errors)
        preview_data.append(row_data)

    total = len(preview_data)
    error_rows = {e["row"] for e in errors}
    success_count = total - len(error_rows)
    error_count = len(error_rows)

    if mode == "preview":
        return {
            "total": total,
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors,
            "preview_data": preview_data
        }

    # confirm 模式 — 只导入无错误的行
    if error_count > 0:
        raise HTTPException(status_code=400, detail={
            "message": "存在校验错误，无法导入",
            "errors": errors
        })

    created_ids = []
    try:
        for row_data in preview_data:
            # 清理非模型字段
            row_data.pop("_cabinet_code", None)
            # 设置 asset_type 枚举
            if "asset_type" in row_data and row_data["asset_type"]:
                row_data["asset_type"] = AssetType(row_data["asset_type"])

            asset = Asset(**{k: v for k, v in row_data.items() if v is not None})
            db.add(asset)
            await db.flush()

            # 添加生命周期记录
            lifecycle = AssetLifecycle(
                asset_id=asset.id,
                action="purchase",
                action_date=datetime.now(),
                operator=current_user.username,
                remark="批量导入创建"
            )
            db.add(lifecycle)
            created_ids.append(asset.id)

        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败，已回滚: {str(e)}")

    return {
        "total": total,
        "success_count": len(created_ids),
        "error_count": 0,
        "errors": [],
        "created_ids": created_ids
    }


@router.get("/assets/export", summary="导出资产列表")
async def export_assets(
    asset_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    cabinet_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    template: bool = Query(False, description="是否只下载空模板"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """导出资产列表为 Excel，template=true 时只返回表头"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产列表"

    # 写表头
    headers = [col[0] for col in EXPORT_COLUMNS]
    ws.append(headers)

    if not template:
        query = select(Asset)

        conditions = []
        if asset_type:
            conditions.append(Asset.asset_type == asset_type)
        if status:
            conditions.append(Asset.status == status)
        if cabinet_id:
            conditions.append(Asset.cabinet_id == cabinet_id)
        if keyword:
            keyword_filter = or_(
                Asset.asset_code.contains(keyword),
                Asset.asset_name.contains(keyword),
                Asset.brand.contains(keyword),
                Asset.model.contains(keyword)
            )
            conditions.append(keyword_filter)

        if conditions:
            query = query.where(and_(*conditions))

        query = query.order_by(Asset.created_at.desc())
        result = await db.execute(query)
        assets = result.scalars().all()

        # 预加载机柜映射
        cab_result = await db.execute(select(Cabinet))
        cab_map = {c.id: c.cabinet_code for c in cab_result.scalars().all()}

        # 资产类型反向映射
        type_cn_map = {
            "server": "服务器", "network": "网络设备", "storage": "存储设备",
            "ups": "UPS", "pdu": "PDU", "ac": "空调", "cabinet": "机柜",
            "sensor": "传感器", "other": "其他",
        }

        # 写数据
        for asset in assets:
            row = []
            for _, field in EXPORT_COLUMNS:
                if field == "_cabinet_code":
                    row.append(cab_map.get(asset.cabinet_id, ""))
                elif field == "asset_type":
                    val = asset.asset_type.value if asset.asset_type else ""
                    row.append(type_cn_map.get(val, val))
                elif field == "status":
                    val = asset.status.value if asset.status else ""
                    row.append(STATUS_CN_MAP.get(val, val))
                elif field in ("purchase_date", "warranty_start", "warranty_end"):
                    val = getattr(asset, field, None)
                    row.append(str(val) if val else "")
                else:
                    row.append(getattr(asset, field, "") or "")
            ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets_export.xlsx"}
    )


@router.get("/assets/{asset_id}", response_model=AssetResponse, summary="获取资产详情")
async def get_asset(
    asset_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    根据ID获取资产详情
    """
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()

    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    asset_data = AssetResponse.model_validate(asset)

    # 获取机柜名称
    if asset.cabinet_id:
        cabinet_result = await db.execute(
            select(Cabinet).where(Cabinet.id == asset.cabinet_id)
        )
        cabinet = cabinet_result.scalar_one_or_none()
        if cabinet:
            asset_data.cabinet_name = cabinet.cabinet_name

    # 计算保修状态
    asset_data.warranty_status = _calculate_warranty_status(asset.warranty_end)

    return asset_data


@router.post("/assets", response_model=AssetResponse, summary="创建资产")
async def create_asset(
    data: AssetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """
    创建新资产
    """
    # 检查编码是否已存在
    existing = await db.execute(
        select(Asset).where(Asset.asset_code == data.asset_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="资产编码已存在")

    # 如果指定了机柜，检查机柜是否存在
    cabinet_name = None
    if data.cabinet_id:
        cabinet_result = await db.execute(
            select(Cabinet).where(Cabinet.id == data.cabinet_id)
        )
        cabinet = cabinet_result.scalar_one_or_none()
        if not cabinet:
            raise HTTPException(status_code=400, detail="指定的机柜不存在")
        cabinet_name = cabinet.cabinet_name

    # U 位冲突校验
    if data.cabinet_id and data.u_position and data.u_height:
        conflict = await _check_u_position_conflict(db, data.cabinet_id, data.u_position, data.u_height)
        if conflict:
            raise HTTPException(status_code=400, detail=conflict)

    asset = Asset(**data.model_dump())
    db.add(asset)
    await db.commit()
    await db.refresh(asset)

    # 添加生命周期记录
    to_location = None
    if cabinet_name:
        to_location = f"{cabinet_name} U{asset.u_position}" if asset.u_position else cabinet_name

    lifecycle = AssetLifecycle(
        asset_id=asset.id,
        action="purchase",
        action_date=datetime.now(),
        operator=current_user.username,
        from_location=None,
        to_location=to_location,
        remark="资产创建入库"
    )
    db.add(lifecycle)
    await db.commit()

    asset_data = AssetResponse.model_validate(asset)
    asset_data.cabinet_name = cabinet_name
    asset_data.warranty_status = _calculate_warranty_status(asset.warranty_end)

    return asset_data


@router.put("/assets/{asset_id}", response_model=AssetResponse, summary="更新资产")
async def update_asset(
    asset_id: int,
    data: AssetUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """
    更新资产信息
    """
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()

    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    update_data = data.model_dump(exclude_unset=True)

    # 如果更新编码，检查是否已存在
    if "asset_code" in update_data and update_data["asset_code"] != asset.asset_code:
        existing = await db.execute(
            select(Asset).where(Asset.asset_code == update_data["asset_code"])
        )
        if existing.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="资产编码已存在")

    # U 位冲突校验
    check_cabinet = update_data.get("cabinet_id", asset.cabinet_id)
    check_u_pos = update_data.get("u_position", asset.u_position)
    check_u_height = update_data.get("u_height", asset.u_height)
    if check_cabinet and check_u_pos and check_u_height:
        conflict = await _check_u_position_conflict(db, check_cabinet, check_u_pos, check_u_height, exclude_asset_id=asset_id)
        if conflict:
            raise HTTPException(status_code=400, detail=conflict)

    # 记录位置变更
    old_cabinet_id = asset.cabinet_id
    old_u_position = asset.u_position
    new_cabinet_id = update_data.get("cabinet_id", old_cabinet_id)
    new_u_position = update_data.get("u_position", old_u_position)
    location_changed = (old_cabinet_id != new_cabinet_id) or (old_u_position != new_u_position)

    # 记录状态变更
    old_status = asset.status
    new_status = update_data.get("status", old_status)
    status_changed = old_status != new_status

    # 更新资产属性
    for key, value in update_data.items():
        if value is not None:
            setattr(asset, key, value)

    asset.updated_at = datetime.now()
    await db.commit()
    await db.refresh(asset)

    # 添加位置变更生命周期记录
    if location_changed:
        from_location = None
        to_location = None

        if old_cabinet_id:
            old_cab_result = await db.execute(
                select(Cabinet).where(Cabinet.id == old_cabinet_id)
            )
            old_cabinet = old_cab_result.scalar_one_or_none()
            if old_cabinet:
                from_location = f"{old_cabinet.cabinet_name} U{old_u_position}" if old_u_position else old_cabinet.cabinet_name

        if new_cabinet_id:
            new_cab_result = await db.execute(
                select(Cabinet).where(Cabinet.id == new_cabinet_id)
            )
            new_cabinet = new_cab_result.scalar_one_or_none()
            if new_cabinet:
                to_location = f"{new_cabinet.cabinet_name} U{new_u_position}" if new_u_position else new_cabinet.cabinet_name

        lifecycle = AssetLifecycle(
            asset_id=asset_id,
            action="move",
            action_date=datetime.now(),
            operator=current_user.username,
            from_location=from_location,
            to_location=to_location,
            remark="资产位置变更"
        )
        db.add(lifecycle)
        await db.commit()

    # 添加状态变更生命周期记录
    if status_changed:
        action = "status_change"
        remark = f"状态变更: {old_status.value if old_status else 'None'} -> {new_status.value if new_status else 'None'}"

        if new_status == AssetStatus.scrapped:
            action = "scrap"
            remark = "资产报废"
        elif new_status == AssetStatus.maintenance:
            action = "maintain"
            remark = "资产送修"
        elif new_status == AssetStatus.in_use:
            action = "deploy"
            remark = "资产部署上线"

        lifecycle = AssetLifecycle(
            asset_id=asset_id,
            action=action,
            action_date=datetime.now(),
            operator=current_user.username,
            from_location=None,
            to_location=None,
            remark=remark
        )
        db.add(lifecycle)
        await db.commit()

    # 获取机柜名称
    cabinet_name = None
    if asset.cabinet_id:
        cabinet_result = await db.execute(
            select(Cabinet).where(Cabinet.id == asset.cabinet_id)
        )
        cabinet = cabinet_result.scalar_one_or_none()
        if cabinet:
            cabinet_name = cabinet.cabinet_name

    asset_data = AssetResponse.model_validate(asset)
    asset_data.cabinet_name = cabinet_name
    asset_data.warranty_status = _calculate_warranty_status(asset.warranty_end)

    return asset_data


@router.delete("/assets/{asset_id}", summary="删除资产")
async def delete_asset(
    asset_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    删除资产及其关联记录
    """
    result = await db.execute(select(Asset).where(Asset.id == asset_id))
    asset = result.scalar_one_or_none()

    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    # 删除关联的生命周期记录
    await db.execute(
        select(AssetLifecycle).where(AssetLifecycle.asset_id == asset_id)
    )
    await db.execute(
        AssetLifecycle.__table__.delete().where(AssetLifecycle.asset_id == asset_id)
    )

    # 删除关联的维护记录
    await db.execute(
        MaintenanceRecord.__table__.delete().where(MaintenanceRecord.asset_id == asset_id)
    )

    # 删除关联的盘点明细
    await db.execute(
        AssetInventoryItem.__table__.delete().where(AssetInventoryItem.asset_id == asset_id)
    )

    await db.delete(asset)
    await db.commit()

    return {"message": "资产删除成功"}


@router.get("/assets/{asset_id}/lifecycle", response_model=List[LifecycleResponse], summary="获取资产生命周期记录")
async def get_asset_lifecycle(
    asset_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取指定资产的生命周期记录
    """
    # 检查资产是否存在
    asset_result = await db.execute(select(Asset).where(Asset.id == asset_id))
    if not asset_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="资产不存在")

    result = await db.execute(
        select(AssetLifecycle).where(
            AssetLifecycle.asset_id == asset_id
        ).order_by(AssetLifecycle.action_date.desc())
    )
    records = result.scalars().all()

    return [LifecycleResponse.model_validate(r) for r in records]


# ==================== 维护管理 ====================

@router.post("/maintenance", response_model=MaintenanceResponse, summary="创建维护记录")
async def create_maintenance(
    data: MaintenanceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """
    创建维护记录并更新资产状态
    """
    # 检查资产是否存在
    asset_result = await db.execute(select(Asset).where(Asset.id == data.asset_id))
    asset = asset_result.scalar_one_or_none()

    if not asset:
        raise HTTPException(status_code=404, detail="资产不存在")

    # 创建维护记录
    record = MaintenanceRecord(**data.model_dump())
    db.add(record)
    await db.commit()
    await db.refresh(record)

    # 更新资产状态为维护中
    asset.status = AssetStatus.maintenance
    asset.updated_at = datetime.now()
    await db.commit()

    # 添加生命周期记录
    lifecycle = AssetLifecycle(
        asset_id=asset.id,
        action="maintain",
        action_date=datetime.now(),
        operator=data.technician or current_user.username,
        from_location=None,
        to_location=None,
        remark=f"开始维护: {data.maintenance_type} - {data.description or ''}"
    )
    db.add(lifecycle)
    await db.commit()

    return MaintenanceResponse.model_validate(record)


@router.put("/maintenance/{record_id}/complete", response_model=MaintenanceResponse, summary="完成维护")
async def complete_maintenance(
    record_id: int,
    result: Optional[str] = Query(None, description="维护结果"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    完成维护并恢复资产状态
    """
    record_result = await db.execute(
        select(MaintenanceRecord).where(MaintenanceRecord.id == record_id)
    )
    record = record_result.scalar_one_or_none()

    if not record:
        raise HTTPException(status_code=404, detail="维护记录不存在")

    # 更新维护记录
    record.end_time = datetime.now()
    if result:
        record.result = result
    await db.commit()
    await db.refresh(record)

    # 恢复资产状态为使用中
    asset_result = await db.execute(select(Asset).where(Asset.id == record.asset_id))
    asset = asset_result.scalar_one_or_none()

    if asset:
        asset.status = AssetStatus.in_use
        asset.updated_at = datetime.now()
        await db.commit()

        # 添加生命周期记录
        lifecycle = AssetLifecycle(
            asset_id=asset.id,
            action="deploy",
            action_date=datetime.now(),
            operator=record.technician,
            from_location=None,
            to_location=None,
            remark=f"维护完成: {result or '正常'}"
        )
        db.add(lifecycle)
        await db.commit()

    return MaintenanceResponse.model_validate(record)


@router.get("/maintenance", response_model=List[MaintenanceResponse], summary="获取维护记录列表")
async def get_maintenance_records(
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回记录数"),
    asset_id: Optional[int] = Query(None, description="资产ID"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取维护记录列表（可按资产ID筛选）
    """
    query = select(MaintenanceRecord)

    if asset_id:
        query = query.where(MaintenanceRecord.asset_id == asset_id)

    query = query.order_by(MaintenanceRecord.start_time.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    records = result.scalars().all()

    return [MaintenanceResponse.model_validate(r) for r in records]


# ==================== 盘点管理 ====================

@router.post("/inventory", response_model=InventoryResponse, summary="创建资产盘点")
async def create_inventory(
    data: InventoryCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    创建资产盘点任务，自动生成盘点明细
    """
    # 检查盘点编码是否已存在
    existing = await db.execute(
        select(AssetInventory).where(AssetInventory.inventory_code == data.inventory_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="盘点编码已存在")

    # 创建盘点主记录
    inventory = AssetInventory(**data.model_dump())
    inventory.status = "pending"
    db.add(inventory)
    await db.commit()
    await db.refresh(inventory)

    # 获取所有在用和借出的资产，创建盘点明细
    assets_result = await db.execute(
        select(Asset).where(
            Asset.status.in_([AssetStatus.in_use, AssetStatus.borrowed])
        )
    )
    assets = assets_result.scalars().all()

    for asset in assets:
        expected_location = None
        if asset.cabinet_id:
            cabinet_result = await db.execute(
                select(Cabinet).where(Cabinet.id == asset.cabinet_id)
            )
            cabinet = cabinet_result.scalar_one_or_none()
            if cabinet:
                expected_location = f"{cabinet.cabinet_name} U{asset.u_position}" if asset.u_position else cabinet.cabinet_name

        item = AssetInventoryItem(
            inventory_id=inventory.id,
            asset_id=asset.id,
            expected_location=expected_location,
            is_matched=False
        )
        db.add(item)

    await db.commit()

    # 更新统计信息
    await _update_inventory_stats(db, inventory.id)
    await db.refresh(inventory)

    return InventoryResponse.model_validate(inventory)


@router.get("/inventory", response_model=List[InventoryResponse], summary="获取盘点列表")
async def get_inventory_list(
    skip: int = Query(0, ge=0, description="跳过记录数"),
    limit: int = Query(100, ge=1, le=1000, description="返回记录数"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取资产盘点列表
    """
    query = select(AssetInventory).order_by(
        AssetInventory.created_at.desc()
    ).offset(skip).limit(limit)

    result = await db.execute(query)
    inventories = result.scalars().all()

    return [InventoryResponse.model_validate(inv) for inv in inventories]


@router.get("/inventory/{inventory_id}/items", response_model=List[InventoryItemResponse], summary="获取盘点明细")
async def get_inventory_items(
    inventory_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取指定盘点任务的明细列表
    """
    # 检查盘点是否存在
    inventory_result = await db.execute(
        select(AssetInventory).where(AssetInventory.id == inventory_id)
    )
    if not inventory_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="盘点任务不存在")

    result = await db.execute(
        select(AssetInventoryItem).where(
            AssetInventoryItem.inventory_id == inventory_id
        )
    )
    items = result.scalars().all()

    return [InventoryItemResponse.model_validate(item) for item in items]


@router.put("/inventory/items/{item_id}", response_model=InventoryItemResponse, summary="更新盘点明细")
async def update_inventory_item(
    item_id: int,
    data: InventoryItemUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator)
):
    """
    更新盘点明细（录入盘点结果）
    """
    result = await db.execute(
        select(AssetInventoryItem).where(AssetInventoryItem.id == item_id)
    )
    item = result.scalar_one_or_none()

    if not item:
        raise HTTPException(status_code=404, detail="盘点明细不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        if value is not None:
            setattr(item, key, value)

    if not item.check_time:
        item.check_time = datetime.now()

    await db.commit()
    await db.refresh(item)

    # 更新盘点统计信息
    await _update_inventory_stats(db, item.inventory_id)

    return InventoryItemResponse.model_validate(item)


# ==================== 统计分析 ====================

@router.get("/statistics", response_model=AssetStatistics, summary="获取资产统计信息")
async def get_statistics(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取资产统计信息
    """
    # 资产总数
    total_result = await db.execute(select(func.count(Asset.id)))
    total_count = total_result.scalar() or 0

    # 按状态统计
    status_result = await db.execute(
        select(Asset.status, func.count(Asset.id)).group_by(Asset.status)
    )
    by_status = {
        status.value if status else "unknown": count
        for status, count in status_result.all()
    }

    # 按类型统计
    type_result = await db.execute(
        select(Asset.asset_type, func.count(Asset.id)).group_by(Asset.asset_type)
    )
    by_type = {
        asset_type.value if asset_type else "unknown": count
        for asset_type, count in type_result.all()
    }

    # 按部门统计
    dept_result = await db.execute(
        select(Asset.department, func.count(Asset.id)).where(
            Asset.department.isnot(None)
        ).group_by(Asset.department)
    )
    by_department = {
        dept or "未分配": count
        for dept, count in dept_result.all()
    }

    # 资产总价值
    value_result = await db.execute(select(func.sum(Asset.purchase_price)))
    total_value = value_result.scalar() or 0

    # 保修即将到期数量（30天内）
    expiring_date = date.today() + timedelta(days=30)
    expiring_result = await db.execute(
        select(func.count(Asset.id)).where(
            Asset.warranty_end.isnot(None),
            Asset.warranty_end <= expiring_date,
            Asset.warranty_end >= date.today(),
            Asset.status != AssetStatus.scrapped
        )
    )
    warranty_expiring_count = expiring_result.scalar() or 0

    return AssetStatistics(
        total_count=total_count,
        by_status=by_status,
        by_type=by_type,
        by_department=by_department,
        total_value=float(total_value),
        warranty_expiring_count=warranty_expiring_count
    )


@router.get("/warranty-alerts", response_model=WarrantyAlertResponse, summary="获取保修预警汇总")
async def get_warranty_alerts(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """返回 30/60/90 天三个阈值的过保预警资产列表"""
    today = date.today()
    result = await db.execute(
        select(Asset).where(
            Asset.warranty_end.isnot(None),
            Asset.warranty_end >= today,
            Asset.warranty_end <= today + timedelta(days=90),
            Asset.status != AssetStatus.scrapped
        ).order_by(Asset.warranty_end.asc())
    )
    assets = result.scalars().all()

    alerts_30, alerts_60, alerts_90 = [], [], []
    for asset in assets:
        days_remaining = (asset.warranty_end - today).days
        item = WarrantyAlertItem(
            asset_id=asset.id,
            asset_code=asset.asset_code,
            asset_name=asset.asset_name,
            asset_type=asset.asset_type.value if asset.asset_type else None,
            warranty_end=asset.warranty_end.isoformat(),
            days_remaining=days_remaining,
            status=asset.status.value if asset.status else None,
        )
        if days_remaining <= 30:
            alerts_30.append(item)
        elif days_remaining <= 60:
            alerts_60.append(item)
        else:
            alerts_90.append(item)

    return WarrantyAlertResponse(
        within_30_days=alerts_30,
        within_60_days=alerts_60,
        within_90_days=alerts_90,
        total_count=len(assets),
    )


@router.get("/warranty-expiring", response_model=List[AssetResponse], summary="获取即将过保资产")
async def get_warranty_expiring_assets(
    days: int = Query(30, ge=1, le=365, description="天数范围"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer)
):
    """
    获取保修即将到期的资产
    """
    expiring_date = date.today() + timedelta(days=days)

    result = await db.execute(
        select(Asset).where(
            Asset.warranty_end.isnot(None),
            Asset.warranty_end <= expiring_date,
            Asset.warranty_end >= date.today(),
            Asset.status != AssetStatus.scrapped
        ).order_by(Asset.warranty_end.asc())
    )
    assets = result.scalars().all()

    asset_list = []
    for asset in assets:
        asset_data = AssetResponse.model_validate(asset)

        # 获取机柜名称
        if asset.cabinet_id:
            cabinet_result = await db.execute(
                select(Cabinet).where(Cabinet.id == asset.cabinet_id)
            )
            cabinet = cabinet_result.scalar_one_or_none()
            if cabinet:
                asset_data.cabinet_name = cabinet.cabinet_name

        asset_data.warranty_status = _calculate_warranty_status(asset.warranty_end)
        asset_list.append(asset_data)

    return asset_list


# ==================== 辅助函数 ====================

def _calculate_warranty_status(warranty_end: Optional[date]) -> str:
    """
    计算保修状态

    Args:
        warranty_end: 保修结束日期

    Returns:
        保修状态: valid/expiring/expired/unknown
    """
    if not warranty_end:
        return "unknown"

    today = date.today()
    if warranty_end < today:
        return "expired"
    elif warranty_end <= today + timedelta(days=30):
        return "expiring"
    else:
        return "valid"


async def _update_inventory_stats(db: AsyncSession, inventory_id: int) -> None:
    """
    更新盘点统计信息

    Args:
        db: 数据库会话
        inventory_id: 盘点ID
    """
    inventory_result = await db.execute(
        select(AssetInventory).where(AssetInventory.id == inventory_id)
    )
    inventory = inventory_result.scalar_one_or_none()

    if not inventory:
        return

    items_result = await db.execute(
        select(AssetInventoryItem).where(
            AssetInventoryItem.inventory_id == inventory_id
        )
    )
    items = items_result.scalars().all()

    total_count = len(items)
    checked_count = sum(1 for item in items if item.check_time is not None)
    matched_count = sum(1 for item in items if item.is_matched)
    unmatched_count = sum(1 for item in items if item.check_time is not None and not item.is_matched)

    inventory.total_count = total_count
    inventory.checked_count = checked_count
    inventory.matched_count = matched_count
    inventory.unmatched_count = unmatched_count

    # 更新盘点状态
    if checked_count == 0:
        inventory.status = "pending"
    elif checked_count < total_count:
        inventory.status = "in_progress"
    else:
        inventory.status = "completed"
        inventory.completed_at = datetime.now()

    await db.commit()
