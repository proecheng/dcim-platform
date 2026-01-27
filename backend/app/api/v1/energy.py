"""
用电管理 API
"""
from typing import Optional, List, Dict, Any
from datetime import date, datetime, timedelta
from io import BytesIO
from fastapi import APIRouter, Depends, Query, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update, delete, and_, or_

from ..deps import get_db, get_current_user, require_admin, require_operator, require_viewer
from ...models.user import User
from ...models.energy import (
    PowerDevice, EnergyHourly, EnergyDaily, EnergyMonthly,
    ElectricityPricing, EnergySuggestion, PUEHistory,
    Transformer, MeterPoint, DistributionPanel, DistributionCircuit,
    PowerCurveData, DemandHistory, DeviceShiftConfig, DeviceLoadProfile
)
from ...models.point import PointRealtime
from ...models.device import Device
from ...schemas.common import ResponseModel, PageResponse
from ...services.energy_topology import topology_service
from ...services.power_device import power_device_service
from ...services.energy_analysis import demand_analysis_service, load_shift_analysis_service
from ...schemas.energy import (
    PowerDeviceCreate, PowerDeviceUpdate, PowerDeviceResponse, PowerDeviceTree,
    RealtimePowerData, RealtimePowerSummary,
    PUEData, PUETrend, PUEHistoryItem,
    EnergyStatQuery, EnergyStat, EnergyTrend, EnergyTrendItem, EnergyComparison,
    EnergyDailyData, EnergyMonthlyData,
    ElectricityPricingCreate, ElectricityPricingUpdate, ElectricityPricingResponse,
    EnergySuggestionResponse, AcceptSuggestion, CompleteSuggestion, RejectSuggestion, SavingPotential,
    DistributionDiagram, DistributionNode,
    # New imports for meter points and topology
    TransformerCreate, TransformerUpdate, TransformerResponse,
    MeterPointCreate, MeterPointUpdate, MeterPointResponse, MeterPointDetail,
    DistributionPanelCreate, DistributionPanelUpdate, DistributionPanelResponse,
    DistributionCircuitCreate, DistributionCircuitUpdate, DistributionCircuitResponse,
    PowerCurvePoint, PowerCurveResponse,
    DemandHistoryItem, DemandHistoryResponse,
    DeviceShiftPotential, DeviceShiftAnalysisResult,
    DemandConfigAnalysisItem, DemandConfigAnalysisResult,
    TopologyCircuitNode, TopologyPanelNode, TopologyMeterNode, TopologyTransformerNode,
    DistributionTopologyResponse
)

router = APIRouter()


# ==================== 用电设备管理 ====================

@router.get("/devices", response_model=ResponseModel[List[PowerDeviceResponse]], summary="获取用电设备列表")
async def get_power_devices(
    device_type: Optional[str] = Query(None, description="设备类型"),
    area_code: Optional[str] = Query(None, description="区域代码"),
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取用电设备列表"""
    query = select(PowerDevice)

    if device_type:
        query = query.where(PowerDevice.device_type == device_type)
    if area_code:
        query = query.where(PowerDevice.area_code == area_code)
    if is_enabled is not None:
        query = query.where(PowerDevice.is_enabled == is_enabled)

    query = query.order_by(PowerDevice.device_type, PowerDevice.device_code)
    result = await db.execute(query)
    devices = result.scalars().all()

    return ResponseModel(data=[PowerDeviceResponse.model_validate(d) for d in devices])


@router.get("/devices/tree", response_model=ResponseModel[List[PowerDeviceTree]], summary="获取用电设备树")
async def get_power_device_tree(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取用电设备树结构（配电层级）"""
    result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True).order_by(PowerDevice.device_type)
    )
    devices = result.scalars().all()

    # 构建设备字典
    device_dict = {d.id: PowerDeviceTree.model_validate(d) for d in devices}

    # 构建树结构
    root_devices = []
    for device in devices:
        device_tree = device_dict[device.id]
        if device.parent_device_id and device.parent_device_id in device_dict:
            device_dict[device.parent_device_id].children.append(device_tree)
        else:
            root_devices.append(device_tree)

    return ResponseModel(data=root_devices)


@router.post("/devices", response_model=ResponseModel[PowerDeviceResponse], summary="创建用电设备")
async def create_power_device(
    device: PowerDeviceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """
    创建用电设备

    创建设备后自动生成:
    - 设备负荷转移配置 (device_shift_configs)
    - 设备参数调节配置 (load_regulation_configs)
    """
    # 检查设备编码是否已存在
    existing = await db.execute(
        select(PowerDevice).where(PowerDevice.device_code == device.device_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="设备编码已存在")

    # 如果有上级设备，检查是否存在
    if device.parent_device_id:
        parent = await db.execute(
            select(PowerDevice).where(PowerDevice.id == device.parent_device_id)
        )
        if not parent.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="上级设备不存在")

    new_device = PowerDevice(**device.model_dump())
    db.add(new_device)
    await db.commit()
    await db.refresh(new_device)

    # 自动生成设备配置 (转移配置和调节配置)
    try:
        from ...services.device_config_generator import DeviceConfigAutoGenerator

        generator = DeviceConfigAutoGenerator(db)
        await generator.generate_configs_for_device(new_device, force=False)
        await db.commit()
    except Exception as e:
        # 配置生成失败不影响设备创建
        print(f"[警告] 设备配置自动生成失败: {e}")

    return ResponseModel(data=PowerDeviceResponse.model_validate(new_device))


# 注意: 静态路由必须在动态路由 /devices/{device_id} 之前定义
@router.get("/devices/shiftable", response_model=ResponseModel, summary="获取可转移负荷设备列表")
async def get_shiftable_devices(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取所有配置了负荷转移能力的设备

    返回数据来源：device_shift_configs表 + power_devices表
    """
    from ...services.device_regulation_service import DeviceRegulationService

    device_service = DeviceRegulationService(db)
    devices = await device_service.get_shiftable_devices()
    data_source = await device_service.get_regulation_data_source()

    return ResponseModel(data={
        "devices": devices,
        "total_count": len(devices),
        "total_shiftable_power": sum(d["shiftable_power"] for d in devices),
        "data_source": data_source
    })


@router.get("/devices/adjustable", response_model=ResponseModel, summary="获取可调节参数设备列表")
async def get_adjustable_devices(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取所有配置了调节能力的设备（温度/频率/亮度等）

    返回数据来源：load_regulation_configs表 + power_devices表
    """
    from ...services.device_regulation_service import DeviceRegulationService

    device_service = DeviceRegulationService(db)
    devices = await device_service.get_adjustable_devices()
    data_source = await device_service.get_regulation_data_source()

    return ResponseModel(data={
        "devices": devices,
        "total_count": len(devices),
        "by_regulation_type": {
            reg_type: len([d for d in devices if d["regulation_type"] == reg_type])
            for reg_type in set(d["regulation_type"] for d in devices)
        },
        "data_source": data_source
    })


@router.post("/devices/generate-configs", response_model=ResponseModel, summary="批量生成设备配置")
async def batch_generate_device_configs(
    device_ids: Optional[List[int]] = None,
    force: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """
    批量为设备生成转移配置和调节配置

    - **device_ids**: 设备ID列表，不提供则为所有设备生成
    - **force**: 是否强制覆盖已有配置（默认false，跳过已有配置的设备）
    """
    from ...services.device_config_generator import DeviceConfigAutoGenerator

    if device_ids is None:
        result = await db.execute(
            select(PowerDevice.id).where(PowerDevice.is_enabled == True)
        )
        device_ids = [row[0] for row in result.all()]

    if not device_ids:
        return ResponseModel(data={
            "message": "没有找到需要生成配置的设备",
            "total_devices": 0,
            "shift_configs_created": 0,
            "regulation_configs_created": 0
        })

    generator = DeviceConfigAutoGenerator(db)
    result = await generator.batch_generate_configs(device_ids, force)

    return ResponseModel(data={
        "message": f"已为 {result['total_devices']} 个设备生成配置",
        **result
    })


@router.get("/devices/{device_id}", response_model=ResponseModel[PowerDeviceResponse], summary="获取用电设备详情")
async def get_power_device(
    device_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取用电设备详情"""
    result = await db.execute(select(PowerDevice).where(PowerDevice.id == device_id))
    device = result.scalar_one_or_none()

    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    return ResponseModel(data=PowerDeviceResponse.model_validate(device))


@router.put("/devices/{device_id}", response_model=ResponseModel[PowerDeviceResponse], summary="更新用电设备")
async def update_power_device(
    device_id: int,
    device: PowerDeviceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新用电设备"""
    result = await db.execute(select(PowerDevice).where(PowerDevice.id == device_id))
    existing = result.scalar_one_or_none()

    if not existing:
        raise HTTPException(status_code=404, detail="设备不存在")

    update_data = device.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()

    await db.execute(
        update(PowerDevice).where(PowerDevice.id == device_id).values(**update_data)
    )
    await db.commit()

    result = await db.execute(select(PowerDevice).where(PowerDevice.id == device_id))
    updated_device = result.scalar_one()

    return ResponseModel(data=PowerDeviceResponse.model_validate(updated_device))


@router.delete("/devices/{device_id}", response_model=ResponseModel, summary="删除用电设备")
async def delete_power_device(
    device_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除用电设备"""
    result = await db.execute(select(PowerDevice).where(PowerDevice.id == device_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="设备不存在")

    # 检查是否有下级设备
    children = await db.execute(
        select(PowerDevice).where(PowerDevice.parent_device_id == device_id)
    )
    if children.scalars().first():
        raise HTTPException(status_code=400, detail="该设备有下级设备，无法删除")

    await db.execute(delete(PowerDevice).where(PowerDevice.id == device_id))
    await db.commit()

    return ResponseModel(message="删除成功")


# ==================== 实时电力数据 ====================

@router.get("/realtime", response_model=ResponseModel[List[RealtimePowerData]], summary="获取实时电力数据")
async def get_realtime_power(
    device_type: Optional[str] = Query(None, description="设备类型"),
    area_code: Optional[str] = Query(None, description="区域代码"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取所有用电设备的实时电力数据"""
    query = select(PowerDevice).where(PowerDevice.is_enabled == True)

    if device_type:
        query = query.where(PowerDevice.device_type == device_type)
    if area_code:
        query = query.where(PowerDevice.area_code == area_code)

    result = await db.execute(query)
    devices = result.scalars().all()

    power_data = []
    for device in devices:
        # 从 PointRealtime 获取实时数据（如果有对应的电力点位）
        # 这里使用模拟数据作为示例
        import random
        base_power = device.rated_power or 10.0

        data = RealtimePowerData(
            device_id=device.id,
            device_code=device.device_code,
            device_name=device.device_name,
            device_type=device.device_type,
            voltage_a=380 + random.uniform(-5, 5),
            voltage_b=380 + random.uniform(-5, 5) if device.phase_type == "3P" else None,
            voltage_c=380 + random.uniform(-5, 5) if device.phase_type == "3P" else None,
            current_a=base_power * 1.52 * random.uniform(0.5, 0.9),
            current_b=base_power * 1.52 * random.uniform(0.5, 0.9) if device.phase_type == "3P" else None,
            current_c=base_power * 1.52 * random.uniform(0.5, 0.9) if device.phase_type == "3P" else None,
            active_power=base_power * random.uniform(0.5, 0.9),
            reactive_power=base_power * 0.2 * random.uniform(0.5, 1.0),
            apparent_power=base_power * 1.1 * random.uniform(0.5, 0.9),
            power_factor=random.uniform(0.85, 0.98),
            frequency=50 + random.uniform(-0.1, 0.1),
            total_energy=base_power * 24 * 30 * random.uniform(0.5, 1.0),
            load_rate=random.uniform(0.4, 0.85) if device.rated_power else None,
            status="normal",
            update_time=datetime.now()
        )
        power_data.append(data)

    return ResponseModel(data=power_data)


@router.get("/realtime/summary", response_model=ResponseModel[RealtimePowerSummary], summary="获取电力汇总")
async def get_power_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取实时电力汇总（总功率、IT功率、制冷功率、PUE等）"""
    result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True)
    )
    devices = result.scalars().all()

    import random
    total_power = 0.0
    it_power = 0.0
    cooling_power = 0.0
    ups_power = 0.0
    other_power = 0.0

    for device in devices:
        base_power = (device.rated_power or 10.0) * random.uniform(0.5, 0.9)

        if device.device_type == "IT":
            it_power += base_power
        elif device.device_type == "AC":
            cooling_power += base_power
        elif device.device_type == "UPS":
            ups_power += base_power
        elif device.device_type == "MAIN":
            total_power = base_power * 10  # 主进线
        else:
            other_power += base_power

    if total_power == 0:
        total_power = it_power + cooling_power + ups_power + other_power

    current_pue = total_power / it_power if it_power > 0 else 1.0

    # 获取今日能耗
    today = date.today()
    today_result = await db.execute(
        select(func.sum(EnergyDaily.total_energy), func.sum(EnergyDaily.energy_cost))
        .where(EnergyDaily.stat_date == today)
    )
    today_stats = today_result.first()
    today_energy = today_stats[0] or total_power * datetime.now().hour
    today_cost = today_stats[1] or today_energy * 0.8

    # 获取本月能耗
    month_result = await db.execute(
        select(func.sum(EnergyDaily.total_energy), func.sum(EnergyDaily.energy_cost))
        .where(
            EnergyDaily.stat_date >= date(today.year, today.month, 1),
            EnergyDaily.stat_date <= today
        )
    )
    month_stats = month_result.first()
    month_energy = month_stats[0] or today_energy * today.day
    month_cost = month_stats[1] or month_energy * 0.8

    summary = RealtimePowerSummary(
        total_power=round(total_power, 2),
        it_power=round(it_power, 2),
        cooling_power=round(cooling_power, 2),
        ups_power=round(ups_power, 2),
        other_power=round(other_power, 2),
        current_pue=round(current_pue, 2),
        today_energy=round(today_energy, 2),
        today_cost=round(today_cost, 2),
        month_energy=round(month_energy, 2),
        month_cost=round(month_cost, 2)
    )

    return ResponseModel(data=summary)


@router.get("/realtime/{device_id}", response_model=ResponseModel[RealtimePowerData], summary="获取设备实时电力")
async def get_device_realtime_power(
    device_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取指定设备的实时电力数据"""
    result = await db.execute(select(PowerDevice).where(PowerDevice.id == device_id))
    device = result.scalar_one_or_none()

    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")

    import random
    base_power = device.rated_power or 10.0

    data = RealtimePowerData(
        device_id=device.id,
        device_code=device.device_code,
        device_name=device.device_name,
        device_type=device.device_type,
        voltage_a=380 + random.uniform(-5, 5),
        voltage_b=380 + random.uniform(-5, 5) if device.phase_type == "3P" else None,
        voltage_c=380 + random.uniform(-5, 5) if device.phase_type == "3P" else None,
        current_a=base_power * 1.52 * random.uniform(0.5, 0.9),
        current_b=base_power * 1.52 * random.uniform(0.5, 0.9) if device.phase_type == "3P" else None,
        current_c=base_power * 1.52 * random.uniform(0.5, 0.9) if device.phase_type == "3P" else None,
        active_power=base_power * random.uniform(0.5, 0.9),
        reactive_power=base_power * 0.2 * random.uniform(0.5, 1.0),
        apparent_power=base_power * 1.1 * random.uniform(0.5, 0.9),
        power_factor=random.uniform(0.85, 0.98),
        frequency=50 + random.uniform(-0.1, 0.1),
        total_energy=base_power * 24 * 30 * random.uniform(0.5, 1.0),
        load_rate=random.uniform(0.4, 0.85) if device.rated_power else None,
        status="normal",
        update_time=datetime.now()
    )

    return ResponseModel(data=data)


# ==================== PUE ====================

@router.get("/pue", response_model=ResponseModel[PUEData], summary="获取当前PUE")
async def get_current_pue(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取当前PUE值及功率分布"""
    result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True)
    )
    devices = result.scalars().all()

    import random
    total_power = 0.0
    it_power = 0.0
    cooling_power = 0.0
    ups_loss = 0.0
    lighting_power = 0.0
    other_power = 0.0

    for device in devices:
        base_power = (device.rated_power or 10.0) * random.uniform(0.5, 0.9)

        if device.is_it_load or device.device_type == "IT":
            it_power += base_power
        elif device.device_type == "AC":
            cooling_power += base_power
        elif device.device_type == "UPS":
            ups_loss += base_power * 0.05  # UPS损耗约5%
        elif device.device_type == "MAIN":
            total_power = base_power * 10
        else:
            other_power += base_power

    # 照明估算
    lighting_power = total_power * 0.02 if total_power > 0 else 5.0

    if total_power == 0:
        total_power = it_power + cooling_power + ups_loss + lighting_power + other_power

    current_pue = total_power / it_power if it_power > 0 else 1.5

    pue_data = PUEData(
        current_pue=round(current_pue, 3),
        total_power=round(total_power, 2),
        it_power=round(it_power, 2),
        cooling_power=round(cooling_power, 2),
        ups_loss=round(ups_loss, 2),
        lighting_power=round(lighting_power, 2),
        other_power=round(other_power, 2),
        update_time=datetime.now()
    )

    return ResponseModel(data=pue_data)


@router.get("/pue/trend", response_model=ResponseModel[PUETrend], summary="获取PUE趋势")
async def get_pue_trend(
    period: str = Query("day", description="时间段: hour/day/week/month"),
    start_time: Optional[datetime] = Query(None, description="开始时间"),
    end_time: Optional[datetime] = Query(None, description="结束时间"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取PUE历史趋势"""
    now = datetime.now()

    # 根据周期确定时间范围
    if period == "hour":
        if not start_time:
            start_time = now - timedelta(hours=24)
        if not end_time:
            end_time = now
    elif period == "day":
        if not start_time:
            start_time = now - timedelta(days=7)
        if not end_time:
            end_time = now
    elif period == "week":
        if not start_time:
            start_time = now - timedelta(weeks=4)
        if not end_time:
            end_time = now
    else:  # month
        if not start_time:
            start_time = now - timedelta(days=365)
        if not end_time:
            end_time = now

    result = await db.execute(
        select(PUEHistory)
        .where(PUEHistory.record_time >= start_time, PUEHistory.record_time <= end_time)
        .order_by(PUEHistory.record_time)
    )
    history = result.scalars().all()

    # 如果没有历史数据，生成模拟数据
    if not history:
        import random
        data_list = []
        current = start_time
        while current <= end_time:
            pue = 1.4 + random.uniform(-0.2, 0.3)
            total = 500 + random.uniform(-50, 50)
            it = total / pue
            data_list.append(PUEHistoryItem(
                record_time=current,
                pue=round(pue, 3),
                total_power=round(total, 2),
                it_power=round(it, 2)
            ))
            if period == "hour":
                current += timedelta(hours=1)
            elif period == "day":
                current += timedelta(days=1)
            elif period == "week":
                current += timedelta(days=7)
            else:
                current += timedelta(days=30)
    else:
        data_list = [PUEHistoryItem.model_validate(h) for h in history]

    pue_values = [d.pue for d in data_list]
    avg_pue = sum(pue_values) / len(pue_values) if pue_values else 0
    min_pue = min(pue_values) if pue_values else 0
    max_pue = max(pue_values) if pue_values else 0

    trend = PUETrend(
        period=period,
        data=data_list,
        avg_pue=round(avg_pue, 3),
        min_pue=round(min_pue, 3),
        max_pue=round(max_pue, 3)
    )

    return ResponseModel(data=trend)


# ==================== 能耗统计 ====================

@router.get("/statistics/daily", response_model=ResponseModel[List[EnergyDailyData]], summary="获取日能耗统计")
async def get_daily_statistics(
    device_id: Optional[int] = Query(None, description="设备ID"),
    device_type: Optional[str] = Query(None, description="设备类型"),
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取日能耗统计数据"""
    query = select(EnergyDaily).where(
        EnergyDaily.stat_date >= start_date,
        EnergyDaily.stat_date <= end_date
    )

    if device_id:
        query = query.where(EnergyDaily.device_id == device_id)

    query = query.order_by(EnergyDaily.stat_date.desc())
    result = await db.execute(query)
    daily_data = result.scalars().all()

    # 如果没有数据，生成模拟数据
    if not daily_data:
        import random
        data_list = []
        current_date = start_date
        while current_date <= end_date:
            total = 1000 + random.uniform(-200, 200)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            data = EnergyDailyData(
                id=0,
                device_id=device_id or 1,
                stat_date=current_date,
                total_energy=round(total, 2),
                peak_energy=round(peak, 2),
                normal_energy=round(normal, 2),
                valley_energy=round(valley, 2),
                max_power=round(total / 20, 2),
                avg_power=round(total / 24, 2),
                max_power_time=datetime.combine(current_date, datetime.min.time()) + timedelta(hours=14),
                energy_cost=round(peak * 1.2 + normal * 0.8 + valley * 0.4, 2),
                pue=round(1.4 + random.uniform(-0.1, 0.2), 2)
            )
            data_list.append(data)
            current_date += timedelta(days=1)
        return ResponseModel(data=data_list)

    return ResponseModel(data=[EnergyDailyData.model_validate(d) for d in daily_data])


@router.get("/statistics/monthly", response_model=ResponseModel[List[EnergyMonthlyData]], summary="获取月能耗统计")
async def get_monthly_statistics(
    device_id: Optional[int] = Query(None, description="设备ID"),
    device_type: Optional[str] = Query(None, description="设备类型"),
    year: int = Query(..., description="年份"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取月能耗统计数据"""
    query = select(EnergyMonthly).where(EnergyMonthly.stat_year == year)

    if device_id:
        query = query.where(EnergyMonthly.device_id == device_id)

    query = query.order_by(EnergyMonthly.stat_month)
    result = await db.execute(query)
    monthly_data = result.scalars().all()

    # 如果没有数据，生成模拟数据
    if not monthly_data:
        import random
        data_list = []
        for month in range(1, 13):
            total = 30000 + random.uniform(-5000, 5000)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            peak_cost = peak * 1.2
            normal_cost = normal * 0.8
            valley_cost = valley * 0.4
            data = EnergyMonthlyData(
                id=0,
                device_id=device_id or 1,
                stat_year=year,
                stat_month=month,
                total_energy=round(total, 2),
                peak_energy=round(peak, 2),
                normal_energy=round(normal, 2),
                valley_energy=round(valley, 2),
                max_power=round(total / 500, 2),
                avg_power=round(total / 720, 2),
                max_power_date=date(year, month, 15),
                energy_cost=round(peak_cost + normal_cost + valley_cost, 2),
                peak_cost=round(peak_cost, 2),
                normal_cost=round(normal_cost, 2),
                valley_cost=round(valley_cost, 2),
                avg_pue=round(1.45 + random.uniform(-0.1, 0.15), 2)
            )
            data_list.append(data)
        return ResponseModel(data=data_list)

    return ResponseModel(data=[EnergyMonthlyData.model_validate(d) for d in monthly_data])


@router.get("/statistics/summary", response_model=ResponseModel[EnergyStat], summary="获取能耗汇总")
async def get_energy_summary(
    device_id: Optional[int] = Query(None, description="设备ID"),
    device_type: Optional[str] = Query(None, description="设备类型"),
    area_code: Optional[str] = Query(None, description="区域代码"),
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取能耗汇总统计"""
    query = select(
        func.sum(EnergyDaily.total_energy),
        func.sum(EnergyDaily.peak_energy),
        func.sum(EnergyDaily.normal_energy),
        func.sum(EnergyDaily.valley_energy),
        func.sum(EnergyDaily.energy_cost),
        func.avg(EnergyDaily.avg_power),
        func.max(EnergyDaily.max_power),
        func.avg(EnergyDaily.pue)
    ).where(
        EnergyDaily.stat_date >= start_date,
        EnergyDaily.stat_date <= end_date
    )

    if device_id:
        query = query.where(EnergyDaily.device_id == device_id)

    result = await db.execute(query)
    stats = result.first()

    # 如果没有数据，返回模拟数据
    days = (end_date - start_date).days + 1
    if stats[0] is None:
        import random
        total = 1000 * days + random.uniform(-500, 500)
        peak = total * 0.4
        normal = total * 0.35
        valley = total * 0.25
        peak_cost = peak * 1.2
        normal_cost = normal * 0.8
        valley_cost = valley * 0.4

        summary = EnergyStat(
            total_energy=round(total, 2),
            peak_energy=round(peak, 2),
            normal_energy=round(normal, 2),
            valley_energy=round(valley, 2),
            total_cost=round(peak_cost + normal_cost + valley_cost, 2),
            peak_cost=round(peak_cost, 2),
            normal_cost=round(normal_cost, 2),
            valley_cost=round(valley_cost, 2),
            avg_power=round(total / (days * 24), 2),
            max_power=round(total / (days * 20), 2),
            avg_pue=1.45
        )
    else:
        total = stats[0] or 0
        peak = stats[1] or 0
        normal = stats[2] or 0
        valley = stats[3] or 0
        total_cost = stats[4] or 0

        summary = EnergyStat(
            total_energy=round(total, 2),
            peak_energy=round(peak, 2),
            normal_energy=round(normal, 2),
            valley_energy=round(valley, 2),
            total_cost=round(total_cost, 2),
            peak_cost=round(peak * 1.2, 2),
            normal_cost=round(normal * 0.8, 2),
            valley_cost=round(valley * 0.4, 2),
            avg_power=round(stats[5] or 0, 2),
            max_power=round(stats[6] or 0, 2),
            avg_pue=round(stats[7] or 1.5, 2)
        )

    return ResponseModel(data=summary)


@router.get("/statistics/trend", response_model=ResponseModel[EnergyTrend], summary="获取能耗趋势")
async def get_energy_trend(
    device_id: Optional[int] = Query(None, description="设备ID"),
    device_type: Optional[str] = Query(None, description="设备类型"),
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    granularity: str = Query("daily", description="粒度: hourly/daily/monthly"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取能耗趋势数据"""
    import random
    data_list = []
    total_energy = 0.0
    total_cost = 0.0

    if granularity == "daily":
        current = start_date
        while current <= end_date:
            energy = 1000 + random.uniform(-200, 200)
            cost = energy * 0.8
            total_energy += energy
            total_cost += cost
            data_list.append(EnergyTrendItem(
                time_label=current.strftime("%Y-%m-%d"),
                energy=round(energy, 2),
                cost=round(cost, 2),
                power=round(energy / 24, 2)
            ))
            current += timedelta(days=1)
    elif granularity == "monthly":
        year = start_date.year
        for month in range(1, 13):
            energy = 30000 + random.uniform(-5000, 5000)
            cost = energy * 0.8
            total_energy += energy
            total_cost += cost
            data_list.append(EnergyTrendItem(
                time_label=f"{year}-{month:02d}",
                energy=round(energy, 2),
                cost=round(cost, 2),
                power=round(energy / 720, 2)
            ))
    else:  # hourly
        current = datetime.combine(start_date, datetime.min.time())
        end = datetime.combine(end_date, datetime.max.time())
        while current <= end:
            energy = 40 + random.uniform(-10, 10)
            cost = energy * 0.8
            total_energy += energy
            total_cost += cost
            data_list.append(EnergyTrendItem(
                time_label=current.strftime("%Y-%m-%d %H:00"),
                energy=round(energy, 2),
                cost=round(cost, 2),
                power=round(energy, 2)
            ))
            current += timedelta(hours=1)
            if len(data_list) > 168:  # 最多一周的小时数据
                break

    trend = EnergyTrend(
        granularity=granularity,
        data=data_list,
        total_energy=round(total_energy, 2),
        total_cost=round(total_cost, 2)
    )

    return ResponseModel(data=trend)


@router.get("/statistics/comparison", response_model=ResponseModel[EnergyComparison], summary="获取能耗对比")
async def get_energy_comparison(
    device_id: Optional[int] = Query(None, description="设备ID"),
    comparison_type: str = Query("mom", description="对比类型: mom(环比)/yoy(同比)"),
    period: str = Query("month", description="对比周期: day/week/month"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取能耗环比/同比对比"""
    import random
    today = date.today()

    # 生成模拟数据
    current_total = 30000 + random.uniform(-3000, 3000)
    current_peak = current_total * 0.4
    current_normal = current_total * 0.35
    current_valley = current_total * 0.25
    current_cost = current_peak * 1.2 + current_normal * 0.8 + current_valley * 0.4

    if comparison_type == "yoy":  # 同比
        change_rate = random.uniform(-0.15, 0.15)
    else:  # 环比
        change_rate = random.uniform(-0.1, 0.1)

    prev_total = current_total / (1 + change_rate)
    prev_peak = prev_total * 0.4
    prev_normal = prev_total * 0.35
    prev_valley = prev_total * 0.25
    prev_cost = prev_peak * 1.2 + prev_normal * 0.8 + prev_valley * 0.4

    current_stat = EnergyStat(
        total_energy=round(current_total, 2),
        peak_energy=round(current_peak, 2),
        normal_energy=round(current_normal, 2),
        valley_energy=round(current_valley, 2),
        total_cost=round(current_cost, 2),
        peak_cost=round(current_peak * 1.2, 2),
        normal_cost=round(current_normal * 0.8, 2),
        valley_cost=round(current_valley * 0.4, 2),
        avg_power=round(current_total / 720, 2),
        max_power=round(current_total / 500, 2),
        avg_pue=1.45
    )

    prev_stat = EnergyStat(
        total_energy=round(prev_total, 2),
        peak_energy=round(prev_peak, 2),
        normal_energy=round(prev_normal, 2),
        valley_energy=round(prev_valley, 2),
        total_cost=round(prev_cost, 2),
        peak_cost=round(prev_peak * 1.2, 2),
        normal_cost=round(prev_normal * 0.8, 2),
        valley_cost=round(prev_valley * 0.4, 2),
        avg_power=round(prev_total / 720, 2),
        max_power=round(prev_total / 500, 2),
        avg_pue=1.48
    )

    comparison = EnergyComparison(
        current_period=current_stat,
        previous_period=prev_stat,
        energy_change=round(current_total - prev_total, 2),
        energy_change_rate=round(change_rate, 4),
        cost_change=round(current_cost - prev_cost, 2),
        cost_change_rate=round((current_cost - prev_cost) / prev_cost if prev_cost > 0 else 0, 4)
    )

    return ResponseModel(data=comparison)


# ==================== 电费分析 ====================

@router.get("/cost/daily", response_model=ResponseModel, summary="获取日电费统计")
async def get_daily_cost(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取日电费统计（峰谷平分时电费）"""
    result = await db.execute(
        select(EnergyDaily)
        .where(EnergyDaily.stat_date >= start_date, EnergyDaily.stat_date <= end_date)
        .order_by(EnergyDaily.stat_date)
    )
    daily_data = result.scalars().all()

    import random
    if not daily_data:
        # 生成模拟数据
        data_list = []
        current = start_date
        while current <= end_date:
            total = 1000 + random.uniform(-200, 200)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            data_list.append({
                "date": current.isoformat(),
                "total_energy": round(total, 2),
                "peak_energy": round(peak, 2),
                "normal_energy": round(normal, 2),
                "valley_energy": round(valley, 2),
                "peak_cost": round(peak * 1.2, 2),
                "normal_cost": round(normal * 0.8, 2),
                "valley_cost": round(valley * 0.4, 2),
                "total_cost": round(peak * 1.2 + normal * 0.8 + valley * 0.4, 2)
            })
            current += timedelta(days=1)
        return ResponseModel(data={"items": data_list})

    data_list = []
    for d in daily_data:
        data_list.append({
            "date": d.stat_date.isoformat(),
            "total_energy": d.total_energy,
            "peak_energy": d.peak_energy,
            "normal_energy": d.normal_energy,
            "valley_energy": d.valley_energy,
            "peak_cost": round(d.peak_energy * 1.2, 2),
            "normal_cost": round(d.normal_energy * 0.8, 2),
            "valley_cost": round(d.valley_energy * 0.4, 2),
            "total_cost": d.energy_cost
        })

    return ResponseModel(data={"items": data_list})


@router.get("/cost/monthly", response_model=ResponseModel, summary="获取月电费统计")
async def get_monthly_cost(
    year: int = Query(..., description="年份"),
    month: Optional[int] = Query(None, description="月份"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取月电费统计"""
    query = select(EnergyMonthly).where(EnergyMonthly.stat_year == year)
    if month:
        query = query.where(EnergyMonthly.stat_month == month)
    query = query.order_by(EnergyMonthly.stat_month)

    result = await db.execute(query)
    monthly_data = result.scalars().all()

    import random
    if not monthly_data:
        # 生成模拟数据
        data_list = []
        months = [month] if month else range(1, 13)
        for m in months:
            total = 30000 + random.uniform(-5000, 5000)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            data_list.append({
                "year": year,
                "month": m,
                "total_energy": round(total, 2),
                "peak_energy": round(peak, 2),
                "normal_energy": round(normal, 2),
                "valley_energy": round(valley, 2),
                "peak_cost": round(peak * 1.2, 2),
                "normal_cost": round(normal * 0.8, 2),
                "valley_cost": round(valley * 0.4, 2),
                "total_cost": round(peak * 1.2 + normal * 0.8 + valley * 0.4, 2)
            })
        return ResponseModel(data={"items": data_list})

    data_list = []
    for d in monthly_data:
        data_list.append({
            "year": d.stat_year,
            "month": d.stat_month,
            "total_energy": d.total_energy,
            "peak_energy": d.peak_energy,
            "normal_energy": d.normal_energy,
            "valley_energy": d.valley_energy,
            "peak_cost": d.peak_cost,
            "normal_cost": d.normal_cost,
            "valley_cost": d.valley_cost,
            "total_cost": d.energy_cost
        })

    return ResponseModel(data={"items": data_list})


# ==================== 电价配置 ====================

@router.get("/pricing", response_model=ResponseModel[List[ElectricityPricingResponse]], summary="获取电价配置")
async def get_pricing_list(
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取电价配置列表"""
    query = select(ElectricityPricing)
    if is_enabled is not None:
        query = query.where(ElectricityPricing.is_enabled == is_enabled)
    query = query.order_by(ElectricityPricing.period_type, ElectricityPricing.start_time)

    result = await db.execute(query)
    pricing_list = result.scalars().all()

    return ResponseModel(data=[ElectricityPricingResponse.model_validate(p) for p in pricing_list])


@router.post("/pricing", response_model=ResponseModel[ElectricityPricingResponse], summary="创建电价配置")
async def create_pricing(
    pricing: ElectricityPricingCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """创建电价配置"""
    new_pricing = ElectricityPricing(**pricing.model_dump())
    db.add(new_pricing)
    await db.commit()
    await db.refresh(new_pricing)

    return ResponseModel(data=ElectricityPricingResponse.model_validate(new_pricing))


@router.put("/pricing/{pricing_id}", response_model=ResponseModel[ElectricityPricingResponse], summary="更新电价配置")
async def update_pricing(
    pricing_id: int,
    pricing: ElectricityPricingUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新电价配置"""
    result = await db.execute(select(ElectricityPricing).where(ElectricityPricing.id == pricing_id))
    existing = result.scalar_one_or_none()

    if not existing:
        raise HTTPException(status_code=404, detail="电价配置不存在")

    update_data = pricing.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()

    await db.execute(
        update(ElectricityPricing).where(ElectricityPricing.id == pricing_id).values(**update_data)
    )
    await db.commit()

    result = await db.execute(select(ElectricityPricing).where(ElectricityPricing.id == pricing_id))
    updated = result.scalar_one()

    return ResponseModel(data=ElectricityPricingResponse.model_validate(updated))


@router.delete("/pricing/{pricing_id}", response_model=ResponseModel, summary="删除电价配置")
async def delete_pricing(
    pricing_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除电价配置"""
    result = await db.execute(select(ElectricityPricing).where(ElectricityPricing.id == pricing_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="电价配置不存在")

    await db.execute(delete(ElectricityPricing).where(ElectricityPricing.id == pricing_id))
    await db.commit()

    return ResponseModel(message="删除成功")


# ==================== 节能建议 ====================

@router.get("/suggestions", response_model=ResponseModel[List[EnergySuggestionResponse]], summary="获取节能建议")
async def get_suggestions(
    status: Optional[str] = Query(None, description="状态: pending/accepted/rejected/completed"),
    priority: Optional[str] = Query(None, description="优先级: high/medium/low"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取节能建议列表"""
    query = select(EnergySuggestion)

    if status:
        query = query.where(EnergySuggestion.status == status)
    if priority:
        query = query.where(EnergySuggestion.priority == priority)

    query = query.order_by(
        EnergySuggestion.priority.desc(),
        EnergySuggestion.created_at.desc()
    )
    query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    suggestions = result.scalars().all()

    return ResponseModel(data=[EnergySuggestionResponse.model_validate(s) for s in suggestions])


# ==================== V2.3 节能建议引擎（必须在 /suggestions/{suggestion_id} 之前定义）====================

@router.get("/suggestions/templates", response_model=ResponseModel, summary="获取建议模板列表")
async def get_suggestion_templates(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取所有节能建议模板
    """
    from ...services.suggestion_engine import SuggestionEngineService

    service = SuggestionEngineService(db)
    templates = service.get_templates()

    return ResponseModel(data={
        "total": len(templates),
        "templates": templates
    })


@router.post("/suggestions/analyze", response_model=ResponseModel, summary="触发建议分析")
async def trigger_suggestion_analysis(
    categories: Optional[List[str]] = Query(None, description="分析类别"),
    force_refresh: bool = Query(False, description="强制刷新已有建议"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    触发节能建议分析引擎

    分析当前数据并生成节能建议
    """
    from ...services.suggestion_engine import SuggestionEngineService

    service = SuggestionEngineService(db)
    result = await service.analyze_and_generate(categories, force_refresh)

    return ResponseModel(data=result)


@router.get("/suggestions/enhanced", response_model=ResponseModel, summary="获取增强建议列表")
async def get_enhanced_suggestions(
    category: Optional[str] = Query(None, description="建议类别"),
    priority: Optional[str] = Query(None, description="优先级"),
    status: str = Query("pending", description="状态"),
    limit: int = Query(50, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取增强版节能建议列表

    包含详细的问题描述、分析、实施步骤和预期效果
    """
    from ...services.suggestion_engine import SuggestionEngineService

    service = SuggestionEngineService(db)
    suggestions = await service.get_suggestions(category, priority, status, limit)

    result = []
    for s in suggestions:
        result.append({
            "id": s.id,
            "template_id": s.template_id,
            "category": s.category,
            "rule_name": s.rule_name,
            "suggestion": s.suggestion,
            "problem_description": s.problem_description,
            "analysis_detail": s.analysis_detail,
            "implementation_steps": s.implementation_steps,
            "expected_effect": s.expected_effect,
            "priority": s.priority,
            "difficulty": s.difficulty,
            "potential_saving": s.potential_saving,
            "potential_cost_saving": s.potential_cost_saving,
            "status": s.status,
            "created_at": s.created_at.isoformat() if s.created_at else None
        })

    return ResponseModel(data={
        "total": len(result),
        "suggestions": result
    })


@router.get("/suggestions/summary", response_model=ResponseModel, summary="获取建议汇总统计")
async def get_suggestions_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取节能建议汇总统计
    """
    from ...services.suggestion_engine import SuggestionEngineService

    service = SuggestionEngineService(db)
    summary = await service.get_suggestion_summary()

    return ResponseModel(data=summary)


@router.get("/suggestions/enhanced/{suggestion_id}", response_model=ResponseModel, summary="获取增强建议详情")
async def get_enhanced_suggestion_detail(
    suggestion_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取单个节能建议的详细信息（包含完整的调整参数和设备信息）
    """
    from ...services.suggestion_engine import SuggestionEngineService

    service = SuggestionEngineService(db)
    detail = await service.get_suggestion_detail(suggestion_id)

    if not detail:
        raise HTTPException(status_code=404, detail="建议不存在")

    return ResponseModel(data=detail)


# ==================== 节能建议详情路由（具体路径必须在参数路径之前）====================

@router.get("/suggestions/detail/{suggestion_id}", response_model=ResponseModel, summary="获取建议完整详情")
async def get_suggestion_full_detail(
    suggestion_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取建议的完整详情，包含：
    - 建议基本信息
    - 可调整参数（含实时设备和电价数据）
    - 计算公式和步骤
    - 涉及设备列表
    - 数据来源信息

    返回的数据来自：
    - electricity_pricing表（电价）
    - device_shift_configs表（设备转移能力）
    - power_devices表（设备基础信息）
    """
    from ...services.suggestion_engine import SuggestionEngineService

    engine = SuggestionEngineService(db)
    detail = await engine.get_suggestion_detail(suggestion_id)

    if not detail:
        raise HTTPException(status_code=404, detail="建议不存在")

    return ResponseModel(data=detail)


# ==================== 节能建议详情路由（带参数，必须在具体路径之后）====================

@router.get("/suggestions/{suggestion_id}", response_model=ResponseModel[EnergySuggestionResponse], summary="获取建议详情")
async def get_suggestion(
    suggestion_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取节能建议详情"""
    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    suggestion = result.scalar_one_or_none()

    if not suggestion:
        raise HTTPException(status_code=404, detail="建议不存在")

    return ResponseModel(data=EnergySuggestionResponse.model_validate(suggestion))


@router.put("/suggestions/{suggestion_id}/accept", response_model=ResponseModel[EnergySuggestionResponse], summary="接受建议")
async def accept_suggestion(
    suggestion_id: int,
    data: AcceptSuggestion,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """接受节能建议"""
    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    suggestion = result.scalar_one_or_none()

    if not suggestion:
        raise HTTPException(status_code=404, detail="建议不存在")

    if suggestion.status != "pending":
        raise HTTPException(status_code=400, detail="只能接受待处理的建议")

    await db.execute(
        update(EnergySuggestion).where(EnergySuggestion.id == suggestion_id).values(
            status="accepted",
            accepted_by=current_user.id,
            accepted_at=datetime.now(),
            remark=data.remark,
            updated_at=datetime.now()
        )
    )
    await db.commit()

    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    updated = result.scalar_one()

    return ResponseModel(data=EnergySuggestionResponse.model_validate(updated))


@router.put("/suggestions/{suggestion_id}/reject", response_model=ResponseModel[EnergySuggestionResponse], summary="拒绝建议")
async def reject_suggestion(
    suggestion_id: int,
    data: RejectSuggestion,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """拒绝节能建议"""
    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    suggestion = result.scalar_one_or_none()

    if not suggestion:
        raise HTTPException(status_code=404, detail="建议不存在")

    if suggestion.status != "pending":
        raise HTTPException(status_code=400, detail="只能拒绝待处理的建议")

    await db.execute(
        update(EnergySuggestion).where(EnergySuggestion.id == suggestion_id).values(
            status="rejected",
            remark=data.remark,
            updated_at=datetime.now()
        )
    )
    await db.commit()

    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    updated = result.scalar_one()

    return ResponseModel(data=EnergySuggestionResponse.model_validate(updated))


@router.put("/suggestions/{suggestion_id}/complete", response_model=ResponseModel[EnergySuggestionResponse], summary="完成建议")
async def complete_suggestion(
    suggestion_id: int,
    data: CompleteSuggestion,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """标记建议为已完成"""
    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    suggestion = result.scalar_one_or_none()

    if not suggestion:
        raise HTTPException(status_code=404, detail="建议不存在")

    if suggestion.status != "accepted":
        raise HTTPException(status_code=400, detail="只能完成已接受的建议")

    await db.execute(
        update(EnergySuggestion).where(EnergySuggestion.id == suggestion_id).values(
            status="completed",
            completed_at=datetime.now(),
            actual_saving=data.actual_saving,
            remark=data.remark if data.remark else suggestion.remark,
            updated_at=datetime.now()
        )
    )
    await db.commit()

    result = await db.execute(select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id))
    updated = result.scalar_one()

    return ResponseModel(data=EnergySuggestionResponse.model_validate(updated))


@router.get("/saving/potential", response_model=ResponseModel[SavingPotential], summary="获取节能潜力")
async def get_saving_potential(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取节能潜力分析"""
    # 统计各类建议数量
    high_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.priority == "high", EnergySuggestion.status == "pending")
    )
    medium_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.priority == "medium", EnergySuggestion.status == "pending")
    )
    low_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.priority == "low", EnergySuggestion.status == "pending")
    )
    pending_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.status == "pending")
    )
    accepted_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.status == "accepted")
    )
    completed_count = await db.execute(
        select(func.count()).select_from(EnergySuggestion)
        .where(EnergySuggestion.status == "completed")
    )

    # 计算潜在节能
    potential_result = await db.execute(
        select(
            func.sum(EnergySuggestion.potential_saving),
            func.sum(EnergySuggestion.potential_cost_saving)
        ).where(EnergySuggestion.status == "pending")
    )
    potential = potential_result.first()

    # 计算实际节能
    actual_result = await db.execute(
        select(func.sum(EnergySuggestion.actual_saving))
        .where(
            EnergySuggestion.status == "completed",
            EnergySuggestion.completed_at >= date(date.today().year, 1, 1)
        )
    )
    actual_saving = actual_result.scalar() or 0

    saving = SavingPotential(
        total_potential_saving=potential[0] or 0,
        total_cost_saving=potential[1] or 0,
        high_priority_count=high_count.scalar() or 0,
        medium_priority_count=medium_count.scalar() or 0,
        low_priority_count=low_count.scalar() or 0,
        pending_count=pending_count.scalar() or 0,
        accepted_count=accepted_count.scalar() or 0,
        completed_count=completed_count.scalar() or 0,
        actual_saving_ytd=actual_saving
    )

    return ResponseModel(data=saving)


# ==================== V2.4 数据驱动 - 电价和设备能力查询 ====================

@router.get("/pricing/current", response_model=ResponseModel, summary="获取当前电价配置")
async def get_current_pricing(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取当前有效的电价配置，用于前端展示

    返回数据来源：electricity_pricing表
    """
    from ...services.pricing_service import PricingService

    pricing_service = PricingService(db)
    pricing = await pricing_service.get_current_pricing()
    time_periods = await pricing_service.get_time_periods_for_display()
    data_source = await pricing_service.get_pricing_data_source()

    return ResponseModel(data={
        "pricing": pricing,
        "time_periods": time_periods,
        "data_source": data_source
    })


@router.post("/suggestions/{suggestion_id}/recalculate", response_model=ResponseModel, summary="调整参数并重算")
async def recalculate_suggestion(
    suggestion_id: int,
    params: Dict[str, Any],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    根据用户调整的参数重新计算节能效果

    请求体示例:
    {
        "selected_devices": [101, 102],  # 选择的设备ID
        "shift_hours": 3,                 # 转移时长
        "source_period": "sharp",         # 转出时段
        "target_period": "valley"         # 转入时段
    }

    返回重新计算的节能效果，使用真实电价和设备数据
    """
    from ...services.suggestion_engine import SuggestionEngineService
    from pydantic import BaseModel, Field
    from typing import List

    # 定义请求schema
    class RecalculateRequest(BaseModel):
        selected_devices: List[int] = Field(..., description="选择的设备ID列表")
        shift_hours: float = Field(..., ge=0.5, le=24, description="转移时长（小时）")
        source_period: str = Field(..., description="转出时段（sharp/peak/normal）")
        target_period: str = Field(..., description="转入时段（valley/deep_valley/normal）")

    # 验证建议是否存在
    result = await db.execute(
        select(EnergySuggestion).where(EnergySuggestion.id == suggestion_id)
    )
    suggestion = result.scalar_one_or_none()
    if not suggestion:
        raise HTTPException(status_code=404, detail="建议不存在")

    # 重新计算
    engine = SuggestionEngineService(db)
    calc_result = await engine.recalculate_suggestion(suggestion_id, params)

    return ResponseModel(data=calc_result)


# ==================== 配电图 ====================

@router.get("/distribution", response_model=ResponseModel[DistributionDiagram], summary="获取配电图数据")
async def get_distribution_diagram(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取配电图数据（含实时功率和负载率）"""
    result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True)
    )
    devices = result.scalars().all()

    if not devices:
        raise HTTPException(status_code=404, detail="暂无配电设备数据")

    import random

    # 构建设备节点字典
    node_dict = {}
    for device in devices:
        base_power = (device.rated_power or 10.0) * random.uniform(0.5, 0.9)
        load_rate = base_power / device.rated_power if device.rated_power else None

        node = DistributionNode(
            device_id=device.id,
            device_code=device.device_code,
            device_name=device.device_name,
            device_type=device.device_type,
            power=round(base_power, 2),
            load_rate=round(load_rate, 2) if load_rate else None,
            status="normal" if (load_rate or 0) < 0.8 else "warning",
            children=[]
        )
        node_dict[device.id] = node

    # 构建树结构
    root = None
    total_power = 0.0

    for device in devices:
        node = node_dict[device.id]
        if device.parent_device_id and device.parent_device_id in node_dict:
            node_dict[device.parent_device_id].children.append(node)
        else:
            if device.device_type == "MAIN":
                root = node
                total_power = node.power or 0

    if not root:
        # 如果没有MAIN类型设备，取第一个无父级的设备作为根
        for device in devices:
            if not device.parent_device_id:
                root = node_dict[device.id]
                total_power = root.power or 0
                break

    if not root:
        root = list(node_dict.values())[0]
        total_power = root.power or 0

    diagram = DistributionDiagram(
        root=root,
        total_power=round(total_power, 2),
        timestamp=datetime.now()
    )

    return ResponseModel(data=diagram)


# ==================== 数据导出 ====================

@router.get("/export/daily", summary="导出日能耗数据")
async def export_daily_data(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    format: str = Query("excel", description="格式: excel/csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出日能耗数据"""
    result = await db.execute(
        select(EnergyDaily)
        .where(EnergyDaily.stat_date >= start_date, EnergyDaily.stat_date <= end_date)
        .order_by(EnergyDaily.stat_date)
    )
    daily_data = result.scalars().all()

    import random

    # 如果没有数据，生成模拟数据
    if not daily_data:
        data_list = []
        current = start_date
        while current <= end_date:
            total = 1000 + random.uniform(-200, 200)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            data_list.append({
                "日期": current.isoformat(),
                "总电量(kWh)": round(total, 2),
                "峰时电量(kWh)": round(peak, 2),
                "平时电量(kWh)": round(normal, 2),
                "谷时电量(kWh)": round(valley, 2),
                "最大功率(kW)": round(total / 20, 2),
                "平均功率(kW)": round(total / 24, 2),
                "电费(元)": round(peak * 1.2 + normal * 0.8 + valley * 0.4, 2),
                "PUE": round(1.4 + random.uniform(-0.1, 0.2), 2)
            })
            current += timedelta(days=1)
    else:
        data_list = []
        for d in daily_data:
            data_list.append({
                "日期": d.stat_date.isoformat(),
                "总电量(kWh)": d.total_energy,
                "峰时电量(kWh)": d.peak_energy,
                "平时电量(kWh)": d.normal_energy,
                "谷时电量(kWh)": d.valley_energy,
                "最大功率(kW)": d.max_power,
                "平均功率(kW)": d.avg_power,
                "电费(元)": d.energy_cost,
                "PUE": d.pue
            })

    if format == "csv":
        import csv
        output = BytesIO()
        import codecs
        output.write(codecs.BOM_UTF8)

        if data_list:
            writer = csv.DictWriter(output, fieldnames=data_list[0].keys())
            output.write(','.join(data_list[0].keys()).encode('utf-8'))
            output.write(b'\n')
            for row in data_list:
                output.write(','.join(str(v) for v in row.values()).encode('utf-8'))
                output.write(b'\n')

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=energy_daily_{start_date}_{end_date}.csv"}
        )
    else:
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "日能耗统计"

            if data_list:
                headers = list(data_list[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                for row_idx, row_data in enumerate(data_list, 2):
                    for col, value in enumerate(row_data.values(), 1):
                        ws.cell(row=row_idx, column=col, value=value)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=energy_daily_{start_date}_{end_date}.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel导出功能需要安装openpyxl库")


@router.get("/export/monthly", summary="导出月能耗数据")
async def export_monthly_data(
    year: int = Query(..., description="年份"),
    format: str = Query("excel", description="格式: excel/csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """导出月能耗数据"""
    result = await db.execute(
        select(EnergyMonthly)
        .where(EnergyMonthly.stat_year == year)
        .order_by(EnergyMonthly.stat_month)
    )
    monthly_data = result.scalars().all()

    import random

    if not monthly_data:
        data_list = []
        for month in range(1, 13):
            total = 30000 + random.uniform(-5000, 5000)
            peak = total * 0.4
            normal = total * 0.35
            valley = total * 0.25
            data_list.append({
                "年份": year,
                "月份": month,
                "总电量(kWh)": round(total, 2),
                "峰时电量(kWh)": round(peak, 2),
                "平时电量(kWh)": round(normal, 2),
                "谷时电量(kWh)": round(valley, 2),
                "最大功率(kW)": round(total / 500, 2),
                "平均功率(kW)": round(total / 720, 2),
                "峰时电费(元)": round(peak * 1.2, 2),
                "平时电费(元)": round(normal * 0.8, 2),
                "谷时电费(元)": round(valley * 0.4, 2),
                "总电费(元)": round(peak * 1.2 + normal * 0.8 + valley * 0.4, 2),
                "平均PUE": round(1.45 + random.uniform(-0.1, 0.15), 2)
            })
    else:
        data_list = []
        for d in monthly_data:
            data_list.append({
                "年份": d.stat_year,
                "月份": d.stat_month,
                "总电量(kWh)": d.total_energy,
                "峰时电量(kWh)": d.peak_energy,
                "平时电量(kWh)": d.normal_energy,
                "谷时电量(kWh)": d.valley_energy,
                "最大功率(kW)": d.max_power,
                "平均功率(kW)": d.avg_power,
                "峰时电费(元)": d.peak_cost,
                "平时电费(元)": d.normal_cost,
                "谷时电费(元)": d.valley_cost,
                "总电费(元)": d.energy_cost,
                "平均PUE": d.avg_pue
            })

    if format == "csv":
        import csv
        output = BytesIO()
        import codecs
        output.write(codecs.BOM_UTF8)

        if data_list:
            output.write(','.join(data_list[0].keys()).encode('utf-8'))
            output.write(b'\n')
            for row in data_list:
                output.write(','.join(str(v) for v in row.values()).encode('utf-8'))
                output.write(b'\n')

        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=energy_monthly_{year}.csv"}
        )
    else:
        try:
            import openpyxl
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "月能耗统计"

            if data_list:
                headers = list(data_list[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

                for row_idx, row_data in enumerate(data_list, 2):
                    for col, value in enumerate(row_data.values(), 1):
                        ws.cell(row=row_idx, column=col, value=value)

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=energy_monthly_{year}.xlsx"}
            )
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel导出功能需要安装openpyxl库")


# ==================== 变压器管理 ====================

@router.get("/transformers", response_model=ResponseModel[List[TransformerResponse]], summary="获取变压器列表")
async def get_transformers(
    status: Optional[str] = Query(None, description="状态"),
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取变压器列表"""
    query = select(Transformer)
    if status:
        query = query.where(Transformer.status == status)
    if is_enabled is not None:
        query = query.where(Transformer.is_enabled == is_enabled)
    query = query.order_by(Transformer.transformer_code)

    result = await db.execute(query)
    transformers = result.scalars().all()
    return ResponseModel(data=[TransformerResponse.model_validate(t) for t in transformers])


@router.post("/transformers", response_model=ResponseModel[TransformerResponse], summary="创建变压器")
async def create_transformer(
    transformer: TransformerCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """创建变压器"""
    existing = await db.execute(
        select(Transformer).where(Transformer.transformer_code == transformer.transformer_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="变压器编码已存在")

    new_transformer = Transformer(**transformer.model_dump())
    db.add(new_transformer)
    await db.commit()
    await db.refresh(new_transformer)
    return ResponseModel(data=TransformerResponse.model_validate(new_transformer))


@router.get("/transformers/{transformer_id}", response_model=ResponseModel[TransformerResponse], summary="获取变压器详情")
async def get_transformer(
    transformer_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取变压器详情"""
    result = await db.execute(select(Transformer).where(Transformer.id == transformer_id))
    transformer = result.scalar_one_or_none()
    if not transformer:
        raise HTTPException(status_code=404, detail="变压器不存在")
    return ResponseModel(data=TransformerResponse.model_validate(transformer))


@router.put("/transformers/{transformer_id}", response_model=ResponseModel[TransformerResponse], summary="更新变压器")
async def update_transformer(
    transformer_id: int,
    data: TransformerUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新变压器"""
    result = await db.execute(select(Transformer).where(Transformer.id == transformer_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="变压器不存在")

    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()
    await db.execute(update(Transformer).where(Transformer.id == transformer_id).values(**update_data))
    await db.commit()

    result = await db.execute(select(Transformer).where(Transformer.id == transformer_id))
    return ResponseModel(data=TransformerResponse.model_validate(result.scalar_one()))


@router.delete("/transformers/{transformer_id}", response_model=ResponseModel, summary="删除变压器")
async def delete_transformer(
    transformer_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除变压器"""
    result = await db.execute(select(Transformer).where(Transformer.id == transformer_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="变压器不存在")

    # 检查是否有关联的计量点
    meters = await db.execute(select(MeterPoint).where(MeterPoint.transformer_id == transformer_id))
    if meters.scalars().first():
        raise HTTPException(status_code=400, detail="该变压器有关联的计量点，无法删除")

    await db.execute(delete(Transformer).where(Transformer.id == transformer_id))
    await db.commit()
    return ResponseModel(message="删除成功")


# ==================== 计量点管理 ====================

@router.get("/meter-points", response_model=ResponseModel[List[MeterPointResponse]], summary="获取计量点列表")
async def get_meter_points(
    transformer_id: Optional[int] = Query(None, description="变压器ID"),
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取计量点列表"""
    query = select(MeterPoint)
    if transformer_id:
        query = query.where(MeterPoint.transformer_id == transformer_id)
    if is_enabled is not None:
        query = query.where(MeterPoint.is_enabled == is_enabled)
    query = query.order_by(MeterPoint.meter_code)

    result = await db.execute(query)
    meter_points = result.scalars().all()
    return ResponseModel(data=[MeterPointResponse.model_validate(m) for m in meter_points])


@router.post("/meter-points", response_model=ResponseModel[MeterPointResponse], summary="创建计量点")
async def create_meter_point(
    meter_point: MeterPointCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """创建计量点"""
    existing = await db.execute(
        select(MeterPoint).where(MeterPoint.meter_code == meter_point.meter_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="计量点编码已存在")

    if meter_point.transformer_id:
        transformer = await db.execute(
            select(Transformer).where(Transformer.id == meter_point.transformer_id)
        )
        if not transformer.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="关联的变压器不存在")

    new_meter = MeterPoint(**meter_point.model_dump())
    db.add(new_meter)
    await db.commit()
    await db.refresh(new_meter)
    return ResponseModel(data=MeterPointResponse.model_validate(new_meter))


@router.get("/meter-points/{meter_point_id}", response_model=ResponseModel[MeterPointDetail], summary="获取计量点详情")
async def get_meter_point(
    meter_point_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取计量点详情（含变压器信息）"""
    result = await db.execute(select(MeterPoint).where(MeterPoint.id == meter_point_id))
    meter_point = result.scalar_one_or_none()
    if not meter_point:
        raise HTTPException(status_code=404, detail="计量点不存在")

    response_data = MeterPointResponse.model_validate(meter_point).model_dump()

    # 获取变压器信息
    if meter_point.transformer_id:
        transformer_result = await db.execute(
            select(Transformer).where(Transformer.id == meter_point.transformer_id)
        )
        transformer = transformer_result.scalar_one_or_none()
        if transformer:
            response_data["transformer_name"] = transformer.transformer_name
            response_data["transformer_capacity"] = transformer.rated_capacity

    return ResponseModel(data=MeterPointDetail(**response_data))


@router.put("/meter-points/{meter_point_id}", response_model=ResponseModel[MeterPointResponse], summary="更新计量点")
async def update_meter_point(
    meter_point_id: int,
    data: MeterPointUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新计量点"""
    result = await db.execute(select(MeterPoint).where(MeterPoint.id == meter_point_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="计量点不存在")

    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()
    await db.execute(update(MeterPoint).where(MeterPoint.id == meter_point_id).values(**update_data))
    await db.commit()

    result = await db.execute(select(MeterPoint).where(MeterPoint.id == meter_point_id))
    return ResponseModel(data=MeterPointResponse.model_validate(result.scalar_one()))


@router.delete("/meter-points/{meter_point_id}", response_model=ResponseModel, summary="删除计量点")
async def delete_meter_point(
    meter_point_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除计量点"""
    result = await db.execute(select(MeterPoint).where(MeterPoint.id == meter_point_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="计量点不存在")

    # 检查是否有关联的配电柜
    panels = await db.execute(select(DistributionPanel).where(DistributionPanel.meter_point_id == meter_point_id))
    if panels.scalars().first():
        raise HTTPException(status_code=400, detail="该计量点有关联的配电柜，无法删除")

    await db.execute(delete(MeterPoint).where(MeterPoint.id == meter_point_id))
    await db.commit()
    return ResponseModel(message="删除成功")


# ==================== 配电柜管理 ====================

@router.get("/panels", response_model=ResponseModel[List[DistributionPanelResponse]], summary="获取配电柜列表")
async def get_distribution_panels(
    panel_type: Optional[str] = Query(None, description="配电柜类型"),
    meter_point_id: Optional[int] = Query(None, description="计量点ID"),
    is_enabled: Optional[bool] = Query(None, description="是否启用"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取配电柜列表"""
    query = select(DistributionPanel)
    if panel_type:
        query = query.where(DistributionPanel.panel_type == panel_type)
    if meter_point_id:
        query = query.where(DistributionPanel.meter_point_id == meter_point_id)
    if is_enabled is not None:
        query = query.where(DistributionPanel.is_enabled == is_enabled)
    query = query.order_by(DistributionPanel.panel_code)

    result = await db.execute(query)
    panels = result.scalars().all()
    return ResponseModel(data=[DistributionPanelResponse.model_validate(p) for p in panels])


@router.get("/panels/{panel_id}", response_model=ResponseModel[DistributionPanelResponse], summary="获取配电柜详情")
async def get_distribution_panel(
    panel_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取配电柜详情"""
    result = await db.execute(select(DistributionPanel).where(DistributionPanel.id == panel_id))
    panel = result.scalar_one_or_none()
    if not panel:
        raise HTTPException(status_code=404, detail="配电柜不存在")
    return ResponseModel(data=DistributionPanelResponse.model_validate(panel))


@router.post("/panels", response_model=ResponseModel[DistributionPanelResponse], summary="创建配电柜")
async def create_distribution_panel(
    panel: DistributionPanelCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """创建配电柜"""
    existing = await db.execute(
        select(DistributionPanel).where(DistributionPanel.panel_code == panel.panel_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="配电柜编码已存在")

    new_panel = DistributionPanel(**panel.model_dump())
    db.add(new_panel)
    await db.commit()
    await db.refresh(new_panel)
    return ResponseModel(data=DistributionPanelResponse.model_validate(new_panel))


@router.put("/panels/{panel_id}", response_model=ResponseModel[DistributionPanelResponse], summary="更新配电柜")
async def update_distribution_panel(
    panel_id: int,
    data: DistributionPanelUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新配电柜"""
    result = await db.execute(select(DistributionPanel).where(DistributionPanel.id == panel_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="配电柜不存在")

    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()
    await db.execute(update(DistributionPanel).where(DistributionPanel.id == panel_id).values(**update_data))
    await db.commit()

    result = await db.execute(select(DistributionPanel).where(DistributionPanel.id == panel_id))
    return ResponseModel(data=DistributionPanelResponse.model_validate(result.scalar_one()))


@router.delete("/panels/{panel_id}", response_model=ResponseModel, summary="删除配电柜")
async def delete_distribution_panel(
    panel_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除配电柜"""
    result = await db.execute(select(DistributionPanel).where(DistributionPanel.id == panel_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="配电柜不存在")

    # 检查是否有关联的回路
    circuits = await db.execute(select(DistributionCircuit).where(DistributionCircuit.panel_id == panel_id))
    if circuits.scalars().first():
        raise HTTPException(status_code=400, detail="该配电柜有关联的回路，无法删除")

    await db.execute(delete(DistributionPanel).where(DistributionPanel.id == panel_id))
    await db.commit()
    return ResponseModel(message="删除成功")


# ==================== 配电回路管理 ====================

@router.get("/circuits", response_model=ResponseModel[List[DistributionCircuitResponse]], summary="获取配电回路列表")
async def get_distribution_circuits(
    panel_id: Optional[int] = Query(None, description="配电柜ID"),
    load_type: Optional[str] = Query(None, description="负载类型"),
    is_shiftable: Optional[bool] = Query(None, description="是否可转移"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取配电回路列表"""
    query = select(DistributionCircuit)
    if panel_id:
        query = query.where(DistributionCircuit.panel_id == panel_id)
    if load_type:
        query = query.where(DistributionCircuit.load_type == load_type)
    if is_shiftable is not None:
        query = query.where(DistributionCircuit.is_shiftable == is_shiftable)
    query = query.order_by(DistributionCircuit.circuit_code)

    result = await db.execute(query)
    circuits = result.scalars().all()
    return ResponseModel(data=[DistributionCircuitResponse.model_validate(c) for c in circuits])


@router.get("/circuits/{circuit_id}", response_model=ResponseModel[DistributionCircuitResponse], summary="获取配电回路详情")
async def get_distribution_circuit(
    circuit_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取配电回路详情"""
    result = await db.execute(select(DistributionCircuit).where(DistributionCircuit.id == circuit_id))
    circuit = result.scalar_one_or_none()
    if not circuit:
        raise HTTPException(status_code=404, detail="配电回路不存在")
    return ResponseModel(data=DistributionCircuitResponse.model_validate(circuit))


@router.post("/circuits", response_model=ResponseModel[DistributionCircuitResponse], summary="创建配电回路")
async def create_distribution_circuit(
    circuit: DistributionCircuitCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """创建配电回路"""
    existing = await db.execute(
        select(DistributionCircuit).where(DistributionCircuit.circuit_code == circuit.circuit_code)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="回路编码已存在")

    # 检查配电柜是否存在
    panel = await db.execute(select(DistributionPanel).where(DistributionPanel.id == circuit.panel_id))
    if not panel.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="关联的配电柜不存在")

    new_circuit = DistributionCircuit(**circuit.model_dump())
    db.add(new_circuit)
    await db.commit()
    await db.refresh(new_circuit)
    return ResponseModel(data=DistributionCircuitResponse.model_validate(new_circuit))


@router.put("/circuits/{circuit_id}", response_model=ResponseModel[DistributionCircuitResponse], summary="更新配电回路")
async def update_distribution_circuit(
    circuit_id: int,
    data: DistributionCircuitUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """更新配电回路"""
    result = await db.execute(select(DistributionCircuit).where(DistributionCircuit.id == circuit_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="回路不存在")

    update_data = data.model_dump(exclude_unset=True)
    update_data["updated_at"] = datetime.now()
    await db.execute(update(DistributionCircuit).where(DistributionCircuit.id == circuit_id).values(**update_data))
    await db.commit()

    result = await db.execute(select(DistributionCircuit).where(DistributionCircuit.id == circuit_id))
    return ResponseModel(data=DistributionCircuitResponse.model_validate(result.scalar_one()))


@router.delete("/circuits/{circuit_id}", response_model=ResponseModel, summary="删除配电回路")
async def delete_distribution_circuit(
    circuit_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """删除配电回路"""
    result = await db.execute(select(DistributionCircuit).where(DistributionCircuit.id == circuit_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="回路不存在")

    # 检查是否有关联的设备
    devices = await db.execute(select(PowerDevice).where(PowerDevice.circuit_id == circuit_id))
    if devices.scalars().first():
        raise HTTPException(status_code=400, detail="该回路有关联的设备，无法删除")

    await db.execute(delete(DistributionCircuit).where(DistributionCircuit.id == circuit_id))
    await db.commit()
    return ResponseModel(message="删除成功")


# ==================== 配电系统拓扑 ====================

@router.get("/topology", response_model=ResponseModel[DistributionTopologyResponse], summary="获取配电系统拓扑")
async def get_distribution_topology(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取完整的配电系统拓扑结构
    变压器 → 计量点 → 配电柜 → 回路 → 设备
    """
    # 获取所有变压器
    transformers_result = await db.execute(
        select(Transformer).where(Transformer.is_enabled == True).order_by(Transformer.transformer_code)
    )
    transformers = transformers_result.scalars().all()

    # 获取所有计量点
    meters_result = await db.execute(
        select(MeterPoint).where(MeterPoint.is_enabled == True)
    )
    meters = meters_result.scalars().all()

    # 获取所有配电柜
    panels_result = await db.execute(
        select(DistributionPanel).where(DistributionPanel.is_enabled == True)
    )
    panels = panels_result.scalars().all()

    # 获取所有回路
    circuits_result = await db.execute(
        select(DistributionCircuit).where(DistributionCircuit.is_enabled == True)
    )
    circuits = circuits_result.scalars().all()

    # 获取所有设备
    devices_result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True)
    )
    devices = devices_result.scalars().all()

    # 构建拓扑结构
    # 设备按回路分组
    circuit_devices = {}
    for device in devices:
        if device.circuit_id:
            if device.circuit_id not in circuit_devices:
                circuit_devices[device.circuit_id] = []
            circuit_devices[device.circuit_id].append(
                PowerDeviceResponse.model_validate(device)
            )

    # 回路按配电柜分组
    panel_circuits = {}
    for circuit in circuits:
        if circuit.panel_id not in panel_circuits:
            panel_circuits[circuit.panel_id] = []
        panel_circuits[circuit.panel_id].append(TopologyCircuitNode(
            circuit_id=circuit.id,
            circuit_code=circuit.circuit_code,
            circuit_name=circuit.circuit_name,
            load_type=circuit.load_type,
            is_shiftable=circuit.is_shiftable,
            devices=circuit_devices.get(circuit.id, [])
        ))

    # 配电柜按计量点分组
    meter_panels = {}
    for panel in panels:
        if panel.meter_point_id:
            if panel.meter_point_id not in meter_panels:
                meter_panels[panel.meter_point_id] = []
            meter_panels[panel.meter_point_id].append(TopologyPanelNode(
                panel_id=panel.id,
                panel_code=panel.panel_code,
                panel_name=panel.panel_name,
                panel_type=panel.panel_type,
                circuits=panel_circuits.get(panel.id, [])
            ))

    # 计量点按变压器分组
    transformer_meters = {}
    for meter in meters:
        if meter.transformer_id:
            if meter.transformer_id not in transformer_meters:
                transformer_meters[meter.transformer_id] = []
            transformer_meters[meter.transformer_id].append(TopologyMeterNode(
                meter_point_id=meter.id,
                meter_code=meter.meter_code,
                meter_name=meter.meter_name,
                declared_demand=meter.declared_demand,
                demand_type=meter.demand_type,
                panels=meter_panels.get(meter.id, [])
            ))

    # 构建变压器节点
    transformer_nodes = []
    total_capacity = 0
    for transformer in transformers:
        total_capacity += transformer.rated_capacity
        transformer_nodes.append(TopologyTransformerNode(
            transformer_id=transformer.id,
            transformer_code=transformer.transformer_code,
            transformer_name=transformer.transformer_name,
            rated_capacity=transformer.rated_capacity,
            meter_points=transformer_meters.get(transformer.id, [])
        ))

    return ResponseModel(data=DistributionTopologyResponse(
        transformers=transformer_nodes,
        total_capacity=total_capacity,
        total_meter_points=len(meters),
        total_devices=len(devices)
    ))


# ==================== 功率曲线 ====================

@router.get("/power-curve", response_model=ResponseModel[PowerCurveResponse], summary="获取功率曲线")
async def get_power_curve(
    start_time: datetime = Query(..., description="开始时间"),
    end_time: datetime = Query(..., description="结束时间"),
    meter_point_id: Optional[int] = Query(None, description="计量点ID"),
    device_id: Optional[int] = Query(None, description="设备ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取功率曲线数据
    支持按计量点或设备查询
    """
    query = select(PowerCurveData).where(
        PowerCurveData.timestamp >= start_time,
        PowerCurveData.timestamp <= end_time
    )

    if meter_point_id:
        query = query.where(PowerCurveData.meter_point_id == meter_point_id)
    if device_id:
        query = query.where(PowerCurveData.device_id == device_id)

    query = query.order_by(PowerCurveData.timestamp)
    result = await db.execute(query)
    curve_data = result.scalars().all()

    # 如果没有数据，生成模拟数据
    import random
    if not curve_data:
        data_list = []
        current = start_time
        max_power = 0.0
        total_power = 0.0
        max_demand = 0.0
        count = 0

        while current <= end_time:
            hour = current.hour
            # 根据时段确定负载系数
            if 10 <= hour < 12 or 19 <= hour < 21:
                time_period = "sharp"
                load_factor = random.uniform(0.85, 1.0)
            elif 8 <= hour < 10 or 12 <= hour < 14 or 17 <= hour < 19 or 21 <= hour < 23:
                time_period = "peak"
                load_factor = random.uniform(0.7, 0.9)
            elif 0 <= hour < 7:
                time_period = "valley"
                load_factor = random.uniform(0.3, 0.5)
            else:
                time_period = "flat"
                load_factor = random.uniform(0.5, 0.7)

            base_power = 100  # 基准功率
            power = base_power * load_factor
            demand = power * random.uniform(0.95, 1.05)

            max_power = max(max_power, power)
            max_demand = max(max_demand, demand)
            total_power += power
            count += 1

            data_list.append(PowerCurvePoint(
                timestamp=current,
                meter_point_id=meter_point_id,
                device_id=device_id,
                active_power=round(power, 2),
                reactive_power=round(power * 0.2, 2),
                power_factor=round(random.uniform(0.85, 0.95), 3),
                demand_15min=round(demand, 2),
                time_period=time_period
            ))
            current += timedelta(minutes=15)

        avg_power = total_power / count if count > 0 else 0

        return ResponseModel(data=PowerCurveResponse(
            meter_point_id=meter_point_id,
            device_id=device_id,
            data=data_list,
            max_power=round(max_power, 2),
            avg_power=round(avg_power, 2),
            max_demand=round(max_demand, 2)
        ))

    # 有真实数据时
    data_list = []
    max_power = 0.0
    total_power = 0.0
    max_demand = 0.0

    for d in curve_data:
        max_power = max(max_power, d.active_power or 0)
        max_demand = max(max_demand, d.demand_15min or 0)
        total_power += d.active_power or 0

        data_list.append(PowerCurvePoint(
            timestamp=d.timestamp,
            meter_point_id=d.meter_point_id,
            device_id=d.device_id,
            active_power=d.active_power or 0,
            reactive_power=d.reactive_power or 0,
            power_factor=d.power_factor or 0.9,
            demand_15min=d.demand_15min or 0,
            time_period=d.time_period or "flat"
        ))

    avg_power = total_power / len(curve_data) if curve_data else 0

    return ResponseModel(data=PowerCurveResponse(
        meter_point_id=meter_point_id,
        device_id=device_id,
        data=data_list,
        max_power=round(max_power, 2),
        avg_power=round(avg_power, 2),
        max_demand=round(max_demand, 2)
    ))


# ==================== 需量历史 ====================

@router.get("/demand-history/{meter_point_id}", response_model=ResponseModel[DemandHistoryResponse], summary="获取需量历史")
async def get_demand_history(
    meter_point_id: int,
    months: int = Query(12, description="历史月数", ge=1, le=36),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """获取指定计量点的需量历史数据"""
    # 获取计量点信息
    meter_result = await db.execute(select(MeterPoint).where(MeterPoint.id == meter_point_id))
    meter = meter_result.scalar_one_or_none()
    if not meter:
        raise HTTPException(status_code=404, detail="计量点不存在")

    # 获取需量历史
    now = datetime.now()
    start_year = now.year if now.month > months else now.year - 1
    start_month = (now.month - months) % 12 or 12

    query = select(DemandHistory).where(
        DemandHistory.meter_point_id == meter_point_id
    ).order_by(DemandHistory.stat_year.desc(), DemandHistory.stat_month.desc())

    result = await db.execute(query)
    history_data = result.scalars().all()

    import random
    if not history_data:
        # 生成模拟数据
        history_list = []
        declared = meter.declared_demand or 500
        for i in range(months):
            month_offset = now.month - i - 1
            year = now.year + (month_offset // 12)
            month = month_offset % 12 + 1

            max_d = declared * random.uniform(0.7, 1.05)
            avg_d = max_d * random.uniform(0.6, 0.8)
            d_95th = max_d * random.uniform(0.9, 0.98)
            utilization = max_d / declared if declared > 0 else 0

            history_list.append(DemandHistoryItem(
                month=f"{year}-{month:02d}",
                declared_demand=declared,
                max_demand=round(max_d, 2),
                avg_demand=round(avg_d, 2),
                demand_95th=round(d_95th, 2),
                over_declared_times=1 if max_d > declared else 0,
                demand_cost=round(declared * 38, 2),
                utilization_rate=round(utilization * 100, 1)
            ))

        return ResponseModel(data=DemandHistoryResponse(
            meter_point_id=meter_point_id,
            meter_name=meter.meter_name,
            declared_demand=meter.declared_demand or 0,
            history=history_list
        ))

    # 转换真实数据
    history_list = []
    for h in history_data[:months]:
        utilization = h.max_demand / h.declared_demand if h.declared_demand and h.declared_demand > 0 else 0
        history_list.append(DemandHistoryItem(
            month=f"{h.stat_year}-{h.stat_month:02d}",
            declared_demand=h.declared_demand or 0,
            max_demand=h.max_demand or 0,
            avg_demand=h.avg_demand or 0,
            demand_95th=h.demand_95th or 0,
            over_declared_times=h.over_declared_times or 0,
            demand_cost=h.demand_cost or 0,
            utilization_rate=round(utilization * 100, 1)
        ))

    return ResponseModel(data=DemandHistoryResponse(
        meter_point_id=meter_point_id,
        meter_name=meter.meter_name,
        declared_demand=meter.declared_demand or 0,
        history=history_list
    ))


# ==================== 设备负荷转移分析 ====================

@router.get("/analysis/device-shift", response_model=ResponseModel[DeviceShiftAnalysisResult], summary="设备负荷转移分析")
async def analyze_device_shift(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    分析设备负荷转移潜力
    识别哪些设备可以从峰时转移到谷时运行
    """
    # 获取所有设备及其转移配置
    devices_result = await db.execute(
        select(PowerDevice).where(PowerDevice.is_enabled == True)
    )
    devices = devices_result.scalars().all()

    # 获取设备转移配置
    shift_configs_result = await db.execute(select(DeviceShiftConfig))
    shift_configs = {sc.device_id: sc for sc in shift_configs_result.scalars().all()}

    import random

    device_potentials = []
    total_shiftable_power = 0.0
    total_potential_saving = 0.0
    shiftable_count = 0

    for device in devices:
        config = shift_configs.get(device.id)

        # 判断是否可转移
        is_shiftable = config.is_shiftable if config else (
            device.device_type in ["HVAC", "PUMP", "LIGHTING"] and not device.is_critical
        )
        is_critical = config.is_critical if config else device.is_critical

        # 计算当前功率和可转移功率
        rated_power = device.rated_power or 10.0
        current_power = rated_power * random.uniform(0.5, 0.9)
        shiftable_ratio = config.shiftable_power_ratio if config else (0.5 if is_shiftable else 0)
        shiftable_power = current_power * shiftable_ratio

        # 计算5时段用电比例（模拟数据）
        # 确保5个时段占比之和为100%
        if is_shiftable:
            # 可转移设备：高价时段占比较高
            sharp_ratio = random.uniform(0.08, 0.15)
            peak_ratio = random.uniform(0.30, 0.40)
            flat_ratio = random.uniform(0.25, 0.35)
            valley_ratio = random.uniform(0.10, 0.18)
        else:
            # 不可转移设备：分布相对均匀
            sharp_ratio = random.uniform(0.05, 0.10)
            peak_ratio = random.uniform(0.25, 0.35)
            flat_ratio = random.uniform(0.30, 0.40)
            valley_ratio = random.uniform(0.12, 0.20)

        # 深谷占比为剩余部分，确保总和为100%
        deep_valley_ratio = max(0, 1.0 - sharp_ratio - peak_ratio - flat_ratio - valley_ratio)

        # 计算节省潜力（基于5时段价差）
        # 尖峰1.4, 峰时1.0, 平时0.65, 谷时0.35, 深谷0.2
        high_cost_ratio = sharp_ratio + peak_ratio  # 高价时段占比
        low_cost_ratio = valley_ratio + deep_valley_ratio  # 低价时段占比
        avg_high_price = (1.4 * sharp_ratio + 1.0 * peak_ratio) / high_cost_ratio if high_cost_ratio > 0 else 1.0
        avg_low_price = (0.35 * valley_ratio + 0.2 * deep_valley_ratio) / low_cost_ratio if low_cost_ratio > 0 else 0.3
        price_diff = avg_high_price - avg_low_price

        monthly_hours = 30 * 24
        shift_saving = shiftable_power * (high_cost_ratio - low_cost_ratio) * monthly_hours * price_diff if is_shiftable else 0

        if is_shiftable:
            shiftable_count += 1
            total_shiftable_power += shiftable_power
            total_potential_saving += shift_saving

        device_potentials.append(DeviceShiftPotential(
            device_id=device.id,
            device_name=device.device_name,
            device_type=device.device_type,
            rated_power=round(rated_power, 2),
            current_power=round(current_power, 2),
            is_shiftable=is_shiftable,
            shiftable_power=round(shiftable_power, 2),
            sharp_energy_ratio=round(sharp_ratio * 100, 1),
            peak_energy_ratio=round(peak_ratio * 100, 1),
            flat_energy_ratio=round(flat_ratio * 100, 1),
            valley_energy_ratio=round(valley_ratio * 100, 1),
            deep_valley_energy_ratio=round(deep_valley_ratio * 100, 1),
            shift_potential_saving=round(shift_saving, 2),
            allowed_shift_hours=config.allowed_shift_hours if config and config.allowed_shift_hours else [0, 1, 2, 3, 4, 5, 6],
            forbidden_shift_hours=config.forbidden_shift_hours if config and config.forbidden_shift_hours else [10, 11, 14, 15],
            is_critical=is_critical
        ))

    # 生成建议
    recommendations = []
    # 按节省潜力排序
    sorted_devices = sorted(device_potentials, key=lambda x: x.shift_potential_saving, reverse=True)
    top_devices = [d for d in sorted_devices if d.is_shiftable][:5]

    for d in top_devices:
        if d.shift_potential_saving > 100:
            recommendations.append(
                f"建议将 {d.device_name} 的部分负荷({d.shiftable_power:.1f}kW)从峰时转移到谷时运行，"
                f"预计每月可节省 {d.shift_potential_saving:.0f} 元"
            )

    return ResponseModel(data=DeviceShiftAnalysisResult(
        analysis_time=datetime.now(),
        total_devices=len(devices),
        shiftable_devices=shiftable_count,
        total_shiftable_power=round(total_shiftable_power, 2),
        total_potential_saving=round(total_potential_saving, 2),
        devices=device_potentials,
        recommendations=recommendations
    ))


# ==================== 需量配置分析 ====================

@router.get("/analysis/demand-config", response_model=ResponseModel[DemandConfigAnalysisResult], summary="需量配置合理性分析")
async def analyze_demand_config(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    分析各计量点的需量配置是否合理
    识别申报过高或过低的计量点
    """
    # 获取所有计量点
    meters_result = await db.execute(
        select(MeterPoint).where(MeterPoint.is_enabled == True)
    )
    meters = meters_result.scalars().all()

    import random

    analysis_items = []
    over_declared_count = 0
    under_declared_count = 0
    optimal_count = 0
    total_saving = 0.0

    for meter in meters:
        declared = meter.declared_demand or 500

        # 获取或模拟需量历史数据
        history_result = await db.execute(
            select(DemandHistory)
            .where(DemandHistory.meter_point_id == meter.id)
            .order_by(DemandHistory.stat_year.desc(), DemandHistory.stat_month.desc())
            .limit(12)
        )
        history = history_result.scalars().all()

        if history:
            max_demands = [h.max_demand for h in history if h.max_demand]
            avg_demands = [h.avg_demand for h in history if h.avg_demand]
            max_demand_12m = max(max_demands) if max_demands else declared * 0.8
            avg_demand_12m = sum(avg_demands) / len(avg_demands) if avg_demands else max_demand_12m * 0.7
            demand_95th = sorted(max_demands)[int(len(max_demands) * 0.95)] if len(max_demands) > 1 else max_demand_12m * 0.95
        else:
            # 模拟数据
            max_demand_12m = declared * random.uniform(0.6, 1.1)
            avg_demand_12m = max_demand_12m * random.uniform(0.6, 0.8)
            demand_95th = max_demand_12m * random.uniform(0.9, 0.98)

        utilization = max_demand_12m / declared if declared > 0 else 0

        # 计算最优需量（基于95%分位数，留5%余量）
        optimal_demand = demand_95th * 1.05

        # 判断配置状态
        is_over_declared = utilization < 0.7  # 利用率低于70%视为申报过高
        is_under_declared = max_demand_12m > declared  # 最大需量超过申报值

        # 计算潜在节省
        demand_price = 38  # 元/kW·月
        if is_over_declared:
            saving = (declared - optimal_demand) * demand_price * 12
            over_declared_count += 1
        elif is_under_declared:
            # 超需量罚款估算（假设超量部分按2倍计费）
            over_amount = max_demand_12m - declared
            penalty = over_amount * demand_price * 2 * 3  # 假设一年超3次
            saving = penalty - (optimal_demand - declared) * demand_price * 12
            under_declared_count += 1
        else:
            saving = 0
            optimal_count += 1

        total_saving += max(saving, 0)

        # 生成建议
        if is_over_declared:
            recommendation = f"建议将申报需量从 {declared:.0f}kW 下调至 {optimal_demand:.0f}kW，可节省需量费用"
        elif is_under_declared:
            recommendation = f"建议将申报需量从 {declared:.0f}kW 上调至 {optimal_demand:.0f}kW，避免超需量罚款"
        else:
            recommendation = "当前需量配置合理，无需调整"

        # 计算超需量风险概率
        over_risk = min(100, max(0, (max_demand_12m / declared - 0.9) * 100 * 5)) if declared > 0 else 0

        analysis_items.append(DemandConfigAnalysisItem(
            meter_point_id=meter.id,
            meter_name=meter.meter_name,
            declared_demand=round(declared, 2),
            max_demand_12m=round(max_demand_12m, 2),
            avg_demand_12m=round(avg_demand_12m, 2),
            demand_95th=round(demand_95th, 2),
            utilization_rate=round(utilization * 100, 1),
            optimal_demand=round(optimal_demand, 2),
            is_over_declared=is_over_declared,
            is_under_declared=is_under_declared,
            potential_saving=round(max(saving, 0), 2),
            over_demand_risk=round(over_risk, 1),
            recommendation=recommendation
        ))

    return ResponseModel(data=DemandConfigAnalysisResult(
        analysis_time=datetime.now(),
        total_meter_points=len(meters),
        over_declared_count=over_declared_count,
        under_declared_count=under_declared_count,
        optimal_count=optimal_count,
        total_potential_saving=round(total_saving, 2),
        items=analysis_items
    ))

from ...services.analysis_plugins import plugin_manager, register_all_plugins, SuggestionType, PluginPriority

# 初始化插件
_plugins_registered = False


def ensure_plugins_registered():
    """确保插件已注册"""
    global _plugins_registered
    if not _plugins_registered:
        register_all_plugins()
        _plugins_registered = True


@router.get("/analysis/plugins", response_model=ResponseModel, summary="获取分析插件列表")
async def get_analysis_plugins(
    current_user: User = Depends(get_current_user)
):
    """获取所有可用的分析插件"""
    ensure_plugins_registered()
    plugins_info = plugin_manager.get_plugin_info()
    return ResponseModel(data=plugins_info)


@router.post("/analysis/plugins/{plugin_id}/enable", response_model=ResponseModel, summary="启用插件")
async def enable_analysis_plugin(
    plugin_id: str,
    current_user: User = Depends(require_admin)
):
    """启用指定分析插件"""
    ensure_plugins_registered()
    if plugin_manager.enable_plugin(plugin_id):
        return ResponseModel(data={"message": f"插件 {plugin_id} 已启用"})
    raise HTTPException(status_code=404, detail=f"插件 {plugin_id} 不存在")


@router.post("/analysis/plugins/{plugin_id}/disable", response_model=ResponseModel, summary="禁用插件")
async def disable_analysis_plugin(
    plugin_id: str,
    current_user: User = Depends(require_admin)
):
    """禁用指定分析插件"""
    ensure_plugins_registered()
    if plugin_manager.disable_plugin(plugin_id):
        return ResponseModel(data={"message": f"插件 {plugin_id} 已禁用"})
    raise HTTPException(status_code=404, detail=f"插件 {plugin_id} 不存在")


@router.post("/analysis/run", response_model=ResponseModel, summary="执行节能分析")
async def run_energy_analysis(
    plugin_ids: Optional[List[str]] = Query(None, description="要执行的插件ID列表，为空则执行所有启用的插件"),
    days: int = Query(30, description="分析数据天数", ge=7, le=365),
    save_results: bool = Query(True, description="是否保存分析结果到数据库"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """
    执行节能分析

    分析系统运行数据，生成节能建议。

    - **plugin_ids**: 指定要执行的插件，为空则执行所有启用的插件
    - **days**: 分析数据范围（天数）
    - **save_results**: 是否将建议保存到数据库
    """
    ensure_plugins_registered()

    try:
        results = await plugin_manager.run_analysis(
            db=db,
            plugin_ids=plugin_ids,
            days=days,
            save_results=save_results
        )

        # 转换结果格式
        suggestions = []
        for r in results:
            suggestions.append({
                "suggestion_type": r.suggestion_type.value,
                "priority": r.priority.name,
                "title": r.title,
                "description": r.description,
                "detail": r.detail,
                "estimated_saving": r.estimated_saving,
                "estimated_cost_saving": r.estimated_cost_saving,
                "implementation_difficulty": r.implementation_difficulty,
                "payback_period": r.payback_period,
                "related_devices": r.related_devices,
                "confidence": r.confidence,
                "created_at": r.created_at.isoformat()
            })

        return ResponseModel(data={
            "total": len(suggestions),
            "suggestions": suggestions
        })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"分析执行失败: {str(e)}")


@router.post("/analysis/run/{plugin_id}", response_model=ResponseModel, summary="执行单个分析插件")
async def run_single_analysis(
    plugin_id: str,
    days: int = Query(30, description="分析数据天数", ge=7, le=365),
    save_results: bool = Query(True, description="是否保存结果"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_operator)
):
    """执行单个分析插件"""
    ensure_plugins_registered()

    plugin = plugin_manager.get_plugin(plugin_id)
    if not plugin:
        raise HTTPException(status_code=404, detail=f"插件 {plugin_id} 不存在")

    try:
        results = await plugin_manager.run_single_plugin(
            db=db,
            plugin_id=plugin_id,
            days=days,
            save_results=save_results
        )

        suggestions = []
        for r in results:
            suggestions.append({
                "suggestion_type": r.suggestion_type.value,
                "priority": r.priority.name,
                "title": r.title,
                "description": r.description,
                "detail": r.detail,
                "estimated_saving": r.estimated_saving,
                "estimated_cost_saving": r.estimated_cost_saving,
                "implementation_difficulty": r.implementation_difficulty,
                "payback_period": r.payback_period,
                "related_devices": r.related_devices,
                "confidence": r.confidence,
                "created_at": r.created_at.isoformat()
            })

        return ResponseModel(data={
            "plugin_id": plugin_id,
            "plugin_name": plugin.plugin_name,
            "total": len(suggestions),
            "suggestions": suggestions
        })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"分析执行失败: {str(e)}")


@router.get("/analysis/summary", response_model=ResponseModel, summary="获取分析汇总")
async def get_analysis_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取节能分析汇总

    统计各类型建议数量、潜在节能量等
    """
    # 统计各状态建议数量
    status_result = await db.execute(
        select(
            EnergySuggestion.status,
            func.count(EnergySuggestion.id).label('count')
        ).group_by(EnergySuggestion.status)
    )
    status_counts = {row[0]: row[1] for row in status_result.all()}

    # 统计各类型建议数量 (按category分组)
    type_result = await db.execute(
        select(
            EnergySuggestion.category,
            func.count(EnergySuggestion.id).label('count')
        ).group_by(EnergySuggestion.category)
    )
    type_counts = {row[0] or 'unknown': row[1] for row in type_result.all()}

    # 统计潜在节能量
    saving_result = await db.execute(
        select(
            func.sum(EnergySuggestion.potential_saving).label('total_saving'),
            func.sum(EnergySuggestion.potential_cost_saving).label('total_cost_saving')
        ).where(EnergySuggestion.status == 'pending')
    )
    saving_row = saving_result.one()

    # 统计已实现节能量
    completed_result = await db.execute(
        select(
            func.sum(EnergySuggestion.actual_saving).label('actual_saving')
        ).where(EnergySuggestion.status == 'completed')
    )
    completed_row = completed_result.one()

    # 获取插件状态
    ensure_plugins_registered()
    plugins = plugin_manager.get_plugin_info()

    return ResponseModel(data={
        "status_summary": {
            "pending": status_counts.get('pending', 0),
            "accepted": status_counts.get('accepted', 0),
            "rejected": status_counts.get('rejected', 0),
            "completed": status_counts.get('completed', 0)
        },
        "type_summary": type_counts,
        "potential_saving": {
            "energy_kwh": saving_row.total_saving or 0,
            "cost_yuan": saving_row.total_cost_saving or 0
        },
        "actual_saving": {
            "energy_kwh": completed_row.actual_saving or 0,
            "cost_yuan": 0  # No actual_cost_saving field in model
        },
        "plugins": {
            "total": len(plugins),
            "enabled": len([p for p in plugins if p['enabled']])
        }
    })


# ==================== V2.3 需量分析增强 ====================

@router.get("/demand/15min-curve", response_model=ResponseModel, summary="获取15分钟需量曲线")
async def get_demand_15min_curve(
    meter_point_id: int = Query(..., description="计量点ID"),
    date: Optional[date] = Query(None, description="日期，默认今天"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    获取15分钟滑动窗口需量曲线

    返回指定日期的96个15分钟需量数据点
    """
    from ...models.energy import Demand15MinData

    target_date = date or datetime.now().date()
    start_time = datetime.combine(target_date, datetime.min.time())
    end_time = start_time + timedelta(days=1)

    # 查询15分钟数据
    result = await db.execute(
        select(Demand15MinData).where(
            Demand15MinData.meter_point_id == meter_point_id,
            Demand15MinData.timestamp >= start_time,
            Demand15MinData.timestamp < end_time
        ).order_by(Demand15MinData.timestamp)
    )
    records = result.scalars().all()

    # 获取计量点信息
    meter_result = await db.execute(
        select(MeterPoint).where(MeterPoint.id == meter_point_id)
    )
    meter = meter_result.scalar_one_or_none()
    declared_demand = meter.declared_demand if meter else None

    data_points = []
    max_demand = 0
    over_declared_count = 0

    for r in records:
        data_points.append({
            "timestamp": r.timestamp.isoformat(),
            "average_power": r.average_power,
            "rolling_demand": r.rolling_demand,
            "is_over_declared": r.is_over_declared
        })
        if r.rolling_demand and r.rolling_demand > max_demand:
            max_demand = r.rolling_demand
        if r.is_over_declared:
            over_declared_count += 1

    return ResponseModel(data={
        "meter_point_id": meter_point_id,
        "date": target_date.isoformat(),
        "declared_demand": declared_demand,
        "max_demand": max_demand,
        "over_declared_count": over_declared_count,
        "data_points": data_points,
        "total_points": len(data_points)
    })


@router.get("/demand/peak-analysis", response_model=ResponseModel, summary="需量峰值分析")
async def get_demand_peak_analysis(
    meter_point_id: int = Query(..., description="计量点ID"),
    days: int = Query(30, description="分析天数"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    分析需量峰值分布

    统计峰值出现的时段分布、超需量次数等
    """
    from ...models.energy import Demand15MinData, DemandAnalysisRecord

    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)

    # 获取计量点信息
    meter_result = await db.execute(
        select(MeterPoint).where(MeterPoint.id == meter_point_id)
    )
    meter = meter_result.scalar_one_or_none()
    if not meter:
        raise HTTPException(status_code=404, detail="计量点不存在")

    declared_demand = meter.declared_demand or 100

    # 查询历史需量数据
    result = await db.execute(
        select(Demand15MinData).where(
            Demand15MinData.meter_point_id == meter_point_id,
            Demand15MinData.timestamp >= start_date,
            Demand15MinData.timestamp <= end_date
        ).order_by(Demand15MinData.timestamp)
    )
    records = result.scalars().all()

    # 统计分析
    hourly_peaks = {i: [] for i in range(24)}
    over_declared_records = []
    all_demands = []

    for r in records:
        if r.rolling_demand:
            hour = r.timestamp.hour
            hourly_peaks[hour].append(r.rolling_demand)
            all_demands.append(r.rolling_demand)
            if r.is_over_declared:
                over_declared_records.append({
                    "timestamp": r.timestamp.isoformat(),
                    "demand": r.rolling_demand,
                    "over_ratio": (r.rolling_demand - declared_demand) / declared_demand * 100
                })

    # 计算各时段平均峰值
    hourly_avg = {}
    peak_hours = []
    for hour, demands in hourly_peaks.items():
        if demands:
            avg = sum(demands) / len(demands)
            hourly_avg[f"{hour:02d}:00"] = round(avg, 2)
            if avg > declared_demand * 0.9:
                peak_hours.append(hour)

    # 计算统计指标
    max_demand = max(all_demands) if all_demands else 0
    avg_demand = sum(all_demands) / len(all_demands) if all_demands else 0
    utilization = (max_demand / declared_demand * 100) if declared_demand > 0 else 0

    return ResponseModel(data={
        "meter_point_id": meter_point_id,
        "meter_name": meter.meter_name,
        "declared_demand": declared_demand,
        "analysis_period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
            "days": days
        },
        "statistics": {
            "max_demand": round(max_demand, 2),
            "avg_demand": round(avg_demand, 2),
            "utilization_rate": round(utilization, 1),
            "over_declared_count": len(over_declared_records),
            "over_declared_ratio": round(len(over_declared_records) / len(all_demands) * 100, 2) if all_demands else 0
        },
        "hourly_distribution": hourly_avg,
        "peak_hours": peak_hours,
        "over_declared_records": over_declared_records[:20]  # 最近20条超需量记录
    })


@router.get("/demand/optimization-plan", response_model=ResponseModel, summary="需量优化方案")
async def get_demand_optimization_plan(
    meter_point_id: int = Query(..., description="计量点ID"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    生成需量优化方案

    基于历史数据分析，给出申报需量调整建议
    """
    from ...models.energy import Demand15MinData

    # 获取计量点信息
    meter_result = await db.execute(
        select(MeterPoint).where(MeterPoint.id == meter_point_id)
    )
    meter = meter_result.scalar_one_or_none()
    if not meter:
        raise HTTPException(status_code=404, detail="计量点不存在")

    declared_demand = meter.declared_demand or 100

    # 获取最近30天数据
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)

    result = await db.execute(
        select(Demand15MinData.rolling_demand).where(
            Demand15MinData.meter_point_id == meter_point_id,
            Demand15MinData.timestamp >= start_date,
            Demand15MinData.rolling_demand.isnot(None)
        )
    )
    demands = [r[0] for r in result.all()]

    if not demands:
        return ResponseModel(data={
            "meter_point_id": meter_point_id,
            "message": "数据不足，无法生成优化方案"
        })

    # 计算统计指标
    max_demand = max(demands)
    avg_demand = sum(demands) / len(demands)
    p95_demand = sorted(demands)[int(len(demands) * 0.95)]  # 95分位数

    # 计算利用率
    utilization = max_demand / declared_demand * 100 if declared_demand > 0 else 0

    # 生成优化建议
    recommendations = []
    recommended_demand = declared_demand
    annual_saving = 0

    if utilization < 70:
        # 申报过高
        recommended_demand = round(p95_demand * 1.1, 0)  # 95分位数 + 10%余量
        fee_diff = (declared_demand - recommended_demand) * 30  # 假设30元/kW·月
        annual_saving = fee_diff * 12
        recommendations.append({
            "type": "reduce_declared",
            "title": "建议降低申报需量",
            "description": f"当前申报{declared_demand}kW，实际最大需量{max_demand:.0f}kW，利用率仅{utilization:.1f}%",
            "action": f"建议将申报需量调整为{recommended_demand:.0f}kW",
            "saving": f"年节省容量电费约{annual_saving:.0f}元"
        })
    elif utilization > 95:
        # 申报偏低，有超需量风险
        recommended_demand = round(max_demand * 1.1, 0)
        penalty_risk = (max_demand - declared_demand) * 2 * 30  # 超需量罚款风险
        recommendations.append({
            "type": "increase_declared",
            "title": "建议提高申报需量",
            "description": f"当前申报{declared_demand}kW，实际最大需量{max_demand:.0f}kW，存在超需量风险",
            "action": f"建议将申报需量调整为{recommended_demand:.0f}kW",
            "saving": f"可避免超需量罚款风险约{penalty_risk:.0f}元/月"
        })
    else:
        recommendations.append({
            "type": "maintain",
            "title": "当前申报需量合理",
            "description": f"当前申报{declared_demand}kW，利用率{utilization:.1f}%，处于合理区间",
            "action": "建议维持当前申报需量",
            "saving": "无需调整"
        })

    # 负荷调节建议
    if max_demand > declared_demand * 0.85:
        recommendations.append({
            "type": "load_shift",
            "title": "建议错峰用电",
            "description": "高峰时段需量接近申报值，建议将部分负荷转移至低谷时段",
            "action": "识别可调节负荷，制定错峰用电计划",
            "saving": "可降低峰值需量10-20%"
        })

    return ResponseModel(data={
        "meter_point_id": meter_point_id,
        "meter_name": meter.meter_name,
        "current_declared": declared_demand,
        "statistics": {
            "max_demand": round(max_demand, 2),
            "avg_demand": round(avg_demand, 2),
            "p95_demand": round(p95_demand, 2),
            "utilization_rate": round(utilization, 1)
        },
        "optimization": {
            "recommended_demand": recommended_demand,
            "annual_saving": annual_saving,
            "recommendations": recommendations
        }
    })


@router.post("/demand/forecast", response_model=ResponseModel, summary="需量预测")
async def forecast_demand(
    meter_point_id: int = Query(..., description="计量点ID"),
    forecast_hours: int = Query(24, description="预测小时数"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    预测未来需量趋势

    基于历史数据预测未来N小时的需量
    """
    from ...models.energy import Demand15MinData

    # 获取计量点信息
    meter_result = await db.execute(
        select(MeterPoint).where(MeterPoint.id == meter_point_id)
    )
    meter = meter_result.scalar_one_or_none()
    if not meter:
        raise HTTPException(status_code=404, detail="计量点不存在")

    declared_demand = meter.declared_demand or 100

    # 获取最近7天同时段数据用于预测
    now = datetime.now()
    forecast_points = []

    for i in range(forecast_hours * 4):  # 每15分钟一个点
        forecast_time = now + timedelta(minutes=15 * i)
        hour = forecast_time.hour
        minute = (forecast_time.minute // 15) * 15

        # 查询历史同时段数据
        historical_demands = []
        for day_offset in range(1, 8):  # 过去7天
            hist_time = forecast_time - timedelta(days=day_offset)
            start = hist_time - timedelta(minutes=7)
            end = hist_time + timedelta(minutes=7)

            result = await db.execute(
                select(Demand15MinData.rolling_demand).where(
                    Demand15MinData.meter_point_id == meter_point_id,
                    Demand15MinData.timestamp >= start,
                    Demand15MinData.timestamp <= end,
                    Demand15MinData.rolling_demand.isnot(None)
                )
            )
            for r in result.all():
                if r[0]:
                    historical_demands.append(r[0])

        # 简单平均预测
        if historical_demands:
            predicted = sum(historical_demands) / len(historical_demands)
            confidence = min(0.9, 0.5 + len(historical_demands) * 0.05)
        else:
            predicted = declared_demand * 0.7  # 默认70%利用率
            confidence = 0.3

        forecast_points.append({
            "timestamp": forecast_time.isoformat(),
            "predicted_demand": round(predicted, 2),
            "confidence": round(confidence, 2),
            "is_peak_risk": predicted > declared_demand * 0.9
        })

    # 统计预测结果
    peak_risk_count = sum(1 for p in forecast_points if p["is_peak_risk"])
    max_predicted = max(p["predicted_demand"] for p in forecast_points)

    return ResponseModel(data={
        "meter_point_id": meter_point_id,
        "meter_name": meter.meter_name,
        "declared_demand": declared_demand,
        "forecast_period": {
            "start": now.isoformat(),
            "hours": forecast_hours,
            "points": len(forecast_points)
        },
        "summary": {
            "max_predicted": max_predicted,
            "peak_risk_count": peak_risk_count,
            "peak_risk_ratio": round(peak_risk_count / len(forecast_points) * 100, 1)
        },
        "forecast_points": forecast_points
    })


