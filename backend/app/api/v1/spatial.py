"""
空间拓扑管理 API
"""
import json
from io import BytesIO
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError
from starlette.responses import StreamingResponse
import openpyxl

# Excel 导入文件大小限制 (10MB)
MAX_IMPORT_SIZE = 10 * 1024 * 1024

from ..deps import get_db, require_operator, require_viewer, require_site_access, get_user_site_ids
from ...models.user import User
from ...models.spatial import Site, Floor, Room, Row, LayoutTemplate
from ...models.asset import Cabinet
from ...models.gateway import Gateway, DataSource, MqttAclRule
from ...models.device import Device
from ...schemas.spatial import (
    SiteCreate, SiteUpdate, SiteResponse,
    FloorCreate, FloorUpdate, FloorResponse,
    RoomCreate, RoomUpdate, RoomResponse,
    RowCreate, RowUpdate, RowResponse,
    LayoutTemplateResponse,
    CabinetPositionUpdate,
    SpatialTreeResponse,
    ImportResultResponse,
    TemplateApplyRequest, TemplateApplyResponse,
)

router = APIRouter(prefix="/spatial", tags=["空间拓扑"])


# ==================== 预置模板数据 ====================

PRESET_TEMPLATES = [
    {
        "template_code": "2n_cold_aisle",
        "template_name": "2N冷通道",
        "description": "双排面对面，中间冷通道",
        "template_data": json.dumps({
            "name": "2N冷通道",
            "rows": [
                {"row_code": "R1", "aisle_type": "cold", "cabinets": 10},
                {"row_code": "R2", "aisle_type": "cold", "cabinets": 10},
            ],
            "description": "双排面对面，中间冷通道",
        }),
    },
    {
        "template_code": "single_row",
        "template_name": "单排布局",
        "description": "单排机柜布局",
        "template_data": json.dumps({
            "name": "单排布局",
            "rows": [
                {"row_code": "R1", "aisle_type": "none", "cabinets": 10},
            ],
            "description": "单排机柜布局",
        }),
    },
    {
        "template_code": "double_row",
        "template_name": "双排背靠背",
        "description": "双排背靠背，中间热通道",
        "template_data": json.dumps({
            "name": "双排背靠背",
            "rows": [
                {"row_code": "R1", "aisle_type": "hot", "cabinets": 10},
                {"row_code": "R2", "aisle_type": "hot", "cabinets": 10},
            ],
            "description": "双排背靠背，中间热通道",
        }),
    },
]


async def _ensure_preset_templates(db: AsyncSession):
    """确保预置模板存在"""
    result = await db.execute(select(func.count(LayoutTemplate.id)))
    count = result.scalar() or 0
    if count == 0:
        try:
            for tpl in PRESET_TEMPLATES:
                db.add(LayoutTemplate(**tpl))
            await db.commit()
        except IntegrityError:
            await db.rollback()  # 其他请求已插入，忽略


# ==================== Site CRUD ====================

@router.get("/sites", response_model=List[SiteResponse])
async def list_sites(
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    status: Optional[str] = Query(None, description="状态: active/inactive/maintenance"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
    site_ids: Optional[List[int]] = Depends(get_user_site_ids),
):
    """获取站点列表（含网关/设备统计，非admin仅返回授权站点）"""
    stmt = select(Site)
    if site_ids is not None:
        stmt = stmt.where(Site.id.in_(site_ids))
    if keyword:
        stmt = stmt.where(
            Site.site_name.contains(keyword) | Site.site_code.contains(keyword)
        )
    if status:
        stmt = stmt.where(Site.status == status)
    stmt = stmt.order_by(Site.id)
    result = await db.execute(stmt)
    sites = result.scalars().all()

    # 批量查询每个站点的网关和设备数量
    site_ids = [s.id for s in sites]
    gw_counts = {}
    dev_counts = {}
    if site_ids:
        gw_result = await db.execute(
            select(Gateway.site_id, func.count(Gateway.id))
            .where(Gateway.site_id.in_(site_ids))
            .group_by(Gateway.site_id)
        )
        gw_counts = dict(gw_result.all())
        dev_result = await db.execute(
            select(Device.site_id, func.count(Device.id))
            .where(Device.site_id.in_(site_ids))
            .group_by(Device.site_id)
        )
        dev_counts = dict(dev_result.all())

    items = []
    for s in sites:
        resp = SiteResponse.model_validate(s)
        resp.gateway_count = gw_counts.get(s.id, 0)
        resp.device_count = dev_counts.get(s.id, 0)
        items.append(resp)
    return items


@router.post("/sites", response_model=SiteResponse)
async def create_site(
    data: SiteCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """创建站点（自动生成 EMQX ACL 规则）"""
    site = Site(**data.model_dump())
    db.add(site)
    await db.flush()  # 获取 site.id

    # 自动创建 ACL 规则
    from ...services.emqx_acl import emqx_acl_service
    await emqx_acl_service.on_site_created(site.id, site.site_code, db)

    await db.commit()
    await db.refresh(site)

    resp = SiteResponse.model_validate(site)
    resp.gateway_count = 0
    resp.device_count = 0
    return resp


@router.put("/sites/{site_id}", response_model=SiteResponse)
async def update_site(
    data: SiteUpdate,
    site_id: int = Depends(require_site_access),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新站点"""
    result = await db.execute(select(Site).where(Site.id == site_id))
    site = result.scalar_one_or_none()
    if not site:
        raise HTTPException(status_code=404, detail="站点不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(site, k, v)
    site.updated_at = datetime.now()
    await db.commit()
    await db.refresh(site)

    # 返回含统计信息的响应
    gw_cnt = await db.execute(select(func.count(Gateway.id)).where(Gateway.site_id == site_id))
    dev_cnt = await db.execute(select(func.count(Device.id)).where(Device.site_id == site_id))
    resp = SiteResponse.model_validate(site)
    resp.gateway_count = gw_cnt.scalar() or 0
    resp.device_count = dev_cnt.scalar() or 0
    return resp


@router.delete("/sites/{site_id}")
async def delete_site(
    site_id: int = Depends(require_site_access),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """删除站点（检查所有关联数据）"""
    result = await db.execute(select(Site).where(Site.id == site_id))
    site = result.scalar_one_or_none()
    if not site:
        raise HTTPException(status_code=404, detail="站点不存在")

    # 检查关联的楼层
    cnt = await db.execute(select(func.count(Floor.id)).where(Floor.site_id == site_id))
    floor_count = cnt.scalar() or 0

    # 检查关联的网关
    gw_cnt = await db.execute(select(func.count(Gateway.id)).where(Gateway.site_id == site_id))
    gw_count = gw_cnt.scalar() or 0

    # 检查关联的设备
    dev_cnt = await db.execute(select(func.count(Device.id)).where(Device.site_id == site_id))
    dev_count = dev_cnt.scalar() or 0

    # 检查关联的数据源
    ds_cnt = await db.execute(select(func.count(DataSource.id)).where(DataSource.site_id == site_id))
    ds_count = ds_cnt.scalar() or 0

    deps = []
    if floor_count > 0:
        deps.append(f"楼层({floor_count})")
    if gw_count > 0:
        deps.append(f"网关({gw_count})")
    if dev_count > 0:
        deps.append(f"设备({dev_count})")
    if ds_count > 0:
        deps.append(f"数据源({ds_count})")

    if deps:
        raise HTTPException(
            status_code=400,
            detail=f"请先删除该站点下的关联数据: {', '.join(deps)}"
        )

    # 清理 ACL 规则
    await db.execute(delete(MqttAclRule).where(MqttAclRule.site_id == site_id))
    await db.delete(site)
    await db.commit()
    return {"detail": "删除成功"}


VALID_SITE_STATUSES = {"active", "inactive", "maintenance"}


@router.put("/sites/{site_id}/status")
async def update_site_status(
    site_id: int = Depends(require_site_access),
    status: str = Query(..., description="目标状态: active/inactive/maintenance"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新站点状态"""
    if status not in VALID_SITE_STATUSES:
        raise HTTPException(
            status_code=400,
            detail=f"无效状态，可选值: {', '.join(sorted(VALID_SITE_STATUSES))}"
        )
    result = await db.execute(select(Site).where(Site.id == site_id))
    site = result.scalar_one_or_none()
    if not site:
        raise HTTPException(status_code=404, detail="站点不存在")
    old_status = site.status
    site.status = status
    site.updated_at = datetime.now()
    await db.commit()
    return {"detail": f"站点状态已从 {old_status} 更新为 {status}"}


@router.get("/sites/{site_id}/acl-rules")
async def get_site_acl_rules(
    site_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取站点的 MQTT ACL 规则"""
    from ...services.emqx_acl import emqx_acl_service
    rules = await emqx_acl_service.get_site_rules(site_id, db)
    return [
        {
            "id": r.id,
            "site_id": r.site_id,
            "client_id_pattern": r.client_id_pattern,
            "topic_pattern": r.topic_pattern,
            "action": r.action,
            "permission": r.permission,
            "description": r.description,
        }
        for r in rules
    ]


# ==================== Floor CRUD ====================

@router.get("/floors", response_model=List[FloorResponse])
async def list_floors(
    site_id: Optional[int] = Query(None, description="站点ID"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取楼层列表"""
    stmt = select(Floor)
    if site_id is not None:
        stmt = stmt.where(Floor.site_id == site_id)
    stmt = stmt.order_by(Floor.sort_order, Floor.id)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/floors", response_model=FloorResponse)
async def create_floor(
    data: FloorCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """创建楼层"""
    floor = Floor(**data.model_dump())
    db.add(floor)
    await db.commit()
    await db.refresh(floor)
    return floor


@router.put("/floors/{floor_id}", response_model=FloorResponse)
async def update_floor(
    floor_id: int,
    data: FloorUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新楼层"""
    result = await db.execute(select(Floor).where(Floor.id == floor_id))
    floor = result.scalar_one_or_none()
    if not floor:
        raise HTTPException(status_code=404, detail="楼层不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(floor, k, v)
    floor.updated_at = datetime.now()
    await db.commit()
    await db.refresh(floor)
    return floor


@router.delete("/floors/{floor_id}")
async def delete_floor(
    floor_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """删除楼层"""
    result = await db.execute(select(Floor).where(Floor.id == floor_id))
    floor = result.scalar_one_or_none()
    if not floor:
        raise HTTPException(status_code=404, detail="楼层不存在")
    cnt = await db.execute(select(func.count(Room.id)).where(Room.floor_id == floor_id))
    if (cnt.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="请先删除该楼层下的所有房间")
    await db.delete(floor)
    await db.commit()
    return {"detail": "删除成功"}


# ==================== Room CRUD ====================

@router.get("/rooms", response_model=List[RoomResponse])
async def list_rooms(
    floor_id: Optional[int] = Query(None, description="楼层ID"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取房间列表"""
    stmt = select(Room)
    if floor_id is not None:
        stmt = stmt.where(Room.floor_id == floor_id)
    stmt = stmt.order_by(Room.id)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/rooms", response_model=RoomResponse)
async def create_room(
    data: RoomCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """创建房间"""
    if data.grid_cols > 50 or data.grid_rows > 50:
        raise HTTPException(status_code=400, detail="网格行列数不能超过50")
    room = Room(**data.model_dump())
    db.add(room)
    await db.commit()
    await db.refresh(room)
    return room


@router.put("/rooms/{room_id}", response_model=RoomResponse)
async def update_room(
    room_id: int,
    data: RoomUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新房间"""
    result = await db.execute(select(Room).where(Room.id == room_id))
    room = result.scalar_one_or_none()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")
    update_data = data.model_dump(exclude_unset=True)
    if "grid_cols" in update_data and update_data["grid_cols"] > 50:
        raise HTTPException(status_code=400, detail="网格列数不能超过50")
    if "grid_rows" in update_data and update_data["grid_rows"] > 50:
        raise HTTPException(status_code=400, detail="网格行数不能超过50")
    for k, v in update_data.items():
        setattr(room, k, v)
    room.updated_at = datetime.now()
    await db.commit()
    await db.refresh(room)
    return room


@router.delete("/rooms/{room_id}")
async def delete_room(
    room_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """删除房间"""
    result = await db.execute(select(Room).where(Room.id == room_id))
    room = result.scalar_one_or_none()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")
    cnt = await db.execute(select(func.count(Row.id)).where(Row.room_id == room_id))
    if (cnt.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="请先删除该房间下的所有行")
    await db.delete(room)
    await db.commit()
    return {"detail": "删除成功"}


# ==================== Row CRUD ====================

@router.get("/rows", response_model=List[RowResponse])
async def list_rows(
    room_id: Optional[int] = Query(None, description="房间ID"),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取行列表"""
    stmt = select(Row)
    if room_id is not None:
        stmt = stmt.where(Row.room_id == room_id)
    stmt = stmt.order_by(Row.sort_order, Row.id)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.post("/rows", response_model=RowResponse)
async def create_row(
    data: RowCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """创建行"""
    row = Row(**data.model_dump())
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return row


@router.put("/rows/{row_id}", response_model=RowResponse)
async def update_row(
    row_id: int,
    data: RowUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新行"""
    result = await db.execute(select(Row).where(Row.id == row_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="行不存在")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(row, k, v)
    row.updated_at = datetime.now()
    await db.commit()
    await db.refresh(row)
    return row


@router.delete("/rows/{row_id}")
async def delete_row(
    row_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """删除行"""
    result = await db.execute(select(Row).where(Row.id == row_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="行不存在")
    cnt = await db.execute(
        select(func.count(Cabinet.id)).where(Cabinet.row_id == row_id)
    )
    if (cnt.scalar() or 0) > 0:
        raise HTTPException(status_code=400, detail="请先移除该行下的所有机柜")
    await db.delete(row)
    await db.commit()
    return {"detail": "删除成功"}


# ==================== Tree ====================

@router.get("/tree", response_model=List[SpatialTreeResponse])
async def get_spatial_tree(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取完整空间拓扑树"""
    stmt = (
        select(Site)
        .options(
            selectinload(Site.floors)
            .selectinload(Floor.rooms)
            .selectinload(Room.rows)
            .selectinload(Row.cabinets)
        )
        .order_by(Site.id)
    )
    result = await db.execute(stmt)
    return result.scalars().unique().all()


# ==================== Cabinet Position ====================

@router.put("/cabinets/{cabinet_id}/position")
async def update_cabinet_position(
    cabinet_id: int,
    data: CabinetPositionUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """更新机柜空间位置"""
    result = await db.execute(select(Cabinet).where(Cabinet.id == cabinet_id))
    cabinet = result.scalar_one_or_none()
    if not cabinet:
        raise HTTPException(status_code=404, detail="机柜不存在")

    # H1: 网格边界校验
    target_row_id = data.row_id if data.row_id is not None else cabinet.row_id
    if target_row_id and (data.grid_x is not None or data.grid_y is not None):
        row_result = await db.execute(select(Row).where(Row.id == target_row_id))
        row_obj = row_result.scalar_one_or_none()
        if row_obj:
            room_result = await db.execute(select(Room).where(Room.id == row_obj.room_id))
            room_obj = room_result.scalar_one_or_none()
            if room_obj:
                if data.grid_x is not None and data.grid_x >= room_obj.grid_cols:
                    raise HTTPException(status_code=400, detail="grid_x 超出网格范围")
                if data.grid_y is not None and data.grid_y >= room_obj.grid_rows:
                    raise HTTPException(status_code=400, detail="grid_y 超出网格范围")

                # H2: 坐标唯一性校验
                check_x = data.grid_x if data.grid_x is not None else cabinet.grid_x
                check_y = data.grid_y if data.grid_y is not None else cabinet.grid_y
                if check_x is not None and check_y is not None:
                    room_row_ids = select(Row.id).where(Row.room_id == room_obj.id)
                    conflict = await db.execute(
                        select(Cabinet).where(
                            Cabinet.row_id.in_(room_row_ids),
                            Cabinet.grid_x == check_x,
                            Cabinet.grid_y == check_y,
                            Cabinet.id != cabinet_id,
                        )
                    )
                    if conflict.scalar_one_or_none():
                        raise HTTPException(status_code=400, detail="该网格位置已被其他机柜占用")

    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(cabinet, k, v)
    cabinet.updated_at = datetime.now()
    await db.commit()
    await db.refresh(cabinet)
    return {"detail": "更新成功"}


# ==================== Excel Import/Export ====================

@router.post("/import", response_model=ImportResultResponse)
async def import_spatial(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """Excel导入空间拓扑"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx)")

    content = await file.read()
    if len(content) > MAX_IMPORT_SIZE:
        raise HTTPException(status_code=400, detail="文件大小不能超过10MB")
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel文件为空")

    # 读取表头映射
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {}
    header_mapping = {
        "站点编码": "site_code", "楼层编码": "floor_code",
        "房间编码": "room_code", "行编码": "row_code",
        "通道类型": "aisle_type", "机柜编码": "cabinet_code",
        "机柜名称": "cabinet_name", "列号": "column_number",
        "总U数": "total_u", "最大功率": "max_power", "最大承重": "max_weight",
    }
    for idx, h in enumerate(headers):
        if h in header_mapping:
            col_map[header_mapping[h]] = idx

    total = 0
    success = 0
    failed = 0
    skipped = 0
    errors: List[str] = []

    # 阶段1：收集所有行数据
    rows_data = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(excel_row)
        if not any(row_list):
            continue
        rows_data.append(row_list)
    total = len(rows_data)

    if total == 0:
        wb.close()
        return ImportResultResponse(total=0, success=0, failed=0, skipped=0, errors=[])

    def _get_val(rd, key):
        idx = col_map.get(key)
        if idx is not None and idx < len(rd):
            v = rd[idx]
            return str(v).strip() if v else ""
        return ""

    # 收集去重集合
    site_codes = set()
    floor_keys = set()
    room_keys = set()
    row_keys = set()
    for rd in rows_data:
        sc = _get_val(rd, "site_code")
        fc = _get_val(rd, "floor_code")
        rc = _get_val(rd, "room_code")
        rwc = _get_val(rd, "row_code")
        if sc:
            site_codes.add(sc)
        if sc and fc:
            floor_keys.add((sc, fc))
        if sc and fc and rc:
            room_keys.add((sc, fc, rc))
        if sc and fc and rc and rwc:
            row_keys.add((sc, fc, rc, rwc))

    # 阶段2：查询已存在实体，创建缺失的
    existing_sites = {}
    if site_codes:
        res = await db.execute(select(Site).where(Site.site_code.in_(site_codes)))
        for s in res.scalars().all():
            existing_sites[s.site_code] = s
    for sc in site_codes:
        if sc not in existing_sites:
            site_obj = Site(site_code=sc, site_name=sc)
            db.add(site_obj)
            await db.commit()
            existing_sites[sc] = site_obj

    existing_floors = {}
    for s_code, f_code in floor_keys:
        site_obj = existing_sites.get(s_code)
        if not site_obj:
            continue
        res = await db.execute(
            select(Floor).where(Floor.site_id == site_obj.id, Floor.floor_code == f_code)
        )
        f = res.scalar_one_or_none()
        if f:
            existing_floors[(s_code, f_code)] = f
        else:
            f = Floor(floor_code=f_code, floor_name=f_code, site_id=site_obj.id)
            db.add(f)
            await db.commit()
            existing_floors[(s_code, f_code)] = f

    existing_rooms = {}
    for s_code, f_code, r_code in room_keys:
        floor_obj = existing_floors.get((s_code, f_code))
        if not floor_obj:
            continue
        res = await db.execute(
            select(Room).where(Room.floor_id == floor_obj.id, Room.room_code == r_code)
        )
        r = res.scalar_one_or_none()
        if r:
            existing_rooms[(s_code, f_code, r_code)] = r
        else:
            r = Room(room_code=r_code, room_name=r_code, floor_id=floor_obj.id)
            db.add(r)
            await db.commit()
            existing_rooms[(s_code, f_code, r_code)] = r

    existing_rows = {}
    for s_code, f_code, r_code, rw_code in row_keys:
        room_obj = existing_rooms.get((s_code, f_code, r_code))
        if not room_obj:
            continue
        res = await db.execute(
            select(Row).where(Row.room_id == room_obj.id, Row.row_code == rw_code)
        )
        rw = res.scalar_one_or_none()
        if rw:
            existing_rows[(s_code, f_code, r_code, rw_code)] = rw
        else:
            rw = Row(row_code=rw_code, row_name=rw_code, room_id=room_obj.id)
            db.add(rw)
            await db.commit()
            existing_rows[(s_code, f_code, r_code, rw_code)] = rw

    # 阶段3：逐行处理机柜
    for rd in rows_data:
        cab_code = _get_val(rd, "cabinet_code")
        if not cab_code:
            skipped += 1
            continue

        sc = _get_val(rd, "site_code")
        fc = _get_val(rd, "floor_code")
        rc = _get_val(rd, "room_code")
        rwc = _get_val(rd, "row_code")
        row_obj = existing_rows.get((sc, fc, rc, rwc))

        res = await db.execute(select(Cabinet).where(Cabinet.cabinet_code == cab_code))
        cab = res.scalar_one_or_none()
        if cab:
            if row_obj:
                cab.row_id = row_obj.id
            at = _get_val(rd, "aisle_type")
            if at:
                cab.aisle_type = at
            skipped += 1
        else:
            cab_name = _get_val(rd, "cabinet_name") or cab_code
            total_u_str = _get_val(rd, "total_u")
            total_u = int(total_u_str) if total_u_str.isdigit() else 42
            max_power_str = _get_val(rd, "max_power")
            try:
                max_power = float(max_power_str) if max_power_str else None
            except (ValueError, TypeError):
                max_power = None
            max_weight_str = _get_val(rd, "max_weight")
            try:
                max_weight = float(max_weight_str) if max_weight_str else None
            except (ValueError, TypeError):
                max_weight = None
            col_num = _get_val(rd, "column_number") or None
            at = _get_val(rd, "aisle_type") or None

            cab = Cabinet(
                cabinet_code=cab_code,
                cabinet_name=cab_name,
                total_u=total_u,
                max_power=max_power,
                max_weight=max_weight,
                column_number=col_num,
                row_id=row_obj.id if row_obj else None,
                aisle_type=at,
            )
            db.add(cab)
            success += 1

    await db.commit()
    wb.close()
    return ImportResultResponse(
        total=total, success=success, failed=failed, skipped=skipped, errors=errors
    )


@router.get("/export")
async def export_spatial(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """导出空间拓扑Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "空间拓扑"
    headers = [
        "站点编码", "楼层编码", "房间编码", "行编码",
        "通道类型", "机柜编码", "机柜名称", "列号",
        "总U数", "最大功率", "最大承重",
    ]
    ws.append(headers)

    stmt = (
        select(Site)
        .options(
            selectinload(Site.floors)
            .selectinload(Floor.rooms)
            .selectinload(Room.rows)
            .selectinload(Row.cabinets)
        )
    )
    result = await db.execute(stmt)
    sites = result.scalars().unique().all()

    for site in sites:
        for floor_obj in site.floors:
            for room_obj in floor_obj.rooms:
                for row_obj in room_obj.rows:
                    if row_obj.cabinets:
                        for cab in row_obj.cabinets:
                            ws.append([
                                site.site_code, floor_obj.floor_code,
                                room_obj.room_code, row_obj.row_code,
                                row_obj.aisle_type, cab.cabinet_code,
                                cab.cabinet_name, cab.column_number,
                                cab.total_u, cab.max_power, cab.max_weight,
                            ])
                    else:
                        ws.append([
                            site.site_code, floor_obj.floor_code,
                            room_obj.room_code, row_obj.row_code,
                            row_obj.aisle_type, "", "", "", "", "", "",
                        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=spatial_topology.xlsx"},
    )


# ==================== Templates ====================

@router.get("/templates", response_model=List[LayoutTemplateResponse])
async def list_templates(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_viewer),
):
    """获取布局模板列表"""
    await _ensure_preset_templates(db)
    result = await db.execute(select(LayoutTemplate).order_by(LayoutTemplate.id))
    return result.scalars().all()


@router.post("/templates/{template_id}/apply", response_model=TemplateApplyResponse)
async def apply_template(
    template_id: int,
    data: TemplateApplyRequest,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_operator),
):
    """应用模板到房间"""
    result = await db.execute(select(LayoutTemplate).where(LayoutTemplate.id == template_id))
    template = result.scalar_one_or_none()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    result = await db.execute(select(Room).where(Room.id == data.room_id))
    room = result.scalar_one_or_none()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")

    tpl_data = json.loads(template.template_data)
    tpl_rows = tpl_data.get("rows", [])

    created_rows = 0
    created_cabinets = 0
    skipped_cabinets = 0
    apply_errors: List[str] = []

    for tpl_row in tpl_rows:
        row_code = tpl_row["row_code"]
        aisle_type = tpl_row.get("aisle_type", "none")
        cab_count = tpl_row.get("cabinets", 0)

        res = await db.execute(
            select(Row).where(Row.room_id == room.id, Row.row_code == row_code)
        )
        row_obj = res.scalar_one_or_none()
        if not row_obj:
            row_obj = Row(
                row_code=row_code,
                row_name=row_code,
                room_id=room.id,
                aisle_type=aisle_type,
            )
            db.add(row_obj)
            await db.commit()
            created_rows += 1

        prefix = data.cabinet_prefix or room.room_code
        for seq in range(1, cab_count + 1):
            cab_code = f"{prefix}-{row_code}-C{seq:02d}"
            res = await db.execute(
                select(Cabinet).where(Cabinet.cabinet_code == cab_code)
            )
            if res.scalar_one_or_none():
                skipped_cabinets += 1
                continue
            cab = Cabinet(
                cabinet_code=cab_code,
                cabinet_name=cab_code,
                total_u=42,
                row_id=row_obj.id,
                aisle_type=aisle_type,
            )
            db.add(cab)
            created_cabinets += 1

    await db.commit()
    return TemplateApplyResponse(
        created_rows=created_rows,
        created_cabinets=created_cabinets,
        skipped_cabinets=skipped_cabinets,
        errors=apply_errors,
    )
