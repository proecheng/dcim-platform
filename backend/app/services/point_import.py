"""点位批量导入与预校验服务"""

import io
import json
import logging
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from ..models.gateway import DataSourcePoint

logger = logging.getLogger(__name__)

VALID_DATA_TYPES = {
    "int16",
    "uint16",
    "int32",
    "uint32",
    "float32",
    "float64",
    "bool",
    "string",
}

REQUIRED_COLUMNS = {"address", "data_type"}
OPTIONAL_COLUMNS = {"scale", "offset", "enum_mapping", "is_dry_contact"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS


def parse_excel(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """解析 Excel 文件，返回 (列名列表, 行数据列表)"""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    # 第一行为表头
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    data = []
    for row_idx, row in enumerate(rows[1:], start=2):
        row_dict = {}
        for col_idx, header in enumerate(headers):
            if header in ALL_COLUMNS and col_idx < len(row):
                row_dict[header] = row[col_idx]
        if any(v is not None and v != "" for v in row_dict.values()):
            row_dict["_row"] = row_idx
            data.append(row_dict)

    return headers, data


async def validate_points(
    file_bytes: bytes,
    datasource_id: int,
    db: AsyncSession,
) -> dict:
    """校验 Excel 点位数据，返回校验报告"""
    headers, rows = parse_excel(file_bytes)

    if not rows:
        return {
            "total": 0,
            "passed": 0,
            "failed": 0,
            "errors": [{"row": 0, "field": "", "message": "Excel 文件无数据行"}],
        }

    # 检查必填列是否存在
    missing_cols = REQUIRED_COLUMNS - set(headers)
    if missing_cols:
        return {
            "total": 0,
            "passed": 0,
            "failed": 0,
            "errors": [{"row": 0, "field": "", "message": f"缺少必填列: {', '.join(sorted(missing_cols))}"}],
        }

    errors = []
    seen_addresses: set[str] = set()

    # 查询数据库已有地址
    result = await db.execute(select(DataSourcePoint.address).where(DataSourcePoint.datasource_id == datasource_id))
    existing_addresses = {r[0] for r in result.fetchall()}

    for row in rows:
        row_num = row.get("_row", 0)
        address = row.get("address")
        data_type = row.get("data_type")

        # 必填字段
        if not address or str(address).strip() == "":
            errors.append({"row": row_num, "field": "address", "message": "地址不能为空"})
            continue
        address = str(address).strip()

        if not data_type or str(data_type).strip() == "":
            errors.append({"row": row_num, "field": "data_type", "message": "数据类型不能为空"})
            continue
        data_type = str(data_type).strip().lower()

        # 数据类型校验
        if data_type not in VALID_DATA_TYPES:
            errors.append(
                {
                    "row": row_num,
                    "field": "data_type",
                    "message": f"无效数据类型: {data_type}，允许: {', '.join(sorted(VALID_DATA_TYPES))}",
                }
            )

        # Excel 内部地址重复
        if address in seen_addresses:
            errors.append({"row": row_num, "field": "address", "message": f"地址重复: {address}"})
        seen_addresses.add(address)

        # 与数据库已有地址冲突
        if address in existing_addresses:
            errors.append({"row": row_num, "field": "address", "message": f"地址已存在: {address}"})

        # scale 数值校验
        scale = row.get("scale")
        if scale is not None and scale != "":
            try:
                float(scale)
            except (ValueError, TypeError):
                errors.append({"row": row_num, "field": "scale", "message": f"scale 必须为数值: {scale}"})

        # offset 数值校验
        offset_val = row.get("offset")
        if offset_val is not None and offset_val != "":
            try:
                float(offset_val)
            except (ValueError, TypeError):
                errors.append({"row": row_num, "field": "offset", "message": f"offset 必须为数值: {offset_val}"})

        # enum_mapping JSON 校验
        enum_mapping = row.get("enum_mapping")
        if enum_mapping is not None and enum_mapping != "":
            try:
                parsed = json.loads(str(enum_mapping))
                if not isinstance(parsed, dict):
                    errors.append({"row": row_num, "field": "enum_mapping", "message": "enum_mapping 必须为 JSON 对象"})
            except (json.JSONDecodeError, TypeError):
                errors.append({"row": row_num, "field": "enum_mapping", "message": "enum_mapping JSON 格式无效"})

    total = len(rows)
    # 统计有错误的行数
    error_rows = set(e["row"] for e in errors)
    failed = len(error_rows)
    passed = total - failed

    return {"total": total, "passed": passed, "failed": failed, "errors": errors}


async def import_points(
    file_bytes: bytes,
    datasource_id: int,
    db: AsyncSession,
) -> dict:
    """校验并导入点位，返回导入结果"""
    # 先校验
    report = await validate_points(file_bytes, datasource_id, db)
    if report["failed"] > 0:
        return {"success": False, "report": report}

    # 解析并导入
    _, rows = parse_excel(file_bytes)
    imported = 0
    for row in rows:
        address = str(row["address"]).strip()
        data_type = str(row["data_type"]).strip().lower()
        scale = float(row.get("scale") or 1.0)
        offset_val = float(row.get("offset") or 0.0)

        enum_mapping = None
        raw_enum = row.get("enum_mapping")
        if raw_enum and str(raw_enum).strip():
            enum_mapping = json.loads(str(raw_enum))

        is_dry = row.get("is_dry_contact")
        is_dry_contact = str(is_dry).lower() in ("true", "1", "yes") if is_dry else False

        point = DataSourcePoint(
            datasource_id=datasource_id,
            address=address,
            data_type=data_type,
            scale=scale,
            offset=offset_val,
            enum_mapping=enum_mapping,
            is_dry_contact=is_dry_contact,
        )
        db.add(point)
        imported += 1

    await db.commit()
    return {"success": True, "imported": imported, "report": report}
