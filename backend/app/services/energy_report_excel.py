"""
能效报告 Excel 导出 — Story 6-5
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")
_NUM_FMT = "0.00"


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None and hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_num(ws, row: int, col: int, value, fmt: str = _NUM_FMT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    return cell


def generate_energy_report_excel(report_data: dict) -> io.BytesIO:
    wb = Workbook()

    # ---- Sheet 1: 报告概览 ----
    ws1 = wb.active
    ws1.title = "报告概览"

    pue_trend = report_data.get("pue_trend", {})
    cost = report_data.get("cost_comparison", {})
    saving = report_data.get("energy_saving", {})
    overview = report_data.get("energy_overview", {})
    current_cost = cost.get("current_month", {})

    ws1.append(["算力中心月度能效报告"])
    ws1.merge_cells("A1:B1")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws1.append(["报告周期", f"{report_data.get('year', '')}年{report_data.get('month', '')}月"])
    ws1.append(["生成时间", report_data.get("generated_at", "")])
    ws1.append([])
    ws1.append(["指标", "数值"])
    ws1.cell(row=5, column=1).font = Font(bold=True)
    ws1.cell(row=5, column=2).font = Font(bold=True)
    ws1.append(["PUE均值", pue_trend.get("month_avg_pue", 0)])
    ws1.append(["总能耗(kWh)", overview.get("total_energy", 0)])
    ws1.append(["总电费(元)", current_cost.get("total_cost", 0)])
    ws1.append(["节能金额(元)", saving.get("total_saving_cost", 0)])
    _auto_width(ws1)

    # ---- Sheet 2: PUE趋势 ----
    ws2 = wb.create_sheet("PUE趋势")
    headers2 = ["日期", "平均PUE", "最小PUE", "最大PUE"]
    ws2.append(headers2)
    _style_header(ws2, len(headers2))
    for v in pue_trend.get("daily_values", []):
        row_idx = ws2.max_row + 1
        ws2.cell(row=row_idx, column=1, value=v.get("date", ""))
        _write_num(ws2, row_idx, 2, v.get("avg_pue", 0), "0.0000")
        _write_num(ws2, row_idx, 3, v.get("min_pue", 0), "0.0000")
        _write_num(ws2, row_idx, 4, v.get("max_pue", 0), "0.0000")
    _auto_width(ws2)

    # ---- Sheet 3: 电费对比 ----
    ws3 = wb.create_sheet("电费对比")
    headers3 = ["项目", "本月", "上月", "去年同月", "环比%", "同比%"]
    ws3.append(headers3)
    _style_header(ws3, len(headers3))

    lm = cost.get("last_month", {})
    ly = cost.get("last_year_month", {})

    def _rate(cur_val, prev_val):
        if prev_val and prev_val != 0 and cur_val is not None:
            return round((cur_val - prev_val) / prev_val * 100, 2)
        return None

    cost_rows = [
        ("总电费", "total_cost"),
        ("峰时电费", "peak_cost"),
        ("平时电费", "normal_cost"),
        ("谷时电费", "valley_cost"),
        ("总电量", "total_energy"),
        ("峰时电量", "peak_energy"),
        ("平时电量", "normal_energy"),
        ("谷时电量", "valley_energy"),
    ]
    for label, key in cost_rows:
        c_val = current_cost.get(key, 0)
        lm_val = lm.get(key, 0)
        ly_val = ly.get(key, 0)
        row_idx = ws3.max_row + 1
        ws3.cell(row=row_idx, column=1, value=label)
        _write_num(ws3, row_idx, 2, c_val)
        _write_num(ws3, row_idx, 3, lm_val)
        _write_num(ws3, row_idx, 4, ly_val)
        mom = _rate(c_val, lm_val)
        yoy = _rate(c_val, ly_val)
        _write_num(ws3, row_idx, 5, mom)
        _write_num(ws3, row_idx, 6, yoy)
    _auto_width(ws3)

    # ---- Sheet 4: 节能成果 ----
    ws4 = wb.create_sheet("节能成果")
    headers4 = ["方案名称", "类别", "节能(kWh)", "节省费用(元)", "达成率(%)"]
    ws4.append(headers4)
    _style_header(ws4, len(headers4))
    category_map = {1: "电费结构优化", 2: "设备运行优化", 3: "设备改造升级", 4: "综合能效提升"}
    for item in saving.get("details", []):
        row_idx = ws4.max_row + 1
        ws4.cell(row=row_idx, column=1, value=item.get("title", ""))
        ws4.cell(row=row_idx, column=2, value=category_map.get(item.get("category"), str(item.get("category", ""))))
        _write_num(ws4, row_idx, 3, item.get("saving_kwh", 0))
        _write_num(ws4, row_idx, 4, item.get("saving_cost", 0))
        _write_num(ws4, row_idx, 5, item.get("achievement_rate"))
    _auto_width(ws4)

    # ---- Sheet 5: 每日能耗 ----
    ws5 = wb.create_sheet("每日能耗")
    headers5 = ["日期", "总能耗(kWh)"]
    ws5.append(headers5)
    _style_header(ws5, len(headers5))
    for d in overview.get("daily_energy", []):
        row_idx = ws5.max_row + 1
        ws5.cell(row=row_idx, column=1, value=d.get("date", ""))
        _write_num(ws5, row_idx, 2, d.get("total_energy", 0))
    _auto_width(ws5)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
