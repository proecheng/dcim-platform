"""
事件时间线报告服务 — Story 9-5
聚合联动执行 + 恢复数据，生成完整时间线报告；支持 Excel 导出。
"""

import io
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.linkage import (
    LinkageExecution,
    LinkageLog,
    LinkagePolicy,
    LinkageRecovery,
    LinkageRecoveryLog,
)
from ..schemas.linkage import TimelineEvent, TimelineReportResponse

logger = logging.getLogger(__name__)

# ==================== Excel 样式 ====================

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.freeze_panes = "A2"


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None and hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ==================== 时间线聚合 ====================

_ACTION_TYPE_LABELS = {
    "ALARM_NOTIFY": "告警通知",
    "WEBHOOK": "Webhook 回调",
    "MQTT_COMMAND": "MQTT 指令",
    "VIDEO_RECORD": "视频录制",
    "VIDEO_POPUP": "视频弹窗",
}

_STATUS_LABELS = {
    "success": "成功",
    "failed": "失败",
    "timeout": "超时",
    "skipped": "跳过",
    "pending": "待执行",
    "executing": "执行中",
    "completed": "已完成",
    "partial_failure": "部分失败",
    "partial_recovery": "部分恢复",
}

_PHASE_LABELS = {
    "trigger": "触发",
    "action": "联动动作",
    "recovery": "恢复",
}


async def generate_timeline(
    db: AsyncSession,
    execution_id: int,
) -> Optional[TimelineReportResponse]:
    """聚合联动执行 + 恢复数据，生成完整时间线报告"""

    # 1. 查询执行记录
    result = await db.execute(select(LinkageExecution).where(LinkageExecution.id == execution_id))
    execution = result.scalar_one_or_none()
    if execution is None:
        return None

    # 2. 查询策略信息
    policy_result = await db.execute(select(LinkagePolicy).where(LinkagePolicy.id == execution.policy_id))
    policy = policy_result.scalar_one_or_none()
    policy_name = policy.name if policy else f"策略#{execution.policy_id}"
    level = policy.priority if policy else "normal"

    # 3. 查询执行日志
    logs_result = await db.execute(
        select(LinkageLog).where(LinkageLog.execution_id == execution_id).order_by(LinkageLog.id)
    )
    logs = logs_result.scalars().all()

    # 4. 查询恢复记录
    recovery_result = await db.execute(
        select(LinkageRecovery).where(LinkageRecovery.execution_id == execution_id).order_by(LinkageRecovery.id.desc())
    )
    recovery = recovery_result.scalars().first()

    recovery_logs = []
    if recovery is not None:
        rlogs_result = await db.execute(
            select(LinkageRecoveryLog)
            .where(LinkageRecoveryLog.recovery_id == recovery.id)
            .order_by(LinkageRecoveryLog.step_order)
        )
        recovery_logs = rlogs_result.scalars().all()

    # 5. 构建时间线事件列表
    events: list[TimelineEvent] = []

    # 5a. 触发事件
    trigger_detail = f"触发来源: {execution.trigger_source or '未知'}"
    if execution.trigger_event and isinstance(execution.trigger_event, dict):
        alarm_level = execution.trigger_event.get("alarm_level", "")
        if alarm_level:
            trigger_detail += f", 告警级别: {alarm_level}"
    events.append(
        TimelineEvent(
            timestamp=execution.started_at,
            phase="trigger",
            event_type="联动触发",
            detail=trigger_detail,
            status="success",
            duration_ms=None,
        )
    )

    # 5b. 联动动作
    for log in logs:
        action_label = _ACTION_TYPE_LABELS.get(log.action_type, log.action_type or "未知")
        detail = action_label
        if log.action_config and isinstance(log.action_config, dict):
            target = log.action_config.get("target", log.action_config.get("device_type", ""))
            command = log.action_config.get("command", "")
            if target:
                detail += f" → {target}"
            if command:
                detail += f" ({command})"
        if log.error_message:
            detail += f" [错误: {log.error_message}]"

        events.append(
            TimelineEvent(
                timestamp=log.started_at or execution.started_at,
                phase="action",
                event_type=action_label,
                detail=detail,
                status=log.status or "pending",
                duration_ms=log.duration_ms,
            )
        )

    # 5c. 恢复事件
    if recovery is not None:
        events.append(
            TimelineEvent(
                timestamp=recovery.started_at,
                phase="recovery",
                event_type="恢复开始",
                detail=f"模式: {'一键恢复' if recovery.mode == 'auto' else '手动恢复'}, 操作人: {recovery.operator}",
                status="success",
                duration_ms=None,
            )
        )

        for rlog in recovery_logs:
            action_label = _ACTION_TYPE_LABELS.get(rlog.action_type, rlog.action_type or "未知")
            detail = f"恢复: {action_label}"
            if rlog.recovery_command:
                detail += f" ({rlog.recovery_command})"
            if rlog.target_type:
                detail += f" → {rlog.target_type}"
            if rlog.error_message:
                detail += f" [错误: {rlog.error_message}]"

            events.append(
                TimelineEvent(
                    timestamp=rlog.started_at or recovery.started_at,
                    phase="recovery",
                    event_type=f"恢复步骤#{rlog.step_order}",
                    detail=detail,
                    status=rlog.status or "pending",
                    duration_ms=rlog.duration_ms,
                )
            )

        if recovery.completed_at:
            events.append(
                TimelineEvent(
                    timestamp=recovery.completed_at,
                    phase="recovery",
                    event_type="恢复完成",
                    detail=f"最终状态: {_STATUS_LABELS.get(recovery.status, recovery.status)}",
                    status=recovery.status or "completed",
                    duration_ms=recovery.total_duration_ms,
                )
            )

    # 6. 按时间排序（None 排最后）
    events.sort(key=lambda e: e.timestamp or datetime.max)

    # 7. 构建响应
    return TimelineReportResponse(
        execution_id=execution.id,
        event_id=execution.event_id,
        policy_name=policy_name,
        trigger_source=execution.trigger_source,
        trigger_time=execution.started_at,
        level=level,
        total_duration_ms=execution.total_duration_ms,
        recovery_time_ms=recovery.total_duration_ms if recovery else None,
        operator=recovery.operator if recovery else None,
        status=execution.status,
        events=events,
    )


# ==================== Excel 导出 ====================


def _fmt_dt(dt: Optional[datetime]) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S.") + f"{dt.microsecond // 1000:03d}"


def generate_timeline_excel(report: TimelineReportResponse) -> io.BytesIO:
    """将时间线报告导出为 Excel"""
    wb = Workbook()

    # ---- Sheet 1: 事件概要 ----
    ws1 = wb.active
    ws1.title = "事件概要"

    ws1.append(["事件时间线报告"])
    ws1.merge_cells("A1:B1")
    ws1.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws1.append(["事件ID", report.event_id])
    ws1.append(["策略名称", report.policy_name])
    ws1.append(["触发来源", report.trigger_source or ""])
    ws1.append(["级别", report.level])
    ws1.append(["触发时间", _fmt_dt(report.trigger_time)])
    ws1.append(["总耗时(ms)", report.total_duration_ms])
    ws1.append(["恢复耗时(ms)", report.recovery_time_ms])
    ws1.append(["操作人", report.operator or ""])
    ws1.append(["状态", _STATUS_LABELS.get(report.status, report.status)])

    _auto_width(ws1)

    # ---- Sheet 2: 时间线详情 ----
    ws2 = wb.create_sheet("时间线详情")
    headers = ["序号", "时间", "阶段", "事件类型", "详情", "状态", "耗时(ms)"]
    ws2.append(headers)
    _style_header(ws2, len(headers))

    for idx, evt in enumerate(report.events, 1):
        ws2.append(
            [
                idx,
                _fmt_dt(evt.timestamp),
                _PHASE_LABELS.get(evt.phase, evt.phase),
                evt.event_type,
                evt.detail,
                _STATUS_LABELS.get(evt.status, evt.status),
                evt.duration_ms,
            ]
        )

    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
