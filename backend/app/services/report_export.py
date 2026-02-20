"""对接报告导出服务 — Story 3.5"""

import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_integration_report(
    datasources: list[dict],
    points: list[dict],
) -> bytes:
    """生成对接报告 Excel，返回 bytes"""
    wb = Workbook()

    # --- Sheet 1: 数据源清单 ---
    ws1 = wb.active
    ws1.title = "数据源清单"

    headers1 = ["名称", "协议类型", "连接参数", "连接状态", "最后通信时间", "创建时间", "启用状态"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ds in enumerate(datasources, 2):
        ws1.cell(row=row_idx, column=1, value=ds.get("name", ""))
        ws1.cell(row=row_idx, column=2, value=ds.get("protocol_type", ""))
        config = ds.get("connection_config", {})
        config_summary = json.dumps(config, ensure_ascii=False) if config else ""
        ws1.cell(row=row_idx, column=3, value=config_summary)
        ws1.cell(row=row_idx, column=4, value=ds.get("status", ""))
        ws1.cell(row=row_idx, column=5, value=str(ds.get("last_communication", "") or ""))
        ws1.cell(row=row_idx, column=6, value=str(ds.get("created_at", "") or ""))
        ws1.cell(row=row_idx, column=7, value="是" if ds.get("is_enabled") else "否")

    for col in ws1.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # --- Sheet 2: 点位映射表 ---
    ws2 = wb.create_sheet("点位映射表")

    headers2 = ["数据源名称", "地址", "数据类型", "缩放系数", "偏移量", "是否干接点"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, pt in enumerate(points, 2):
        ws2.cell(row=row_idx, column=1, value=pt.get("datasource_name", ""))
        ws2.cell(row=row_idx, column=2, value=pt.get("address", ""))
        ws2.cell(row=row_idx, column=3, value=pt.get("data_type", ""))
        ws2.cell(row=row_idx, column=4, value=pt.get("scale", 1.0))
        ws2.cell(row=row_idx, column=5, value=pt.get("offset", 0.0))
        ws2.cell(row=row_idx, column=6, value="是" if pt.get("is_dry_contact") else "否")

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
