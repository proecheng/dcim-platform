"""
Shift Report Service
负荷转移报表服务 - 生成月度/年度报表并支持导出
"""

import io
from typing import Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models.load_shift import ShiftExecution, ShiftPlan


class ShiftReportService:
    """负荷转移报表服务"""

    @staticmethod
    async def generate_monthly_report(db: AsyncSession, year: int, month: int) -> Dict[str, Any]:
        """
        生成月度报表

        Args:
            db: 数据库会话
            year: 年份
            month: 月份

        Returns:
            月度报表数据
        """
        # 查询该月的所有执行记录
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        result = await db.execute(
            select(ShiftExecution).where(
                and_(ShiftExecution.start_time >= start_date, ShiftExecution.start_time < end_date)
            )
        )
        executions = result.scalars().all()

        # 统计数据
        total_executions = len(executions)
        success_count = sum(1 for e in executions if e.status == "completed")
        failed_count = sum(1 for e in executions if e.status == "failed")

        total_cost_saving = sum(float(e.actual_cost_saving) if e.actual_cost_saving else 0 for e in executions)
        total_energy_saving = sum(float(e.actual_energy_saving) if e.actual_energy_saving else 0 for e in executions)
        total_shift_power = sum(float(e.actual_shift_power) if e.actual_shift_power else 0 for e in executions)

        evaluated_count = success_count + failed_count
        success_rate = (success_count / evaluated_count * 100) if evaluated_count > 0 else None

        # 按日统计
        daily_stats = {}
        for execution in executions:
            date_key = execution.start_time.strftime("%Y-%m-%d")
            if date_key not in daily_stats:
                daily_stats[date_key] = {
                    "date": date_key,
                    "execution_count": 0,
                    "success_count": 0,
                    "failed_count": 0,
                    "total_shift_power": 0.0,
                    "cost_saving": 0.0,
                    "energy_saving": 0.0,
                }

            daily_stats[date_key]["execution_count"] += 1
            if execution.status == "completed":
                daily_stats[date_key]["success_count"] += 1
            elif execution.status == "failed":
                daily_stats[date_key]["failed_count"] += 1

            daily_stats[date_key]["total_shift_power"] += float(execution.actual_shift_power or 0)
            daily_stats[date_key]["cost_saving"] += float(execution.actual_cost_saving or 0)
            daily_stats[date_key]["energy_saving"] += float(execution.actual_energy_saving or 0)

        # 计算每日成功率
        details = []
        for date_key in sorted(daily_stats.keys()):
            stat = daily_stats[date_key]
            evaluated = stat["success_count"] + stat["failed_count"]
            stat["success_rate"] = stat["success_count"] / evaluated * 100 if evaluated > 0 else None
            details.append(stat)

        # 趋势数据
        trend_data = [
            {"date": stat["date"], "cost_saving": stat["cost_saving"], "energy_saving": stat["energy_saving"]}
            for stat in details
        ]

        # 执行统计
        execution_stats = {"success": success_count, "failed": failed_count}

        period_stats = await ShiftReportService._get_period_stats(db, executions)
        warning = ShiftReportService._build_warning(total_executions, evaluated_count)

        return {
            "report_type": "monthly",
            "year": year,
            "month": month,
            "total_cost_saving": total_cost_saving,
            "total_energy_saving": total_energy_saving,
            "total_shift_power": total_shift_power,
            "execution_count": total_executions,
            "success_rate": success_rate,
            "data_sufficient": total_executions > 0,
            "data_source": "shift_executions",
            "warning": warning,
            "details": details,
            "trend_data": trend_data,
            "execution_stats": execution_stats,
            "period_stats": period_stats,
        }

    @staticmethod
    async def generate_yearly_report(db: AsyncSession, year: int) -> Dict[str, Any]:
        """
        生成年度报表

        Args:
            db: 数据库会话
            year: 年份

        Returns:
            年度报表数据
        """
        # 查询该年的所有执行记录
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)

        result = await db.execute(
            select(ShiftExecution).where(
                and_(ShiftExecution.start_time >= start_date, ShiftExecution.start_time < end_date)
            )
        )
        executions = result.scalars().all()

        # 统计数据
        total_executions = len(executions)
        success_count = sum(1 for e in executions if e.status == "completed")
        failed_count = sum(1 for e in executions if e.status == "failed")

        total_cost_saving = sum(float(e.actual_cost_saving) if e.actual_cost_saving else 0 for e in executions)
        total_energy_saving = sum(float(e.actual_energy_saving) if e.actual_energy_saving else 0 for e in executions)
        total_shift_power = sum(float(e.actual_shift_power) if e.actual_shift_power else 0 for e in executions)

        evaluated_count = success_count + failed_count
        success_rate = (success_count / evaluated_count * 100) if evaluated_count > 0 else None

        # 按月统计
        monthly_stats = {}
        for execution in executions:
            month_key = execution.start_time.strftime("%Y-%m")
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {
                    "date": month_key,
                    "execution_count": 0,
                    "success_count": 0,
                    "failed_count": 0,
                    "total_shift_power": 0.0,
                    "cost_saving": 0.0,
                    "energy_saving": 0.0,
                }

            monthly_stats[month_key]["execution_count"] += 1
            if execution.status == "completed":
                monthly_stats[month_key]["success_count"] += 1
            elif execution.status == "failed":
                monthly_stats[month_key]["failed_count"] += 1

            monthly_stats[month_key]["total_shift_power"] += float(execution.actual_shift_power or 0)
            monthly_stats[month_key]["cost_saving"] += float(execution.actual_cost_saving or 0)
            monthly_stats[month_key]["energy_saving"] += float(execution.actual_energy_saving or 0)

        # 计算每月成功率
        details = []
        for month_key in sorted(monthly_stats.keys()):
            stat = monthly_stats[month_key]
            evaluated = stat["success_count"] + stat["failed_count"]
            stat["success_rate"] = stat["success_count"] / evaluated * 100 if evaluated > 0 else None
            details.append(stat)

        # 趋势数据
        trend_data = [
            {"date": stat["date"], "cost_saving": stat["cost_saving"], "energy_saving": stat["energy_saving"]}
            for stat in details
        ]

        # 执行统计
        execution_stats = {"success": success_count, "failed": failed_count}

        period_stats = await ShiftReportService._get_period_stats(db, executions)
        warning = ShiftReportService._build_warning(total_executions, evaluated_count)

        return {
            "report_type": "yearly",
            "year": year,
            "total_cost_saving": total_cost_saving,
            "total_energy_saving": total_energy_saving,
            "total_shift_power": total_shift_power,
            "execution_count": total_executions,
            "success_rate": success_rate,
            "data_sufficient": total_executions > 0,
            "data_source": "shift_executions",
            "warning": warning,
            "details": details,
            "trend_data": trend_data,
            "execution_stats": execution_stats,
            "period_stats": period_stats,
        }

    @staticmethod
    def _build_warning(total_executions: int, evaluated_count: int) -> str | None:
        if total_executions == 0:
            return "所选周期暂无负荷转移执行记录，无法计算实际节能、成本节省或成功率"
        if evaluated_count == 0:
            return "所选周期的执行记录尚未形成完成或失败结果，暂无法计算成功率"
        if evaluated_count < total_executions:
            return f"所选周期有 {total_executions - evaluated_count} 条执行尚未形成完成或失败结果，成功率仅基于已闭环记录计算"
        return None

    @staticmethod
    async def _get_period_stats(db: AsyncSession, executions) -> Dict[str, int]:
        stats = {"peak_to_valley": 0, "valley_to_peak": 0, "peak_to_flat": 0, "other": 0}
        plan_ids = {execution.plan_id for execution in executions if execution.plan_id is not None}
        if not plan_ids:
            return stats

        result = await db.execute(
            select(ShiftPlan.id, ShiftPlan.shift_from_period, ShiftPlan.shift_to_period).where(
                ShiftPlan.id.in_(plan_ids)
            )
        )
        periods = {row.id: (row.shift_from_period, row.shift_to_period) for row in result}
        for execution in executions:
            period = periods.get(execution.plan_id)
            if not period:
                stats["other"] += 1
                continue
            source, target = period
            if source in {"peak", "sharp"} and target == "valley":
                stats["peak_to_valley"] += 1
            elif source == "valley" and target in {"peak", "sharp"}:
                stats["valley_to_peak"] += 1
            elif source in {"peak", "sharp"} and target == "flat":
                stats["peak_to_flat"] += 1
            else:
                stats["other"] += 1
        return stats

    @staticmethod
    def export_report_excel(report_data: Dict[str, Any]) -> io.BytesIO:
        """
        导出 Excel 格式报表

        Args:
            report_data: 报表数据

        Returns:
            Excel 文件流
        """
        workbook = Workbook()
        summary = workbook.active
        summary.title = "报表概览"
        summary.append(["指标", "数值"])
        for cell in summary[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        summary_rows = [
            ("报告类型", "月度" if report_data.get("report_type") == "monthly" else "年度"),
            ("执行次数", report_data.get("execution_count", 0)),
            ("成功率(%)", report_data.get("success_rate")),
            ("实际转移功率(kW)", report_data.get("total_shift_power", 0)),
            ("实际成本节省(元)", report_data.get("total_cost_saving", 0)),
            ("实际节能量(kWh)", report_data.get("total_energy_saving", 0)),
            ("数据来源", report_data.get("data_source", "shift_executions")),
            ("数据提示", report_data.get("warning") or "数据完整"),
        ]
        for row in summary_rows:
            summary.append(row)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 60

        details_sheet = workbook.create_sheet("执行明细")
        details_sheet.append(
            ["日期", "执行次数", "成功次数", "失败次数", "转移功率(kW)", "成本节省(元)", "节能量(kWh)", "成功率(%)"]
        )
        for cell in details_sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for detail in report_data.get("details", []):
            details_sheet.append(
                [
                    detail.get("date"),
                    detail.get("execution_count", 0),
                    detail.get("success_count", 0),
                    detail.get("failed_count", 0),
                    detail.get("total_shift_power", 0),
                    detail.get("cost_saving", 0),
                    detail.get("energy_saving", 0),
                    detail.get("success_rate"),
                ]
            )
        details_sheet.freeze_panes = "A2"
        for column in details_sheet.columns:
            details_sheet.column_dimensions[column[0].column_letter].width = 18

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_report_pdf(report_data: Dict[str, Any]) -> io.BytesIO:
        """
        导出 PDF 格式报表

        Args:
            report_data: 报表数据

        Returns:
            PDF 文件流
        """
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        buffer = io.BytesIO()
        document = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("ShiftReportTitle", parent=styles["Title"], fontName="STSong-Light", fontSize=18)
        normal_style = ParagraphStyle(
            "ShiftReportNormal", parent=styles["Normal"], fontName="STSong-Light", fontSize=10
        )
        elements = [Paragraph("负荷转移收益报表", title_style), Spacer(1, 16)]
        summary_data = [
            ["指标", "数值"],
            ["执行次数", str(report_data.get("execution_count", 0))],
            ["成功率", "--" if report_data.get("success_rate") is None else f"{report_data['success_rate']:.1f}%"],
            ["实际转移功率", f"{report_data.get('total_shift_power', 0):.2f} kW"],
            ["实际成本节省", f"{report_data.get('total_cost_saving', 0):.2f} 元"],
            ["实际节能量", f"{report_data.get('total_energy_saving', 0):.2f} kWh"],
        ]
        table = Table(summary_data, colWidths=[160, 260])
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ]
            )
        )
        elements.append(table)
        if report_data.get("warning"):
            elements.extend([Spacer(1, 12), Paragraph(report_data["warning"], normal_style)])
        document.build(elements)
        buffer.seek(0)
        return buffer
